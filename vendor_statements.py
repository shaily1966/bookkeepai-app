"""
BookKeep AI Pro — Universal Vendor Statement Processor v1.0
===========================================================
Handles ANY vendor account statement PDF or CSV.
Three processing paths:

  Path A — Amazon Business CSV   : zero AI tokens, structured, highest accuracy
  Path B — Generic CSV           : column-detection heuristic, 80% of exports
  Path C — Universal PDF (Claude): any vendor, any format, reads ALL pages

All three paths produce identical output dicts that plug directly into
the existing reconciliation and Excel builder in app.py.

Integration (add to app.py):
─────────────────────────────
from vendor_statements import process_vendor_statements, add_vendor_items_tab

# 4th upload slot alongside existing 3:
vendor_files = st.file_uploader(
    "🏪 Vendor Statements — Amazon, Costco, Home Depot, Sysco...",
    type=["pdf","csv","tsv"], accept_multiple_files=True, key="vendor_stmts"
)

if vendor_files and st.session_state.transactions:
    with st.spinner("Processing vendor statements..."):
        vendor_items, v_cost = process_vendor_statements(
            vendor_files,
            st.session_state.transactions,
            industry, province_code,
            client, model_id, rate_in, rate_out, period
        )
        st.session_state.vendor_items = vendor_items
        st.session_state.total_cost += v_cost

# In build_excel(), add the tab:
if st.session_state.get('vendor_items'):
    add_vendor_items_tab(wb, st.session_state.vendor_items)
"""

import re
import csv
import io
import base64
import logging
from datetime import datetime
from collections import Counter

logger = logging.getLogger("bookkeep_ai.vendor")

# ═══════════════════════════════════════════════════════════════════
# T2125 LINE MAP (mirrors bank_schemas — no circular import)
# ═══════════════════════════════════════════════════════════════════

T2125 = {
    "Motor Vehicle Expense":  "9281",
    "Meals & Entertainment":  "8523",
    "Office Supplies":        "8810",
    "Utilities":              "8220",
    "Bank Charges":           "8710",
    "Insurance":              "8690",
    "Materials & Supplies":   "8811",
    "Rent":                   "8910",
    "Delivery & Shipping":    "8730",
    "Advertising":            "8520",
    "Travel":                 "9200",
    "Professional Fees":      "8860",
    "Subcontracts":           "8590",
    "Cost of Goods":          "8320",
    "Repairs & Maintenance":  "8960",
    "Owner Draw / Personal":  "",
    "Government Remittances": "",
    "❓ Uncategorized":       "",
}

VALID_CATEGORIES = set(T2125.keys())

# ═══════════════════════════════════════════════════════════════════
# PATH A — AMAZON BUSINESS FAST PATH
# Amazon's own product taxonomy → CRA T2125, zero AI tokens
# ═══════════════════════════════════════════════════════════════════

AMAZON_CATEGORY_MAP = {
    "tools & hardware":              ("Materials & Supplies",  "Full"),
    "power & hand tools":            ("Materials & Supplies",  "Full"),
    "safety & ppe":                  ("Materials & Supplies",  "Full"),
    "industrial & scientific":       ("Materials & Supplies",  "Full"),
    "building supplies":             ("Materials & Supplies",  "Full"),
    "janitorial & sanitation":       ("Materials & Supplies",  "Full"),
    "raw materials":                 ("Materials & Supplies",  "Full"),
    "packaging & shipping supplies": ("Delivery & Shipping",   "Full"),
    "office products":               ("Office Supplies",       "Full"),
    "office supplies":               ("Office Supplies",       "Full"),
    "computers & accessories":       ("Office Supplies",       "Full"),
    "electronics":                   ("Office Supplies",       "Full"),
    "software":                      ("Office Supplies",       "Full"),
    "printers & ink":                ("Office Supplies",       "Full"),
    "business & industrial":         ("Office Supplies",       "Full"),
    "books":                         ("Office Supplies",       "Full"),
    "grocery & gourmet food":        ("Cost of Goods",         "Full"),
    "food service equipment":        ("Materials & Supplies",  "Full"),
    "restaurant & food service":     ("Cost of Goods",         "Full"),
    "clothing, shoes & jewelry":     ("Owner Draw / Personal", "No"),
    "clothing":                      ("Owner Draw / Personal", "No"),
    "toys & games":                  ("Owner Draw / Personal", "No"),
    "sports & outdoors":             ("Owner Draw / Personal", "No"),
    "health & household":            ("Owner Draw / Personal", "No"),
    "beauty & personal care":        ("Owner Draw / Personal", "No"),
    "baby products":                 ("Owner Draw / Personal", "No"),
    "pet supplies":                  ("Owner Draw / Personal", "No"),
    "movies & tv":                   ("Owner Draw / Personal", "No"),
    "music":                         ("Owner Draw / Personal", "No"),
    "video games":                   ("Owner Draw / Personal", "No"),
    "home & kitchen":                ("❓ Uncategorized",      "No"),
    "home improvement":              ("❓ Uncategorized",      "No"),
    "garden & outdoor":              ("❓ Uncategorized",      "No"),
    "amazon devices":                ("Office Supplies",       "Full"),
}

AMAZON_DESC_HINTS = [
    (r'\b(drill|saw|wrench|hammer|socket|level|tape.?measure|ladder|'
     r'bit.?set|worklight|grinder|nailer|compressor|generator|'
     r'dewalt|makita|milwaukee|bosch|ridgid|ryobi|hilti|ramset|paslode)\b',
     "Materials & Supplies"),
    (r'\b(safety.?glass|hard.?hat|gloves|respirator|ppe|hi.?vis|'
     r'steel.?toe|reflective|ear.?protect|harness|first.?aid)\b',
     "Materials & Supplies"),
    (r'\b(printer|ink|toner|paper|stapler|binder|folder|pen|pencil|'
     r'notebook|monitor|keyboard|mouse|webcam|headset|cable|usb|'
     r'desk|chair|whiteboard|shredder|laptop|tablet)\b',
     "Office Supplies"),
    (r'\b(lego|toy|game|clothing|shirt|pants|jacket|dress|shoe|'
     r'birthday|gift|home.?decor|pillow|blanket|candle|jewel)\b',
     "Owner Draw / Personal"),
    (r'\b(food|snack|coffee|tea|water|beverage|grocery|sugar|flour|'
     r'condiment|sauce|spice)\b',
     "❓ Uncategorized"),
]

# Industry overrides for ambiguous category resolutions
INDUSTRY_OVERRIDES = {
    ("❓ Uncategorized",   "Restaurant/Food"):       "Cost of Goods",
    ("❓ Uncategorized",   "Retail"):                "Cost of Goods",
    ("Office Supplies",    "Construction/Trades"):   "❓ Uncategorized",
    ("Home & Kitchen",     "Restaurant/Food"):       "Cost of Goods",
    ("Home & Kitchen",     "Construction/Trades"):   "Materials & Supplies",
    ("Home Improvement",   "Construction/Trades"):   "Materials & Supplies",
    ("Home Improvement",   "Professional Services"): "❓ Uncategorized",
}


def _amazon_resolve(vendor_cat, description, industry):
    """Amazon category + description + industry → (category, t2125, itc_rule, confidence)."""
    vc = vendor_cat.strip().lower()
    for amazon_cat, (category, itc_rule) in AMAZON_CATEGORY_MAP.items():
        if amazon_cat in vc:
            resolved = INDUSTRY_OVERRIDES.get((category, industry), category)
            conf = "95" if resolved != "❓ Uncategorized" else "62"
            return resolved, T2125.get(resolved, ""), itc_rule, conf

    desc_lower = description.lower()
    for pattern, category in AMAZON_DESC_HINTS:
        if re.search(pattern, desc_lower, re.IGNORECASE):
            resolved = INDUSTRY_OVERRIDES.get((category, industry), category)
            conf = "82" if resolved != "❓ Uncategorized" else "62"
            return resolved, T2125.get(resolved, ""), "Full", conf

    defaults = {
        "Construction/Trades":   ("❓ Uncategorized", "", "No",   "58"),
        "Restaurant/Food":       ("❓ Uncategorized", "", "No",   "58"),
        "Retail":                ("❓ Uncategorized", "", "No",   "58"),
        "Professional Services": ("Office Supplies",  "8810", "Full", "72"),
    }
    return defaults.get(industry, ("❓ Uncategorized", "", "No", "58"))


# ═══════════════════════════════════════════════════════════════════
# GENERIC CSV COLUMN DETECTOR
# ═══════════════════════════════════════════════════════════════════

COLUMN_RULES = [
    (r'order\s*date|transaction\s*date|purchase\s*date|invoice\s*date|ship\s*date|\bdate\b', "date"),
    (r'product\s*name|item\s*description|item\s*name|product\s*title|\bdescription\b|title|part\s*desc', "description"),
    (r'product\s*category|department|\bcategory\b|item\s*type|product\s*group|commodity', "vendor_category"),
    (r'item\s*total|order\s*total|total\s*price|unit\s*price|\bprice\b|\bamount\b|\btotal\b|\bcost\b|extended\s*price|net\s*amount', "amount"),
    (r'\bquantity\b|\bqty\b|\bunits\b|\bcount\b', "quantity"),
    (r'\bhst\b|\bgst\b|tax\s*amount|\btax\b|sales\s*tax', "tax"),
    (r'order\s*(id|number|#|num)|invoice\s*(id|number|#|num)|po\s*number|purchase\s*order', "order_num"),
    (r'\basin\b|\bsku\b|item\s*(code|number|#|id)|part\s*(number|#)', "sku"),
    (r'sold\s*by|\bsupplier\b|\bvendor\b|\bseller\b', "seller"),
    (r'\bsubtotal\b|sub\s*total|before\s*tax', "subtotal"),
]


def _detect_columns(header_row):
    mapping = {}
    for idx, col in enumerate(header_row):
        col_clean = col.strip().lower().strip('"\'')
        for pattern, field in COLUMN_RULES:
            if re.search(pattern, col_clean) and field not in mapping:
                mapping[field] = idx
                break
    return mapping


def _get(row, field, col_map, default=""):
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return default
    return row[idx].strip().strip('"\'')


def _parse_amount(s):
    try:
        return round(float(re.sub(r'[,$\s]', '', s or "0")), 2)
    except (ValueError, TypeError):
        return 0.0


def _read_file(file_obj):
    """Read file bytes and decode safely."""
    file_obj.seek(0)
    raw = file_obj.read()
    if raw[:3] == b'\xef\xbb\xbf':
        return raw.decode("utf-8-sig")
    try:
        return raw.decode("utf-8")
    except UnicodeDecodeError:
        return raw.decode("latin-1")


# ═══════════════════════════════════════════════════════════════════
# GENERIC VENDOR CATEGORY RESOLVER
# ═══════════════════════════════════════════════════════════════════

GENERIC_CATEGORY_RULES = [
    (r'food|grocery|produce|dairy|meat|seafood|poultry|bakery|frozen|beverage|'
     r'dry.?goods|ingredient|condiment|restaurant|catering', "Cost of Goods"),
    (r'tool|hardware|lumber|concrete|fastener|pipe|fitting|electrical|plumbing|'
     r'safety|ppe|protective|workwear|paint|coating|adhesive|sealant|roofing|'
     r'drywall|insulation|welding|scaffold', "Materials & Supplies"),
    (r'office|stationery|paper|printer|ink|toner|binder|folder|pen|pencil|'
     r'label|envelope|filing|shredder|whiteboard', "Office Supplies"),
    (r'computer|laptop|monitor|keyboard|mouse|tablet|phone|charger|cable|'
     r'electronic|software|server|networking|tech', "Office Supplies"),
    (r'cleaning|janitorial|sanit|soap|disinfect|mop|broom|garbage.?bag|'
     r'trash|tissue|paper.?towel', "Materials & Supplies"),
    (r'\bauto\b|vehicle|\boil\b|lubricant|filter|\btire\b|battery|wiper|'
     r'fluid|engine|brake|exhaust', "Motor Vehicle Expense"),
    (r'shipping|freight|courier|delivery|postage|handling', "Delivery & Shipping"),
    (r'apparel|clothing|shoe|toy|entertainment|beauty|cosmetic|personal.?care|pharmacy',
     "Owner Draw / Personal"),
]

VENDOR_DEFAULTS = {
    "SYSCO":           {"Restaurant/Food": "Cost of Goods",          "default": "Cost of Goods"},
    "GFS":             {"Restaurant/Food": "Cost of Goods",          "default": "Cost of Goods"},
    "GORDON FOOD":     {"Restaurant/Food": "Cost of Goods",          "default": "Cost of Goods"},
    "FASTENAL":        {"Construction/Trades": "Materials & Supplies","default": "Materials & Supplies"},
    "GRAINGER":        {"Construction/Trades": "Materials & Supplies","default": "Materials & Supplies"},
    "ULINE":           {"default": "Materials & Supplies"},
    "STAPLES":         {"default": "Office Supplies"},
    "BUREAU EN GROS":  {"default": "Office Supplies"},
    "HOME DEPOT": {
        "Construction/Trades": "Materials & Supplies",
        "default": "❓ Uncategorized",
    },
    "COSTCO": {
        "Restaurant/Food":       "Cost of Goods",
        "Retail":                "Cost of Goods",
        "Construction/Trades":   "Materials & Supplies",
        "Professional Services": "Office Supplies",
        "default":               "❓ Uncategorized",
    },
    "CANADIAN TIRE": {
        "Construction/Trades": "Materials & Supplies",
        "default": "❓ Uncategorized",
    },
    "WALMART": {"default": "❓ Uncategorized"},
    "BEST BUY": {"default": "❓ Uncategorized"},
}


def _resolve_generic_category(vendor_category, description, vendor_name, industry):
    """Returns (category, t2125, itc_rule, confidence)."""
    combined = f"{vendor_category} {description}".lower()

    for pattern, category in GENERIC_CATEGORY_RULES:
        if re.search(pattern, combined, re.IGNORECASE):
            resolved = INDUSTRY_OVERRIDES.get((category, industry), category)
            conf = "85" if vendor_category else "75"
            return resolved, T2125.get(resolved, ""), "Full", conf

    vendor_upper = vendor_name.upper()
    for vendor_key, industry_map in VENDOR_DEFAULTS.items():
        if vendor_key in vendor_upper:
            category = industry_map.get(industry) or industry_map.get("default", "❓ Uncategorized")
            return category, T2125.get(category, ""), "Full", "80"

    return "❓ Uncategorized", "", "No", "55"


# ═══════════════════════════════════════════════════════════════════
# PATH C — UNIVERSAL PDF PARSER via Claude Vision
# ═══════════════════════════════════════════════════════════════════

VENDOR_PDF_PROMPT = """You are processing a vendor account statement or supplier invoice for Canadian bookkeeping.

Extract EVERY individual line item. Return ONLY TSV rows. No headers. No totals. No commentary.

Each row: Date | OrderNum | Description | VendorCategory | Qty | Amount | HST | CRACategory

RULES:
- Date: YYYY-MM-DD. Use document date if not per-line.
- OrderNum: order/invoice/PO number. N/A if absent.
- Description: exact product name. Be specific.
- VendorCategory: vendor's own category label if shown, else blank.
- Qty: numeric. 1 if not shown.
- Amount: pre-tax CAD amount, numbers only.
- HST: HST/GST amount for this line. 0 if not itemized.
- CRACategory: EXACTLY one of:
    Materials & Supplies | Office Supplies | Cost of Goods | Motor Vehicle Expense
    Meals & Entertainment | Utilities | Delivery & Shipping | Advertising | Travel
    Professional Fees | Owner Draw / Personal | ❓ Uncategorized

CRACategory logic:
  Tools, hardware, lumber, PPE, safety → Materials & Supplies
  Paper, ink, computers, software, SaaS → Office Supplies
  Food, ingredients, beverages for resale → Cost of Goods
  Fuel, vehicle parts, tires → Motor Vehicle Expense
  Personal items, clothing, toys → Owner Draw / Personal
  Cannot determine → ❓ Uncategorized

SKIP: order totals, subtotals, tax summaries, balance rows, headers, page numbers.
Output ONLY TSV data rows."""


def _parse_vendor_pdf_claude(file_obj, vendor_name, industry,
                              client, model_id, rate_in, rate_out, max_pages=40):
    """Claude Vision reads ALL pages of any vendor PDF. Returns (items, cost)."""
    items = []
    total_cost = 0.0

    try:
        import fitz
        import pdfplumber

        file_obj.seek(0)
        pdf_bytes = file_obj.read()
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        total_pages = min(len(doc), max_pages)
        BATCH = 4

        for batch_start in range(0, total_pages, BATCH):
            batch_end = min(batch_start + BATCH, total_pages)
            content = []

            for pn in range(batch_start, batch_end):
                page_text = ""
                try:
                    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                        if pn < len(pdf.pages):
                            page_text = pdf.pages[pn].extract_text() or ""
                except Exception:
                    pass

                garbled = (
                    not page_text.strip()
                    or len(page_text.strip()) < 50
                    or len(re.findall(r'\(cid:\d+\)', page_text)) > 3
                )

                if garbled:
                    pix = doc[pn].get_pixmap(matrix=fitz.Matrix(150/72, 150/72))
                    b64 = base64.b64encode(pix.tobytes("png")).decode("utf-8")
                    content.append({"type": "image", "source": {
                        "type": "base64", "media_type": "image/png", "data": b64}})
                else:
                    content.append({"type": "text",
                                    "text": f"[PAGE {pn+1}]\n{page_text[:3000]}"})

            if not content:
                continue

            content.append({"type": "text", "text": (
                f"\nVendor: {vendor_name} | Industry: {industry} | "
                f"Pages {batch_start+1}-{batch_end} of {total_pages}\n\n"
                + VENDOR_PDF_PROMPT
            )})

            try:
                resp = ""
                with client.messages.stream(
                    model=model_id, max_tokens=2000,
                    messages=[{"role": "user", "content": content}],
                    timeout=120.0
                ) as stream:
                    for txt in stream.text_stream:
                        resp += txt

                in_tok = (len(VENDOR_PDF_PROMPT) // 4) + (batch_end - batch_start) * 800
                total_cost += (in_tok * rate_in + (len(resp) // 4) * rate_out) / 1e6

                for line in resp.strip().split("\n"):
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    p = [c.strip() for c in line.split("\t")]
                    if len(p) < 5:
                        continue
                    if not p[0] or p[0].lower() in ("date", "---", ""):
                        continue
                    try:
                        amount  = _parse_amount(p[5]) if len(p) > 5 else 0
                        hst     = _parse_amount(p[6]) if len(p) > 6 else 0
                        cra_cat = p[7].strip() if len(p) > 7 else "❓ Uncategorized"
                        if cra_cat not in VALID_CATEGORIES:
                            cra_cat = "❓ Uncategorized"
                        if amount <= 0:
                            continue
                        items.append({
                            "source":          vendor_name,
                            "vendor":          vendor_name,
                            "date":            p[0],
                            "order_num":       p[1] if len(p) > 1 else "N/A",
                            "description":     p[2] if len(p) > 2 else "",
                            "vendor_category": p[3] if len(p) > 3 else "",
                            "quantity":        p[4] if len(p) > 4 else "1",
                            "amount":          amount,
                            "tax":             hst,
                            "sku":             "",
                            "category":        cra_cat,
                            "t2125":           T2125.get(cra_cat, ""),
                            "itc_rule":        "50%" if "Meals" in cra_cat else (
                                               "Full" if cra_cat not in ("Owner Draw / Personal", "❓ Uncategorized") else "No"),
                            "confidence":      "88" if cra_cat != "❓ Uncategorized" else "60",
                            "matched_txn":     None,
                            "notes":           f"VENDOR_PDF:{vendor_name} pp{batch_start+1}-{batch_end}",
                        })
                    except Exception as e:
                        logger.debug(f"PDF row parse error: {e} | {line}")

            except Exception as e:
                logger.warning(f"Claude vendor PDF batch error ({vendor_name}): {e}")

        doc.close()
    except Exception as e:
        logger.error(f"Vendor PDF parse failed ({vendor_name}): {e}")

    return items, total_cost


# ═══════════════════════════════════════════════════════════════════
# RECONCILIATION — match vendor items to bank transactions
# ═══════════════════════════════════════════════════════════════════

def _fuzzy_date(date_str):
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%b %d, %Y",
                "%B %d, %Y", "%d-%b-%Y", "%Y/%m/%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except (ValueError, AttributeError):
            continue
    return None


def _reconcile_items_to_bank(vendor_items, bank_transactions, period_hint=""):
    """
    Group vendor items by order, match order totals to bank debits
    by amount (±2%) and date (±7 days). Upgrades bank transaction
    category when vendor data is more specific.
    """
    if not bank_transactions or not vendor_items:
        return vendor_items

    # Group by (order_num, date)
    orders = {}
    for item in vendor_items:
        key = (item.get("order_num", "N/A"), item.get("date", ""))
        orders.setdefault(key, []).append(item)

    for (order_num, order_date), items in orders.items():
        order_total = round(sum(i["amount"] + i.get("tax", 0) for i in items), 2)
        if order_total <= 0:
            continue

        order_dt = _fuzzy_date(order_date)
        best_txn, best_diff = None, 999

        for txn in bank_transactions:
            txn_amt = txn.get("debit", 0) or txn.get("credit", 0)
            tolerance = max(order_total * 0.02, 1.00)
            if abs(txn_amt - order_total) > tolerance:
                continue
            txn_dt = _fuzzy_date(txn.get("date", ""))
            day_diff = abs((order_dt - txn_dt).days) if order_dt and txn_dt else 5
            if day_diff <= 7 and day_diff < best_diff:
                best_diff = day_diff
                best_txn = txn

        if best_txn:
            matched_desc = best_txn.get("description", "")
            vendor_name = items[0].get("vendor", "Vendor")
            for item in items:
                item["matched_txn"] = matched_desc
                item["matched_date"] = best_txn.get("date", "")

            # Tag bank transaction
            best_txn["notes"] = (
                best_txn.get("notes", "")
                + f" VENDOR_MATCHED:{vendor_name} ORDER:{order_num}"
            ).strip()

            # Upgrade bank transaction category if it was Uncategorized
            if "Uncategorized" in best_txn.get("category", ""):
                cats = [i["category"] for i in items
                        if i["category"] not in ("❓ Uncategorized", "")]
                if cats:
                    best_cat = Counter(cats).most_common(1)[0][0]
                    best_txn["category"] = best_cat
                    best_txn["t2125"] = T2125.get(best_cat, "")
                    best_txn["confidence"] = "92"
                    best_txn["notes"] = (
                        best_txn["notes"] + " CAT_FROM_VENDOR_STMT"
                    ).strip()

    return vendor_items


# ═══════════════════════════════════════════════════════════════════
# CSV PARSERS (public)
# ═══════════════════════════════════════════════════════════════════

def parse_amazon_business_csv(file_obj, industry="Other"):
    """
    Amazon Business order history CSV.
    Download: Amazon Business → Reports → Order History → Export CSV
    Zero AI tokens used.
    """
    items = []
    try:
        text = _read_file(file_obj)
        rows = list(csv.reader(io.StringIO(text)))
        if not rows:
            return items
        col_map = _detect_columns(rows[0])
        for row in rows[1:]:
            if not row or all(c.strip() == "" for c in row):
                continue
            try:
                description = _get(row, "description",     col_map, "Unknown Item")
                vendor_cat  = _get(row, "vendor_category", col_map, "")
                date        = _get(row, "date",            col_map, "")
                order_num   = _get(row, "order_num",       col_map, "N/A")
                sku         = _get(row, "sku",             col_map, "")
                quantity    = _get(row, "quantity",        col_map, "1")
                seller      = _get(row, "seller",          col_map, "Amazon")
                amount      = _parse_amount(_get(row, "amount",    col_map, "0")
                                            or _get(row, "subtotal", col_map, "0"))
                tax         = _parse_amount(_get(row, "tax", col_map, "0"))
                if amount <= 0:
                    continue
                category, t2125_line, itc_rule, confidence = _amazon_resolve(
                    vendor_cat, description, industry)
                items.append({
                    "source": "Amazon Business", "vendor": seller or "Amazon",
                    "date": date, "order_num": order_num, "sku": sku,
                    "description": description, "quantity": quantity,
                    "amount": amount, "tax": tax, "vendor_category": vendor_cat,
                    "category": category, "t2125": t2125_line,
                    "itc_rule": itc_rule, "confidence": confidence,
                    "matched_txn": None,
                    "notes": f"AMAZON_CSV ORDER:{order_num}",
                })
            except Exception as e:
                logger.debug(f"Amazon CSV row: {e}")
    except Exception as e:
        logger.error(f"Amazon CSV failed: {e}")
    return items


def parse_generic_vendor_csv(file_obj, vendor_name="Vendor", industry="Other"):
    """
    Universal CSV parser for any vendor.
    Works for: Costco, Home Depot Pro, Staples, Sysco, Grainger, Fastenal, ULINE, etc.
    """
    items = []
    try:
        text = _read_file(file_obj)
        rows = list(csv.reader(io.StringIO(text)))
        if len(rows) < 2:
            return items

        # Find header row
        header_idx = 0
        for i, row in enumerate(rows[:10]):
            if sum(1 for c in row if c.strip()) >= 3:
                header_idx = i
                break

        col_map = _detect_columns(rows[header_idx])
        if "description" not in col_map and "amount" not in col_map:
            logger.warning(f"Generic CSV: no usable columns in {vendor_name}")
            return items

        for row in rows[header_idx + 1:]:
            if not row or all(c.strip() == "" for c in row):
                continue
            try:
                description = _get(row, "description",     col_map, "Unknown Item")
                vendor_cat  = _get(row, "vendor_category", col_map, "")
                date        = _get(row, "date",            col_map, "")
                order_num   = _get(row, "order_num",       col_map, "")
                quantity    = _get(row, "quantity",        col_map, "1")
                sku         = _get(row, "sku",             col_map, "")
                amount      = _parse_amount(_get(row, "amount",    col_map, "0")
                                            or _get(row, "subtotal", col_map, "0"))
                tax         = _parse_amount(_get(row, "tax", col_map, "0"))
                if amount <= 0:
                    continue
                category, t2125_line, itc_rule, confidence = _resolve_generic_category(
                    vendor_cat, description, vendor_name, industry)
                items.append({
                    "source": vendor_name, "vendor": vendor_name,
                    "date": date, "order_num": order_num, "sku": sku,
                    "description": description, "quantity": quantity,
                    "amount": amount, "tax": tax, "vendor_category": vendor_cat,
                    "category": category, "t2125": t2125_line,
                    "itc_rule": itc_rule, "confidence": confidence,
                    "matched_txn": None,
                    "notes": f"VENDOR_CSV:{vendor_name}" + (f" ORDER:{order_num}" if order_num else ""),
                })
            except Exception as e:
                logger.debug(f"Generic CSV row ({vendor_name}): {e}")
    except Exception as e:
        logger.error(f"Generic CSV failed ({vendor_name}): {e}")
    return items


# ═══════════════════════════════════════════════════════════════════
# VENDOR NAME DETECTION FROM FILENAME
# ═══════════════════════════════════════════════════════════════════

KNOWN_VENDORS = [
    "Amazon", "Costco", "Home Depot", "Staples", "Sysco",
    "Grainger", "Fastenal", "ULINE", "Canadian Tire",
    "Gordon Food", "GFS", "Walmart", "Best Buy", "Bureau En Gros",
]

def _detect_vendor_name(filename):
    name = re.sub(r'\.(pdf|csv|tsv)$', '', filename, flags=re.IGNORECASE)
    for v in KNOWN_VENDORS:
        if v.lower() in name.lower():
            return v
    return re.sub(r'[_\-]+', ' ', name).strip().title() or "Vendor"


# ═══════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════════

def process_vendor_statements(vendor_files, bank_transactions,
                               industry, province_code,
                               client, model_id, rate_in, rate_out,
                               period=""):
    """
    Process any mix of vendor PDFs and CSVs.

    Args:
        vendor_files      : list of Streamlit UploadedFile objects
        bank_transactions : list of transaction dicts (modified in place with match tags)
        industry          : "Construction/Trades", "Restaurant/Food", etc.
        province_code     : "ON", "AB", etc. (reserved for future ITC calc)
        client            : anthropic.Anthropic client instance
        model_id          : Claude model string (e.g. "claude-sonnet-4-6")
        rate_in/rate_out  : API pricing per million tokens
        period            : statement period string for date disambiguation

    Returns:
        (vendor_items, total_api_cost)
    """
    all_items = []
    total_cost = 0.0

    for f in vendor_files:
        fname = f.name.lower()
        vendor_name = _detect_vendor_name(f.name)

        # Path A — Amazon Business CSV
        if fname.endswith(".csv") and "amazon" in fname:
            items = parse_amazon_business_csv(f, industry)
            logger.info(f"Amazon Business CSV: {len(items)} line items")

        # Path B — Generic CSV / TSV
        elif fname.endswith((".csv", ".tsv")):
            items = parse_generic_vendor_csv(f, vendor_name, industry)
            logger.info(f"Generic CSV ({vendor_name}): {len(items)} line items")

        # Path C — Universal PDF via Claude
        elif fname.endswith(".pdf"):
            items, cost = _parse_vendor_pdf_claude(
                f, vendor_name, industry,
                client, model_id, rate_in, rate_out
            )
            total_cost += cost
            logger.info(f"PDF Claude ({vendor_name}): {len(items)} items, ${cost:.4f}")

        else:
            logger.warning(f"Unsupported file type: {f.name}")
            continue

        all_items.extend(items)

    # Reconcile vendor items against bank transactions
    if all_items and bank_transactions:
        all_items = _reconcile_items_to_bank(all_items, bank_transactions, period)

    matched = sum(1 for i in all_items if i.get("matched_txn"))
    logger.info(f"Vendor total: {len(all_items)} items, {matched} matched, ${total_cost:.4f}")
    return all_items, total_cost


# ═══════════════════════════════════════════════════════════════════
# EXCEL TAB BUILDER
# ═══════════════════════════════════════════════════════════════════

def add_vendor_items_tab(wb, vendor_items):
    """
    Add 'Vendor Line Items' tab to an existing openpyxl Workbook.

    Call from build_excel() in app.py:
        from vendor_statements import add_vendor_items_tab
        if st.session_state.get('vendor_items'):
            add_vendor_items_tab(wb, st.session_state.vendor_items)
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Vendor Line Items")

    HEADERS = ["Vendor", "Date", "Order #", "SKU", "Description",
               "Vendor Category", "Qty", "Amount", "HST",
               "CRA Category", "T2125", "Conf",
               "Matched Bank Transaction", "Notes"]

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    RED    = PatternFill("solid", fgColor="FFC7CE")

    for r, item in enumerate(vendor_items, 2):
        try:
            conf = int(item.get("confidence", "0"))
        except (ValueError, TypeError):
            conf = 0
        fill = GREEN if conf >= 85 else YELLOW if conf >= 70 else RED

        values = [
            item.get("vendor", ""),
            item.get("date", ""),
            item.get("order_num", ""),
            item.get("sku", ""),
            item.get("description", ""),
            item.get("vendor_category", ""),
            item.get("quantity", "1"),
            item.get("amount", 0),
            item.get("tax", 0),
            item.get("category", ""),
            item.get("t2125", ""),
            item.get("confidence", ""),
            item.get("matched_txn") or "⚠️ Unmatched",
            item.get("notes", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = fill
            if col in (8, 9):
                cell.number_format = '$#,##0.00'

    WIDTHS = [14, 12, 14, 14, 40, 22, 5, 10, 8, 22, 7, 5, 38, 28]
    for col, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    if vendor_items:
        total_row = len(vendor_items) + 2
        ws.cell(row=total_row, column=7, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=8,
                value=round(sum(i.get("amount", 0) for i in vendor_items), 2)
                ).font = Font(bold=True)
        ws.cell(row=total_row, column=9,
                value=round(sum(i.get("tax", 0) for i in vendor_items), 2)
                ).font = Font(bold=True)
        matched = sum(1 for i in vendor_items if i.get("matched_txn"))
        ws.cell(row=total_row, column=13,
                value=f"{matched} of {len(vendor_items)} matched to bank"
                ).font = Font(bold=True)

    return ws

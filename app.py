import streamlit as st
import anthropic
import pdfplumber
import os, time, json, re, io, base64, logging
from datetime import datetime
from exports import build_quickbooks_csv, build_xero_csv
from bank_schemas import (
    detect_and_parse, detect_bank, get_supported_banks,
    pre_categorize_merchant, clean_description, detect_province_from_text,
    detect_province_from_description, PROVINCE_TAX_RATES, INDUSTRY_PROFILES,
    normalize_merchant, get_cca_class,
    run_validation, match_refunds_to_purchases, apply_year_bound,
    check_transaction_coverage, detect_expense_anomalies,
    detect_subcontractor_payment, aggregate_t5018, apply_industry_remaps,
    detect_shareholder_loan, get_itc_rate_fraction,
    # v3.4 additions
    normalize_province, T2125_LINE_MAP, parse_date_fuzzy,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VERSION = "3.11"
logger = logging.getLogger("bookkeep_ai")
logging.basicConfig(level=logging.WARNING)



# VENDOR MEMORY — SQLite backed (v3.5)
# Replaces ephemeral JSON file and volatile session_state-only storage.
# SQLite persists across Streamlit reruns and survives server restarts
# on local deployments. On cloud (Streamlit Cloud / Heroku ephemeral FS),
# the DB survives within a session; Export/Import buttons let accountants
# back up and restore their learned corrections.
# ═══════════════════════════════════════════════════════════════════

import sqlite3

MEMORY_DB = "vendor_memory.db"

def _get_db():
    """Return a SQLite connection, creating the table if needed."""
    conn = sqlite3.connect(MEMORY_DB, check_same_thread=False)
    conn.execute(
        "CREATE TABLE IF NOT EXISTS vendor_memory "
        "(key TEXT PRIMARY KEY, category TEXT, updated TEXT)"
    )
    conn.commit()
    return conn

def load_vendor_memory():
    """Load all vendor memory from SQLite into session_state on first run."""
    if "vendor_memory" not in st.session_state:
        try:
            conn = _get_db()
            rows = conn.execute("SELECT key, category FROM vendor_memory").fetchall()
            conn.close()
            st.session_state.vendor_memory = {r[0]: r[1] for r in rows}
        except Exception:
            st.session_state.vendor_memory = {}

def save_vendor_memory():
    """Persist all in-memory corrections back to SQLite (full sync)."""
    try:
        conn = _get_db()
        from datetime import datetime as _dt
        ts = _dt.now().isoformat()
        for key, cat in st.session_state.get("vendor_memory", {}).items():
            conn.execute(
                "INSERT INTO vendor_memory (key, category, updated) VALUES (?,?,?) "
                "ON CONFLICT(key) DO UPDATE SET category=excluded.category, updated=excluded.updated",
                (key, cat, ts)
            )
        conn.commit()
        conn.close()
    except Exception:
        pass  # Graceful degradation: memory survives in session_state

def remember_vendor(description, category):
    """Store a vendor→category correction in SQLite + session memory."""
    load_vendor_memory()
    key = clean_description(description)
    if key:
        st.session_state.vendor_memory[key] = category
        try:
            conn = _get_db()
            from datetime import datetime as _dt
            conn.execute(
                "INSERT INTO vendor_memory (key, category, updated) VALUES (?,?,?) "
                "ON CONFLICT(key) DO UPDATE SET category=excluded.category, updated=excluded.updated",
                (key, category, _dt.now().isoformat())
            )
            conn.commit()
            conn.close()
        except Exception:
            pass  # session_state write already succeeded

def recall_vendor(description):
    """Look up a vendor — checks session_state (already loaded from SQLite)."""
    load_vendor_memory()
    key = clean_description(description)
    return st.session_state.vendor_memory.get(key)

def export_vendor_memory_json():
    """Return vendor memory as JSON bytes for download."""
    load_vendor_memory()
    return json.dumps(st.session_state.vendor_memory, indent=2, ensure_ascii=False).encode("utf-8")

def import_vendor_memory_json(data_bytes):
    """Import vendor memory from uploaded JSON bytes. Merges with existing data."""
    try:
        incoming = json.loads(data_bytes.decode("utf-8"))
        load_vendor_memory()
        st.session_state.vendor_memory.update(incoming)
        save_vendor_memory()
        return len(incoming)
    except Exception as e:
        return 0




# ═══════════════════════════════════════════════════════════════════




def _get_clients_db():
    """Return SQLite connection with clients + rules tables."""
    conn = sqlite3.connect(MEMORY_DB, check_same_thread=False)
    conn.execute(
        "CREATE TABLE IF NOT EXISTS clients ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT, "
        "name TEXT UNIQUE, industry TEXT, province TEXT, "
        "structure TEXT, account_type TEXT, updated TEXT)"
    )
    conn.execute(
        "CREATE TABLE IF NOT EXISTS vendor_rules ("
        "keyword TEXT PRIMARY KEY, category TEXT, "
        "match_type TEXT, updated TEXT)"
    )
    conn.commit()
    return conn

def load_clients():
    """Return list of saved client dicts."""
    try:
        conn = _get_clients_db()
        rows = conn.execute(
            "SELECT name, industry, province, structure, account_type FROM clients ORDER BY name"
        ).fetchall()
        conn.close()
        return [{"name": r[0], "industry": r[1], "province": r[2],
                 "structure": r[3], "account_type": r[4]} for r in rows]
    except Exception:
        return []

def save_client(name, industry, province, structure, account_type):
    """Upsert a client profile."""
    try:
        conn = _get_clients_db()
        from datetime import datetime as _dt
        conn.execute(
            "INSERT INTO clients (name, industry, province, structure, account_type, updated) "
            "VALUES (?,?,?,?,?,?) ON CONFLICT(name) DO UPDATE SET "
            "industry=excluded.industry, province=excluded.province, "
            "structure=excluded.structure, account_type=excluded.account_type, "
            "updated=excluded.updated",
            (name, industry, province, structure, account_type, _dt.now().isoformat())
        )
        conn.commit()
        conn.close()
        return True
    except Exception:
        return False

def delete_client(name):
    """Remove a client profile."""
    try:
        conn = _get_clients_db()
        conn.execute("DELETE FROM clients WHERE name=?", (name,))
        conn.commit()
        conn.close()
        return True
    except Exception:
        return False


# ═══════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════

VALID_CATEGORIES = [
    "Motor Vehicle Expense", "Meals & Entertainment", "Office Supplies",
    "Utilities", "Bank Charges", "Insurance", "Materials & Supplies",
    "Rent", "Advertising", "Travel", "Professional Fees", "Subcontracts",
    "Cost of Goods", "Repairs & Maintenance", "Delivery & Shipping",
    "Government Remittances", "Owner Draw / Personal",
    "Shareholder Loan (Debit)", "❓ Uncategorized",
]

def load_vendor_rules():
    """Load all vendor rules from SQLite into session_state."""
    if "vendor_rules" not in st.session_state:
        try:
            conn = _get_clients_db()
            rows = conn.execute(
                "SELECT keyword, category, match_type FROM vendor_rules ORDER BY keyword"
            ).fetchall()
            conn.close()
            st.session_state.vendor_rules = [
                {"keyword": r[0], "category": r[1], "match_type": r[2]} for r in rows
            ]
        except Exception:
            st.session_state.vendor_rules = []

def save_vendor_rule(keyword, category, match_type="contains"):
    """Upsert a vendor rule."""
    load_vendor_rules()
    try:
        conn = _get_clients_db()
        from datetime import datetime as _dt
        conn.execute(
            "INSERT INTO vendor_rules (keyword, category, match_type, updated) VALUES (?,?,?,?) "
            "ON CONFLICT(keyword) DO UPDATE SET category=excluded.category, "
            "match_type=excluded.match_type, updated=excluded.updated",
            (keyword.strip().upper(), category, match_type, _dt.now().isoformat())
        )
        conn.commit()
        conn.close()
        # Refresh session state
        st.session_state.vendor_rules = [
            r for r in st.session_state.vendor_rules
            if r["keyword"] != keyword.strip().upper()
        ]
        st.session_state.vendor_rules.append(
            {"keyword": keyword.strip().upper(), "category": category, "match_type": match_type}
        )
        return True
    except Exception:
        return False

def delete_vendor_rule(keyword):
    """Remove a vendor rule."""
    try:
        conn = _get_clients_db()
        conn.execute("DELETE FROM vendor_rules WHERE keyword=?", (keyword.strip().upper(),))
        conn.commit()
        conn.close()
        load_vendor_rules()
        st.session_state.vendor_rules = [
            r for r in st.session_state.get("vendor_rules", [])
            if r["keyword"] != keyword.strip().upper()
        ]
        return True
    except Exception:
        return False

def apply_vendor_rules(transactions):
    """
    Apply permanent vendor rules to transactions BEFORE AI categorization fallback.
    This is the deterministic override layer: rules always win over AI guesses.
    Returns (transactions, rules_applied_count).
    """
    load_vendor_rules()
    rules = st.session_state.get("vendor_rules", [])
    if not rules:
        return transactions, 0

    applied = 0
    for t in transactions:
        desc = t.get("description", "").upper()
        for rule in rules:
            kw = rule["keyword"]
            mt = rule.get("match_type", "contains")
            matched = (
                (mt == "contains" and kw in desc) or
                (mt == "exact" and kw == desc) or
                (mt == "startswith" and desc.startswith(kw))
            )
            if matched:
                old_cat = t.get("category", "")
                if old_cat != rule["category"]:
                    t["category"] = rule["category"]
                    t["notes"] = (t.get("notes", "") + f" RULE:{kw}").strip()
                    applied += 1
                break  # First matching rule wins
    return transactions, applied

# API RETRY HELPER (v3.4)
# Handles 529 overload and transient errors with exponential backoff.
# Without this, a single rate-limit on batch 3-of-10 silently drops
# all transactions in that batch — a financial data integrity failure.


# ═══════════════════════════════════════════════════════════════════
# UTILITY FUNCTIONS


# ═══════════════════════════════════════════════════════════════════
# CSV PARSER — TD, RBC, BMO, CIBC, Scotiabank exports
# ═══════════════════════════════════════════════════════════════════
import csv


# ═══════════════════════════════════════════════════════════════════
# CLIENT PROFILES — Multi-client workspace (v3.8)
# Accountants managing 50-200 clients can save/load business profiles
# so they never retype name, industry, province, structure each session.
# Stored in SQLite alongside vendor memory.
# ═══════════════════════════════════════════════════════════════════



st.set_page_config(page_title=f"BookKeep AI Pro v{VERSION}", layout="wide")

# ── SESSION STATE ───────────────────────────────────────────────────
for k, v in {"transactions": None, "flags": [], "summary": {}, "pass_number": 0,
             "raw_response": "", "total_cost": 0, "receipts_data": [], "detected_bank": "",
             "recon_matches": [], "recon_unmatched": [], "receipt_matches": [], "invoice_data": [],
             "validation_results": [], "validation_report": [], "anomalies": [],
             "t5018_data": [], "coverage_info": {}, "normalization_count": 0, "audit_trail": []}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── LOAD INSTRUCTIONS ───────────────────────────────────────────────
INSTRUCTIONS_PATH = "instructions.txt"
try:
    with open(INSTRUCTIONS_PATH, "r", encoding="utf-8") as f:
        FULL_INSTRUCTIONS = f.read()
    full_tokens = len(FULL_INSTRUCTIONS) // 4
except FileNotFoundError:
    st.error("❌ instructions.txt not found in app directory")
    st.stop()

# ── API KEY ─────────────────────────────────────────────────────────
api_key = ""
if os.path.exists("api_key.txt"):
    with open("api_key.txt", "r") as f:
        api_key = f.read().strip()
if not api_key:
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
if not api_key:
    api_key = st.sidebar.text_input("🔑 Anthropic API Key", type="password")
if not api_key:
    st.warning("Enter your Anthropic API key in the sidebar, api_key.txt, or ANTHROPIC_API_KEY env var.")
    st.stop()

client = anthropic.Anthropic(api_key=api_key)

# ── SMART CONDENSED PROMPT ──────────────────────────────────────────
SMART_CONDENSED = """You are BookKeep AI — deterministic Canadian accounting transformer.
Output ONLY TSV rows. No headings, no markdown, no commentary.
10 columns: Date|Source|Description|Debit|Credit|Type|Category|ITC Amount|Confidence|Notes

Input: Date|Description|Amount (positive=debit, negative=credit)
Type: PURCHASE, REFUND, PAYMENT, FEE, INTEREST, OTHER
Category (PURCHASE/FEE only): Motor Vehicle Expense, Meals & Entertainment, Office Supplies, Utilities, Bank Charges, Materials & Supplies, Insurance, Rent, Government Remittances, Owner Draw / Personal, Uncategorized
ITC: Debit×13/113. Meals: add MEALS_50_RULE. Payment/Personal/Insurance: blank.
If Debit>=500: add POTENTIAL_CCA_ASSET. Confidence: 60-95 realistic.

CRITICAL — DO NOT EXTRACT SECTION HEADERS:
"Payments received [DATE] to [DATE]" lines are SUMMARIES — skip them. Only extract individual payment rows with their own transaction date.
"Returns and other credits" totals are SUMMARIES — skip them. Only extract individual return rows.
Date column: use POSTING DATE when both dates shown, not transaction date.
Return TSV rows only."""

# ── SIDEBAR ─────────────────────────────────────────────────────────
st.sidebar.header("⚙️ Settings")

prompt_mode = st.sidebar.radio("System Prompt",
    ["Full (best accuracy, cached)", "Smart Condensed (fast)"], index=0)
use_full = "Full" in prompt_mode

MODEL = st.sidebar.selectbox("Model",
    ["claude-sonnet-4-6 (Fast — recommended)", "claude-opus-4-6 (Slow — complex only)"], index=0)
model_id = MODEL.split(" ")[0]
max_pages = st.sidebar.slider("Max PDF Pages", 1, 100, 60)

st.sidebar.divider()
st.sidebar.subheader("💰 Cost Estimate")
sys_tokens = full_tokens if use_full else 2000
rate_in = 3.0 if "sonnet" in model_id else 5.0
rate_out = 15.0 if "sonnet" in model_id else 25.0
per_request = (sys_tokens*0.3/1e6*rate_in) + (max_pages*500*rate_in/1e6) + (2000*rate_out/1e6)
first_request = (sys_tokens*1.0/1e6*rate_in) + (max_pages*500*rate_in/1e6) + (2000*rate_out/1e6)
st.sidebar.success(f"**Per statement:** ~${per_request:.3f} (cached)\n**First request:** ~${first_request:.3f}\n**Daily(20):** ${per_request*19 + first_request:.2f}\n**Monthly(440):** ${per_request*439 + first_request:.2f}")

supported = get_supported_banks()
st.sidebar.divider()
st.sidebar.caption(f"🏦 {len(supported)} banks supported — Auto-detect active")
with st.sidebar.expander("View supported banks", expanded=False):
    # Group by type for clean display
    from bank_schemas import BANK_SCHEMAS
    cc_banks = [s["bank"] for s in BANK_SCHEMAS.values() if s.get("type") == "credit_card"]
    chq_banks = [s["bank"] for s in BANK_SCHEMAS.values() if s.get("type") == "chequing"]
    st.caption(f"**Credit Cards ({len(cc_banks)}):** " + " · ".join(cc_banks))
    st.caption(f"**Chequing/Business ({len(chq_banks)}):** " + " · ".join(chq_banks))

# ── VENDOR MEMORY CONTROLS ──────────────────────────────────────────
st.sidebar.divider()
st.sidebar.subheader("🧠 Vendor Memory")
load_vendor_memory()
mem_count = len(st.session_state.get("vendor_memory", {}))
st.sidebar.caption(f"{mem_count} learned corrections")

# Export
if mem_count > 0:
    mem_json = export_vendor_memory_json()
    st.sidebar.download_button(
        label="⬇️ Export Memory (.json)",
        data=mem_json,
        file_name="vendor_memory.json",
        mime="application/json",
        use_container_width=True,
        help="Download your learned vendor→category corrections as a backup."
    )
    if st.sidebar.button("🗑️ Clear All Memory", use_container_width=True):
        st.session_state.vendor_memory = {}
        try:
            import sqlite3
            conn = sqlite3.connect(MEMORY_DB, check_same_thread=False)
            conn.execute("DELETE FROM vendor_memory")
            conn.commit(); conn.close()
        except Exception:
            pass
        st.sidebar.success("Memory cleared.")

# Import
uploaded_mem = st.sidebar.file_uploader(
    "⬆️ Import Memory (.json)", type=["json"],
    key="mem_upload",
    help="Restore a previously exported vendor_memory.json to continue learning."
)
if uploaded_mem is not None:
    n = import_vendor_memory_json(uploaded_mem.read())
    if n:
        st.sidebar.success(f"Imported {n} corrections.")
    else:
        st.sidebar.error("Import failed — invalid JSON.")

# ── VENDOR RULE EDITOR ─────────────────────────────────────────────
st.sidebar.divider()
st.sidebar.subheader("📋 Vendor Rules")
st.sidebar.caption("Rules override AI — applied before categorization.")
load_vendor_rules()
rules_list = st.session_state.get("vendor_rules", [])

# Show existing rules with delete buttons
if rules_list:
    for rule in rules_list:
        _rc1, _rc2 = st.sidebar.columns([4, 1])
        _rc1.caption(f"`{rule['keyword']}` → **{rule['category']}** ({rule['match_type']})")
        if _rc2.button("✕", key=f"del_rule_{rule['keyword']}"):
            delete_vendor_rule(rule['keyword'])
            st.rerun()
else:
    st.sidebar.caption("No rules yet.")

# Add new rule form
with st.sidebar.expander("➕ Add Rule"):
    _rule_kw = st.text_input("Keyword (e.g. NETFLIX)", key="rule_kw").strip().upper()
    _rule_cat = st.selectbox("Category", VALID_CATEGORIES, key="rule_cat")
    _rule_mt  = st.selectbox("Match type", ["contains", "exact", "startswith"], key="rule_mt")
    if st.button("Save Rule", use_container_width=True) and _rule_kw:
        if save_vendor_rule(_rule_kw, _rule_cat, _rule_mt):
            st.success(f"Rule saved: {_rule_kw} → {_rule_cat}")
            st.rerun()
        else:
            st.error("Save failed.")

# ── HEADER ──────────────────────────────────────────────────────────
st.title("📊 BookKeep AI Pro")
st.caption(f"v{VERSION} | 31 Banks & Cards | Client Profiles | Rule Editor | Merchant Normalization | T5018 | CCA Classes | Validation Engine | Claude 4.6 | ~${per_request:.3f}/statement")

# ═══════════════════════════════════════════════════════════════════
# DEMO MODE — Pre-built sample data, zero API cost
# ═══════════════════════════════════════════════════════════════════
with st.expander("🎯 Try a Free Demo — No Upload Required", expanded=False):
    st.markdown("""
    **See exactly what BookKeep AI Pro produces** — before uploading your own statement.  
    This demo uses a realistic TD Business Chequing statement for a construction client (Ontario).  
    Click below to generate a full 10-tab Excel instantly — **no PDF upload needed, zero API cost.**
    """)
    if st.button("🚀 Run Demo Statement", type="primary", use_container_width=True):
        demo_transactions = [
            {"date":"2025-01-03","source":"TD Business Chequing","description":"PETRO CANADA #4821 BRAMPTON ON","debit":82.40,"credit":0,"balance":0,"type":"PURCHASE","category":"Motor Vehicle Expense","t2125":"9281","biz_pct":100,"itc_rule":"Full","itc_amount":9.50,"confidence":"95","notes":""},
            {"date":"2025-01-06","source":"TD Business Chequing","description":"HOME DEPOT #7823 BRAMPTON ON","debit":347.82,"credit":0,"balance":0,"type":"PURCHASE","category":"Materials & Supplies","t2125":"8811","biz_pct":100,"itc_rule":"Full","itc_amount":40.11,"confidence":"92","notes":""},
            {"date":"2025-01-08","source":"TD Business Chequing","description":"TIM HORTONS #0912 BRAMPTON ON","debit":18.75,"credit":0,"balance":0,"type":"PURCHASE","category":"Meals & Entertainment","t2125":"8523","biz_pct":50,"itc_rule":"50%","itc_amount":1.08,"confidence":"95","notes":"MEALS_50_RULE"},
            {"date":"2025-01-10","source":"TD Business Chequing","description":"WSIB ONTARIO PMT 28374","debit":420.00,"credit":0,"balance":0,"type":"FEE","category":"Government Remittances","t2125":"","biz_pct":100,"itc_rule":"No","itc_amount":0,"confidence":"95","notes":""},
            {"date":"2025-01-12","source":"TD Business Chequing","description":"AMZN MKTP CA*Z99335U2","debit":124.99,"credit":0,"balance":0,"type":"PURCHASE","category":"\u2753 Uncategorized","t2125":"","biz_pct":100,"itc_rule":"No","itc_amount":0,"confidence":"65","notes":"VERIFY_RECEIPT_AMAZON"},
            {"date":"2025-01-14","source":"TD Business Chequing","description":"COSTCO WHOLESALE W126 MISSISSAUGA ON","debit":231.74,"credit":0,"balance":0,"type":"PURCHASE","category":"Materials & Supplies","t2125":"8811","biz_pct":100,"itc_rule":"Full","itc_amount":26.71,"confidence":"91","notes":""},
            {"date":"2025-01-15","source":"TD Business Chequing","description":"INTERAC E-TFR MIKE JOHNSON","debit":650.00,"credit":0,"balance":0,"type":"PURCHASE","category":"Subcontracts","t2125":"8590","biz_pct":100,"itc_rule":"No","itc_amount":0,"confidence":"88","notes":"T5018_CANDIDATE"},
            {"date":"2025-01-17","source":"TD Business Chequing","description":"BELL CANADA 8003102355","debit":89.99,"credit":0,"balance":0,"type":"PURCHASE","category":"Utilities","t2125":"8220","biz_pct":100,"itc_rule":"Full","itc_amount":10.37,"confidence":"95","notes":""},
            {"date":"2025-01-20","source":"TD Business Chequing","description":"RONA #7234 BRAMPTON ON","debit":412.55,"credit":0,"balance":0,"type":"PURCHASE","category":"Materials & Supplies","t2125":"8811","biz_pct":100,"itc_rule":"Full","itc_amount":47.55,"confidence":"93","notes":""},
            {"date":"2025-01-22","source":"TD Business Chequing","description":"CANADIAN TIRE #0234 BRAMPTON ON","debit":156.30,"credit":0,"balance":0,"type":"PURCHASE","category":"Materials & Supplies","t2125":"8811","biz_pct":100,"itc_rule":"Full","itc_amount":18.02,"confidence":"88","notes":""},
            {"date":"2025-01-24","source":"TD Business Chequing","description":"DEWALT TOOLS #8831 TORONTO ON","debit":612.00,"credit":0,"balance":0,"type":"PURCHASE","category":"Materials & Supplies","t2125":"8811","biz_pct":100,"itc_rule":"Full","itc_amount":70.55,"confidence":"93","notes":"CCA_ASSET CCA_CLASS_8"},
            {"date":"2025-01-25","source":"TD Business Chequing","description":"TD SERVICE FEE","debit":16.95,"credit":0,"balance":0,"type":"FEE","category":"Bank Charges","t2125":"8710","biz_pct":100,"itc_rule":"No","itc_amount":0,"confidence":"95","notes":""},
            {"date":"2025-01-27","source":"TD Business Chequing","description":"WALMART #3821 BRAMPTON ON","debit":87.43,"credit":0,"balance":0,"type":"PURCHASE","category":"\u2753 Uncategorized","t2125":"","biz_pct":100,"itc_rule":"No","itc_amount":0,"confidence":"62","notes":""},
            {"date":"2025-01-28","source":"TD Business Chequing","description":"ESSO #4421 MISSISSAUGA ON","debit":94.20,"credit":0,"balance":0,"type":"PURCHASE","category":"Motor Vehicle Expense","t2125":"9281","biz_pct":100,"itc_rule":"Full","itc_amount":10.86,"confidence":"95","notes":""},
            {"date":"2025-01-29","source":"TD Business Chequing","description":"SUBWAY #32981 BRAMPTON ON","debit":23.45,"credit":0,"balance":0,"type":"PURCHASE","category":"Meals & Entertainment","t2125":"8523","biz_pct":50,"itc_rule":"50%","itc_amount":1.35,"confidence":"95","notes":"MEALS_50_RULE"},
            {"date":"2025-01-30","source":"TD Business Chequing","description":"E-TRANSFER RECEIVED CLIENT PMT","debit":0,"credit":4500.00,"balance":0,"type":"PAYMENT","category":"","t2125":"","biz_pct":100,"itc_rule":"No","itc_amount":0,"confidence":"95","notes":""},
            {"date":"2025-01-31","source":"TD Business Chequing","description":"HOME DEPOT #7823 BRAMPTON ON","debit":189.90,"credit":0,"balance":0,"type":"PURCHASE","category":"Materials & Supplies","t2125":"8811","biz_pct":100,"itc_rule":"Full","itc_amount":21.89,"confidence":"92","notes":""},
            {"date":"2025-02-03","source":"TD Business Chequing","description":"INTERAC E-TFR STEVE WILLIAMS","debit":800.00,"credit":0,"balance":0,"type":"PURCHASE","category":"Subcontracts","t2125":"8590","biz_pct":100,"itc_rule":"No","itc_amount":0,"confidence":"88","notes":"T5018_CANDIDATE"},
            {"date":"2025-02-05","source":"TD Business Chequing","description":"STAPLES #1234 BRAMPTON ON","debit":67.80,"credit":0,"balance":0,"type":"PURCHASE","category":"Office Supplies","t2125":"8810","biz_pct":100,"itc_rule":"Full","itc_amount":7.81,"confidence":"94","notes":""},
            {"date":"2025-02-07","source":"TD Business Chequing","description":"HYDRO ONE NETWORKS INC","debit":234.56,"credit":0,"balance":0,"type":"PURCHASE","category":"Utilities","t2125":"8220","biz_pct":100,"itc_rule":"Full","itc_amount":27.03,"confidence":"95","notes":""},
        ]

        demo_t5018 = [
            {"payee":"MIKE JOHNSON","count":1,"total":650.00,"t5018_required":True},
            {"payee":"STEVE WILLIAMS","count":1,"total":800.00,"t5018_required":True},
        ]

        T2125_MAP = {
            "Motor Vehicle Expense": "9281", "Meals & Entertainment": "8523",
            "Office Supplies": "8810", "Utilities": "8220", "Bank Charges": "8710",
            "Insurance": "8690", "Materials & Supplies": "8811", "Rent": "8910",
            "Delivery & Shipping": "8730", "Advertising": "8520",
            "Travel": "9200", "Professional Fees": "8860", "Subcontracts": "8590",
            "Cost of Goods": "8320", "Repairs & Maintenance": "8960",
        }
        BIZ_USE = {"Motor Vehicle Expense": 100, "Meals & Entertainment": 50, "Owner Draw / Personal": 0}

        _wb = build_excel(
            demo_transactions, [],
            {"period": "Jan-Feb 2025", "transactions": str(len(demo_transactions))},
            "Demo Construction Co.", "Construction/Trades", "Ontario", "Jan-Feb 2025",
            recon_matches=None, recon_unmatched=None, receipt_matches=None,
            invoice_data=None, validation_results=None,
            t5018_data=demo_t5018, validation_report=None,
            anomalies=None, audit_trail=None
        )

        import io as _io
        _buf = _io.BytesIO()
        _wb.save(_buf)
        _buf.seek(0)

        total_exp_d = sum(t["debit"] for t in demo_transactions if t["debit"])
        total_itc_d = sum(t.get("itc_amount",0) for t in demo_transactions)
        needs_review_d = len([t for t in demo_transactions if "Uncategorized" in t.get("category","")])

        st.success(f"\u2705 Demo complete! {len(demo_transactions)} transactions — 10-tab Excel ready to download.")
        _dc1, _dc2, _dc3, _dc4 = st.columns(4)
        _dc1.metric("Transactions", len(demo_transactions))
        _dc2.metric("Total Expenses", f"${total_exp_d:,.2f}")
        _dc3.metric("HST ITC Claimable", f"${total_itc_d:,.2f}")
        _dc4.metric("Needs Review", needs_review_d)

        st.info("\U0001f4cb **2 T5018 subcontractor payments** flagged (Mike Johnson $650 + Steve Williams $800) · **1 CCA asset** detected (DeWalt Tools $612 \u2192 Class 8) · **2 items** in Needs Review (Amazon + Walmart \u2014 ambiguous without receipts)")

        st.download_button(
            "\u2b07\ufe0f Download Demo Excel \u2014 10 Tabs, CRA-Ready",
            data=_buf,
            file_name=f"Demo_Construction_BookKeepAI_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )
        st.markdown("---")
        st.markdown("**Ready to process your own statement?** Upload a PDF above and click Process Statement \u2014 results in under 2 minutes.")



# ── STEP 1: BUSINESS INFO ──────────────────────────────────────────
st.header("1️⃣ Business Info")

# v3.8: Client profile loader
_clients = load_clients()
_client_names = ["— New Client —"] + [c["name"] for c in _clients]
_col_load, _col_save, _col_del = st.columns([3, 1, 1])
with _col_load:
    _selected_client = st.selectbox("👤 Client Profile", _client_names,
        help="Load a saved client to pre-fill all fields below.")
with _col_save:
    _save_btn = st.button("💾 Save Profile", use_container_width=True,
        help="Save current fields as a reusable client profile.")
with _col_del:
    _del_btn = st.button("🗑️ Delete", use_container_width=True,
        disabled=(_selected_client == "— New Client —"),
        help="Delete the selected client profile.")

# Pre-fill defaults from selected client
_defaults = {}
if _selected_client != "— New Client —":
    _defaults = next((c for c in _clients if c["name"] == _selected_client), {})

c1, c2, c3 = st.columns(3)
with c1:
    business_name = st.text_input("Business Name", _defaults.get("name", "Client"))
    industry = st.selectbox("Industry",
        ["Construction/Trades", "Retail", "Restaurant/Food", "Professional Services", "Rental Properties", "Other"],
        index=["Construction/Trades", "Retail", "Restaurant/Food", "Professional Services", "Rental Properties", "Other"].index(_defaults["industry"]) if _defaults.get("industry") in ["Construction/Trades", "Retail", "Restaurant/Food", "Professional Services", "Rental Properties", "Other"] else 0)
with c2:
    _acct_opts = ["Business Chequing", "Business Credit Card", "Personal (flag business items)", "PayPal/Square/Stripe"]
    account_type = st.selectbox("Account Type", _acct_opts,
        index=_acct_opts.index(_defaults["account_type"]) if _defaults.get("account_type") in _acct_opts else 0)
    period = st.text_input("Period (optional)", placeholder="e.g. Jan-Mar 2024")
with c3:
    bank_override = st.selectbox("Bank (auto-detected, override if wrong)",
        ["Auto-Detect"] + supported + ["Other"])
    _prov_opts = ["Auto-Detect", "Ontario", "Quebec", "Alberta", "BC", "Saskatchewan", "Manitoba", "Nova Scotia", "New Brunswick", "Other"]
    province_override = st.selectbox("Province (auto-detected from statement)", _prov_opts,
        index=_prov_opts.index(_defaults["province"]) if _defaults.get("province") in _prov_opts else 0)

# v3.3: Business structure for shareholder loan detection
c4, c5 = st.columns(2)
with c4:
    _struct_opts = ["Sole Proprietor", "Corporation", "Inc.", "Ltd.", "Partnership"]
    business_structure = st.selectbox("Business Structure", _struct_opts,
        index=_struct_opts.index(_defaults["structure"]) if _defaults.get("structure") in _struct_opts else 0,
        help="For corporations: personal expenses are flagged as Shareholder Loan (Debit) instead of Owner Draw")
with c5:
    anomaly_detection = st.checkbox("Expense Anomaly Detection", value=True,
        help="Flag transactions unusually large vs category median")

# Handle Save / Delete profile actions
if _save_btn and business_name.strip():
    if save_client(business_name.strip(), industry, province_override, business_structure, account_type):
        st.success(f"✅ Profile saved: {business_name}")
    else:
        st.error("Failed to save profile.")
if _del_btn and _selected_client != "— New Client —":
    if delete_client(_selected_client):
        st.success(f"🗑️ Deleted profile: {_selected_client}")
    else:
        st.error("Failed to delete profile.")

# ── STEP 2: UPLOAD ──────────────────────────────────────────────────
st.header("2️⃣ Upload Files")
c1, c2, c3 = st.columns(3)
with c1:
    statement_files = st.file_uploader("📄 Statements (PDF/CSV) — multi-file for reconciliation",
        type=["pdf","csv","tsv"], key="stmt", accept_multiple_files=True)
with c2:
    receipt_files = st.file_uploader("📸 Receipts — matched to transactions",
        type=["pdf","png","jpg","jpeg"], accept_multiple_files=True, key="rcpts")
with c3:
    invoice_files = st.file_uploader("📋 Invoices — Accounts Payable",
        type=["pdf","png","jpg","jpeg"], accept_multiple_files=True, key="inv")



def call_claude_with_retry(client, model_id, max_tokens, system_config, messages,
                            timeout=300.0, max_retries=3, label="API call"):
    """
    Stream a Claude API call with exponential backoff retry.
    Returns the full response text, or raises on final failure.
    Raises a descriptive exception so the caller can show a targeted warning.
    """
    for attempt in range(max_retries):
        try:
            resp = ""
            with client.messages.stream(
                model=model_id, max_tokens=max_tokens,
                system=system_config, messages=messages, timeout=timeout
            ) as stream:
                for txt in stream.text_stream:
                    resp += txt
            return resp
        except Exception as e:
            err_str = str(e).lower()
            is_retryable = any(x in err_str for x in
                               ("529", "overloaded", "rate limit", "timeout", "connection", "503", "502"))
            if is_retryable and attempt < max_retries - 1:
                wait = 5 * (2 ** attempt)   # 5s, 10s, 20s
                st.warning(f"⏳ {label}: {type(e).__name__} — retrying in {wait}s (attempt {attempt+2}/{max_retries})")
                time.sleep(wait)
            else:
                raise RuntimeError(f"{label} failed after {attempt+1} attempt(s): {e}") from e
    return ""   # unreachable, but satisfies linter


def validate_tsv_line(parts, min_cols=4):
    """
    Validate that a TSV response line has enough columns and a plausible date.
    Returns True if the line looks like a real transaction row.
    Filters out header echoes, markdown fences, and malformed rows.
    """
    if len(parts) < min_cols:
        return False
    date_col = parts[0].strip()
    # Must have something date-like in the first column
    if not date_col or date_col.lower() in ("date", "---", ""):
        return False
    # Reject markdown fences or explanation lines
    if date_col.startswith(("```", "#", "//", "Note", "Total", "Summary")):
        return False
    # Must have at least one non-empty amount-ish column
    for i in range(1, min(5, len(parts))):
        if parts[i].strip() and any(c.isdigit() for c in parts[i]):
            return True
    return False


# ═══════════════════════════════════════════════════════════════════
def reconcile_all(all_txns_by_source, receipt_matches=None, invoice_data=None, period_hint=""):
    """Full cross-document reconciliation across statements, receipts, and invoices.
    period_hint is passed to date parsing to fix the year-1900 reconciliation bug.
    Returns (matches, unmatched, duplicates)."""
    matches = []
    unmatched = []
    duplicates = []

    # Tag all transactions with source
    all_txns = []
    for source, txns in all_txns_by_source.items():
        for t in txns:
            t["_recon_source"] = source
            all_txns.append(t)

    # ── 1. CC Payment ↔ Chequing Debit ──
    cc_payments = [t for t in all_txns if t.get("type") == "PAYMENT" and t.get("credit", 0) > 0]
    cc_keywords = ["VISA","MASTERCARD","MC PAYMENT","CIBC","AMEX","CREDIT CARD","CC PAYMENT","TRIANGLE"]
    cheq_cc_debits = [t for t in all_txns if t.get("debit", 0) > 0
                      and any(kw in t.get("description","").upper() for kw in cc_keywords)]

    used_cheq = set()
    for cc in cc_payments:
        best, best_diff = _find_best_match(cc["credit"], cc["date"], cheq_cc_debits, used_cheq, "debit", 0.50, 5, period_hint)
        if best is not None:
            ch = cheq_cc_debits[best]
            used_cheq.add(best)
            matches.append(_make_match("CC_PAYMENT↔CHEQUING", cc, ch, best_diff))
            cc["notes"] = (cc.get("notes","") + f" RECON:CC↔{ch['_recon_source']}").strip()
            ch["notes"] = (ch.get("notes","") + f" RECON:CC↔{cc['_recon_source']}").strip()
        else:
            unmatched.append({"type":"CC_PAYMENT","date":cc["date"],"source":cc.get("_recon_source",""),
                "desc":cc["description"],"amount":cc["credit"],"status":"UNMATCHED"})
    for i, ch in enumerate(cheq_cc_debits):
        if i not in used_cheq:
            unmatched.append({"type":"CHEQ_CC_DEBIT","date":ch["date"],"source":ch.get("_recon_source",""),
                "desc":ch["description"],"amount":ch["debit"],"status":"UNMATCHED"})

    # ── 2. Invoice ↔ Statement Transaction ──
    if invoice_data:
        used_txn_inv = set()
        for inv in invoice_data:
            inv_total = inv.get("total", 0)
            inv_vendor = inv.get("vendor", "").upper()
            if inv_total <= 0: continue

            best_idx = None; best_score = -1
            for i, t in enumerate(all_txns):
                if i in used_txn_inv: continue
                t_amt = t.get("debit", 0) or t.get("credit", 0)
                if abs(t_amt - inv_total) > 2.0: continue
                t_desc = t.get("description", "").upper()
                vendor_words = [w for w in inv_vendor.split() if len(w) > 3]
                word_matches = sum(1 for w in vendor_words if w in t_desc)
                inv_dt = _parse_date_fuzzy(inv.get("date",""), period_hint)
                t_dt = _parse_date_fuzzy(t.get("date",""), period_hint)
                date_diff = abs((inv_dt - t_dt).days) if inv_dt and t_dt else 0
                if date_diff <= 30:
                    score = word_matches * 10 + (30 - date_diff)
                    if score > best_score:
                        best_score = score; best_idx = i

            if best_idx is not None and best_score > 0:
                t = all_txns[best_idx]
                used_txn_inv.add(best_idx)
                inv_dt = _parse_date_fuzzy(inv.get("date",""), period_hint) or datetime.now()
                t_dt = _parse_date_fuzzy(t["date"], period_hint) or datetime.now()
                matches.append({
                    "match_type": "INVOICE↔STATEMENT",
                    "doc_a": f"Invoice: {inv['vendor']} #{inv.get('invoice_num','')}",
                    "date_a": inv.get("date",""), "amount_a": inv_total,
                    "source_a": inv.get("file",""),
                    "doc_b": t["description"], "date_b": t["date"],
                    "amount_b": t.get("debit",0) or t.get("credit",0),
                    "source_b": t.get("_recon_source",""),
                    "day_diff": abs((inv_dt - t_dt).days),
                    "status": "MATCHED"
                })
                t["notes"] = (t.get("notes","") + f" RECON:INV#{inv.get('invoice_num','')}").strip()
            else:
                unmatched.append({"type":"INVOICE","date":inv.get("date",""),"source":inv.get("file",""),
                    "desc":f"{inv['vendor']} #{inv.get('invoice_num','')}","amount":inv_total,"status":"UNMATCHED"})

    # ── 3. Receipt ↔ Transaction ──
    if receipt_matches:
        for rm in receipt_matches:
            if rm["status"] == "MATCHED":
                matches.append({
                    "match_type": "RECEIPT↔STATEMENT",
                    "doc_a": f"Receipt: {rm['vendor']}", "date_a": rm["date"],
                    "amount_a": rm["total"], "source_a": rm["file"],
                    "doc_b": rm["matched_txn"], "date_b": rm["matched_date"],
                    "amount_b": rm["total"], "source_b": "Statement",
                    "day_diff": 0, "status": "MATCHED"
                })
            else:
                unmatched.append({"type":"RECEIPT","date":rm["date"],"source":rm["file"],
                    "desc":rm["vendor"],"amount":rm["total"],"status":"UNMATCHED"})

    # ── 4. Cross-Statement Duplicate Detection ──
    if len(all_txns_by_source) > 1:
        seen = {}
        for t in all_txns:
            amt = t.get("debit",0) or t.get("credit",0)
            key = f"{t.get('description','')[:20]}|{amt:.2f}"
            t_dt = _parse_date_fuzzy(t.get("date",""), period_hint)
            if key in seen:
                prev_t, prev_dt = seen[key]
                if prev_t.get("_recon_source") != t.get("_recon_source"):
                    if t_dt and prev_dt and abs((t_dt - prev_dt).days) <= 2:
                        if t.get("type") not in ("PAYMENT","FEE_REBATE"):
                            duplicates.append({
                                "desc": t["description"], "amount": amt,
                                "source_1": prev_t.get("_recon_source",""), "date_1": prev_t["date"],
                                "source_2": t.get("_recon_source",""), "date_2": t["date"],
                                "status": "POTENTIAL_DUPLICATE"
                            })
                            t["notes"] = (t.get("notes","") + " POTENTIAL_DUPLICATE").strip()
                            prev_t["notes"] = (prev_t.get("notes","") + " POTENTIAL_DUPLICATE").strip()
            else:
                seen[key] = (t, t_dt)

    return matches, unmatched, duplicates

def _find_best_match(amount, date_str, candidates, used_set, amt_field, amt_tol, day_tol, period_hint=""):
    """Find best matching transaction by amount and date proximity."""
    dt = _parse_date_fuzzy(date_str, period_hint)
    best_idx = None; best_diff = 999
    for i, c in enumerate(candidates):
        if i in used_set: continue
        c_amt = c.get(amt_field, 0)
        if abs(amount - c_amt) > amt_tol: continue
        c_dt = _parse_date_fuzzy(c.get("date",""), period_hint)
        day_diff = abs((dt - c_dt).days) if dt and c_dt else 0
        if day_diff <= day_tol and day_diff < best_diff:
            best_diff = day_diff; best_idx = i
    return best_idx, best_diff

def _make_match(match_type, a, b, day_diff):
    return {
        "match_type": match_type,
        "doc_a": a["description"], "date_a": a["date"],
        "amount_a": a.get("credit",0) or a.get("debit",0), "source_a": a.get("_recon_source",""),
        "doc_b": b["description"], "date_b": b["date"],
        "amount_b": b.get("debit",0) or b.get("credit",0), "source_b": b.get("_recon_source",""),
        "day_diff": day_diff, "status": "MATCHED"
    }

def _parse_date_fuzzy(date_str, period_hint=""):
    """
    Wrapper around bank_schemas.parse_date_fuzzy.
    Passes period_hint so year-less dates (e.g. 'Dec 23') resolve correctly
    instead of defaulting to 1900 and breaking reconciliation day-diffs.
    """
    return parse_date_fuzzy(date_str, period_hint=period_hint)


def extract_pdf_hybrid(uploaded_file, page_limit=30):
    """Page-by-page PDF extraction with CID garble detection.
    Returns (good_text, bad_page_numbers, total_pages).
    - good_text: concatenated text from readable pages
    - bad_page_numbers: list of 0-indexed pages that need Vision fallback
    """
    good_text = ""
    bad_pages = []
    total_pages = 0
    page_texts = {}  # page_num -> text (for readable pages)
    uploaded_file.seek(0)
    try:
        with pdfplumber.open(uploaded_file) as pdf:
            total_pages = len(pdf.pages)
            for i, page in enumerate(pdf.pages):
                if i >= page_limit: break
                try:
                    t = page.extract_text() or ""
                except Exception:
                    t = ""
                if is_page_garbled(t):
                    bad_pages.append(i)
                else:
                    good_text += t + "\n"
                    page_texts[i] = t
    except Exception:
        pass
    return good_text.strip(), bad_pages, total_pages, page_texts

def is_page_garbled(text):
    """Detect CID-encoded or garbled pages that pdfplumber can't read."""
    if not text or len(text.strip()) < 30:
        return True
    # CID font markers: (cid:###) patterns
    cid_count = len(re.findall(r'\(cid:\d+\)', text))
    if cid_count > 5:
        return True
    # High ratio of non-printable or replacement characters
    garbage = sum(1 for c in text if ord(c) > 65535 or c == '\ufffd' or (ord(c) < 32 and c not in '\n\r\t'))
    if len(text) > 0 and garbage / len(text) > 0.15:
        return True
    # Too few alphanumeric vs total (garbled PDFs have lots of symbols)
    alnum = sum(1 for c in text if c.isalnum())
    if len(text) > 50 and alnum / len(text) < 0.3:
        return True
    return False

def render_pages_to_images(uploaded_file, page_numbers, dpi=150):
    """Render specific PDF pages to base64 PNG images using PyMuPDF."""
    import fitz
    images = {}
    uploaded_file.seek(0)
    doc = fitz.open(stream=uploaded_file.read(), filetype="pdf")
    for pn in page_numbers:
        if pn < len(doc):
            pix = doc[pn].get_pixmap(matrix=fitz.Matrix(dpi/72, dpi/72))
            images[pn] = base64.b64encode(pix.tobytes("png")).decode("utf-8")
    doc.close()
    return images

def parse_td_statement(raw_text):
    """Legacy TD EasyWeb parser (balance-change method)."""
    lines = raw_text.split('\n')
    transactions = []
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        m = re.match(
            r'^((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2},\s+\d{4})\s+'
            r'(?:View (?:more|Cheque)\s+)?(.+?)\s+([\d,]+\.\d{2})\s+\$([\d,]+\.\d{2})', line)
        if m:
            desc = re.sub(r'\s+', ' ', re.sub(r'^(?:View (?:more|Cheque)\s*)', '', m.group(2)).strip())
            transactions.append({'date': m.group(1), 'description': desc,
                'amount': float(m.group(3).replace(',','')), 'balance': float(m.group(4).replace(',',''))})
            while i+1 < len(lines) and re.match(r'^\d{5,}$|^SUPP$|^BONU$', lines[i+1].strip()): i += 1
        i += 1
    # Determine debit/credit from balance changes
    for idx in range(len(transactions)):
        t = transactions[idx]
        prev_bal = transactions[idx+1]['balance'] if idx+1 < len(transactions) else t['balance']
        change = t['balance'] - prev_bal
        if abs(change + t['amount']) < 1.0:
            t['debit'], t['credit'] = t['amount'], 0
        elif abs(change - t['amount']) < 1.0:
            t['debit'], t['credit'] = 0, t['amount']
        else:
            desc_up = t['description'].upper()
            if any(kw in desc_up for kw in ['DEPOSIT','TFR-FR','E-TRANSFER ***','CHQ#','TD ATM DEP']):
                t['debit'], t['credit'] = 0, t['amount']
            else:
                t['debit'], t['credit'] = t['amount'], 0
    return transactions

def chunk_transactions(transactions, chunk_size=80):
    return [transactions[i:i+chunk_size] for i in range(0, len(transactions), chunk_size)]



def parse_csv_statement(uploaded_file):
    """Parse bank CSV exports. Returns (transactions, detected_bank).
    Handles TD, RBC, BMO, CIBC, Scotiabank CSV formats."""
    uploaded_file.seek(0)
    raw = uploaded_file.read()
    # Try UTF-8, then latin-1
    try: text = raw.decode("utf-8")
    except UnicodeDecodeError: text = raw.decode("latin-1")
    lines = text.strip().split("\n")
    if len(lines) < 2:
        return [], "Unknown"

    # Detect format from header or first few lines
    header = lines[0].lower()
    transactions = []
    bank = "Unknown"

    # TD: Date, Description, Debit, Credit, Balance (no header sometimes)
    # RBC: "Account Type","Account Number","Transaction Date","Cheque Number","Description 1","Description 2","CAD$","USD$"
    # BMO: "Following data is valid as of ..."; then "First Bank Card","Transaction Date","Transaction Amount","Description"
    # CIBC: "Date","Description","Debit","Credit"
    # Scotiabank: Date, Description, Withdrawal, Deposit, Balance

    if "account type" in header and "description 1" in header:
        # RBC format
        bank = "RBC"
        reader = csv.DictReader(lines)
        for row in reader:
            try:
                date = row.get("Transaction Date","").strip()
                desc = (row.get("Description 1","") + " " + row.get("Description 2","")).strip()
                amt_str = row.get("CAD$","0").strip().replace(",","")
                amt = float(amt_str) if amt_str else 0
                debit = abs(amt) if amt < 0 else 0
                credit = amt if amt > 0 else 0
                transactions.append({"date":date,"description":desc,"debit":debit,"credit":credit,"amount":amt,"balance":0})
            except Exception as e:
                logger.warning(f"RBC CSV row parse error: {e}")
                continue

    elif "first bank card" in header or "transaction amount" in header:
        # BMO format
        bank = "BMO"
        reader = csv.DictReader(lines)
        for row in reader:
            try:
                date = row.get("Transaction Date","").strip()
                desc = row.get("Description","").strip()
                amt = float(row.get("Transaction Amount","0").strip().replace(",",""))
                debit = amt if amt > 0 else 0
                credit = abs(amt) if amt < 0 else 0
                transactions.append({"date":date,"description":desc,"debit":debit,"credit":credit,"amount":amt,"balance":0})
            except Exception as e:
                logger.warning(f"BMO CSV row parse error: {e}")
                continue

    elif "debit" in header and "credit" in header:
        # CIBC or generic debit/credit format
        bank = "CIBC" if "cibc" in text.lower() else "Generic"
        reader = csv.DictReader(lines)
        for row in reader:
            try:
                date = row.get("Date","").strip()
                desc = row.get("Description","").strip()
                deb_str = row.get("Debit","0").strip().replace(",","").replace("$","")
                cred_str = row.get("Credit","0").strip().replace(",","").replace("$","")
                debit = float(deb_str) if deb_str and deb_str != "" else 0
                credit = float(cred_str) if cred_str and cred_str != "" else 0
                amt = debit if debit else -credit
                transactions.append({"date":date,"description":desc,"debit":debit,"credit":credit,"amount":amt,"balance":0})
            except Exception as e:
                logger.warning(f"CIBC CSV row parse error: {e}")
                continue

    elif "withdrawal" in header and "deposit" in header:
        # Scotiabank format
        bank = "Scotiabank"
        reader = csv.DictReader(lines)
        for row in reader:
            try:
                date = row.get("Date","").strip()
                desc = row.get("Description","").strip()
                w = row.get("Withdrawal","0").strip().replace(",","").replace("$","")
                d = row.get("Deposit","0").strip().replace(",","").replace("$","")
                debit = float(w) if w and w != "" else 0
                credit = float(d) if d and d != "" else 0
                transactions.append({"date":date,"description":desc,"debit":debit,"credit":credit,"amount":debit or -credit,"balance":0})
            except Exception as e:
                logger.warning(f"Scotiabank CSV row parse error: {e}")
                continue

    else:
        # TD or generic no-header: Date,Description,Debit,Credit,Balance
        bank = "TD"
        for line in lines:
            parts = line.split(",")
            if len(parts) >= 4:
                try:
                    date = parts[0].strip().strip('"')
                    desc = parts[1].strip().strip('"')
                    deb = parts[2].strip().strip('"').replace(",","")
                    cred = parts[3].strip().strip('"').replace(",","")
                    debit = float(deb) if deb else 0
                    credit = float(cred) if cred else 0
                    bal = float(parts[4].strip().strip('"').replace(",","")) if len(parts) > 4 and parts[4].strip() else 0
                    if date and any(c.isdigit() for c in date):
                        transactions.append({"date":date,"description":desc,"debit":debit,"credit":credit,"amount":debit or -credit,"balance":bal})
                except Exception as e:
                    logger.warning(f"TD CSV row parse error: {e}")
                    continue

    return transactions, bank

# ═══════════════════════════════════════════════════════════════════
# RECEIPT MATCHER — matches receipt photos to transactions
# ═══════════════════════════════════════════════════════════════════

def process_receipts(receipt_files_list, transactions, client, model_id, system_config, rate_in, rate_out, period_hint=""):
    """Send receipt images to Claude, extract vendor+amount+date, match to transactions.
    period_hint: statement period string (e.g. 'Dec 2024 - Jan 2025') used to infer
    year for year-less receipt dates — prevents cross-year receipts matching to wrong year.
    Returns (matched_receipts, total_cost)."""
    matched = []
    cost = 0

    for rf in receipt_files_list:
        rf.seek(0)
        data = rf.read()
        rf.seek(0)

        # Determine media type
        name = rf.name.lower()
        if name.endswith(".pdf"):
            try:
                import fitz
                doc = fitz.open(stream=data, filetype="pdf")
                pix = doc[0].get_pixmap(matrix=fitz.Matrix(150/72, 150/72))
                b64 = base64.b64encode(pix.tobytes("png")).decode("utf-8")
                doc.close()
                media = "image/png"
            except Exception as e:
                continue
        else:
            b64 = base64.b64encode(data).decode("utf-8")
            media = "image/jpeg" if "jpg" in name or "jpeg" in name else "image/png"

        content = [
            {"type":"image","source":{"type":"base64","media_type":media,"data":b64}},
            {"type":"text","text":"Extract from this receipt: Vendor\\tDate\\tTotal\\tHST\\tItems\nOne line only. No explanation. If unreadable, output: UNREADABLE\\t\\t\\t\\t"}
        ]
        try:
            resp = ""
            with client.messages.stream(model=model_id, max_tokens=200, system=system_config,
                messages=[{"role":"user","content":content}], timeout=60.0) as stream:
                for txt in stream.text_stream: resp += txt
            cost += (1600*rate_in + (len(resp)//4)*rate_out) / 1e6

            parts = resp.strip().split("\t")
            if len(parts) >= 3 and parts[0] != "UNREADABLE":
                vendor = parts[0].strip()
                r_date = parts[1].strip()
                try: total = float(re.sub(r'[,$]','',parts[2]))
                except (ValueError, TypeError): total = 0
                try: hst = float(re.sub(r'[,$]','',parts[3])) if len(parts) > 3 else 0
                except (ValueError, TypeError): hst = 0
                items = parts[4] if len(parts) > 4 else ""

                # Match to transaction: same amount ±2% or ±$0.50 (whichever is greater), date ±3 days
                best = None
                r_dt = _parse_date_fuzzy(r_date, period_hint)
                for t in transactions:
                    t_amt = t.get("debit",0) or t.get("credit",0)
                    tolerance = max(t_amt * 0.02, 0.50) if t_amt > 0 else 0.50
                    if abs(t_amt - total) <= tolerance:
                        t_dt = _parse_date_fuzzy(t.get("date",""), period_hint)
                        if r_dt and t_dt and abs((r_dt - t_dt).days) <= 3:
                            best = t
                            break
                        elif not r_dt or not t_dt:
                            best = t  # can't compare dates, match on amount

                matched.append({
                    "file": rf.name, "vendor": vendor, "date": r_date,
                    "total": total, "hst": hst, "items": items,
                    "matched_txn": best.get("description","") if best else "NO MATCH",
                    "matched_date": best.get("date","") if best else "",
                    "status": "MATCHED" if best else "UNMATCHED"
                })
                if best:
                    best["notes"] = (best.get("notes","") + f" RECEIPT:{rf.name}").strip()
        except Exception as e:
            matched.append({"file":rf.name,"vendor":"ERROR","date":"","total":0,"hst":0,"items":"","matched_txn":str(e),"matched_date":"","status":"ERROR"})

    return matched, cost

# ═══════════════════════════════════════════════════════════════════
# INVOICE PARSER — extract AP data from vendor invoices
# ═══════════════════════════════════════════════════════════════════

def process_invoices(invoice_files_list, client, model_id, system_config, rate_in, rate_out):
    """Send invoice images to Claude, extract vendor/amount/HST/due date.
    Returns (invoice_data, total_cost)."""
    invoices = []
    cost = 0

    for inv in invoice_files_list:
        inv.seek(0)
        data = inv.read()
        inv.seek(0)

        name = inv.name.lower()
        if name.endswith(".pdf"):
            try:
                import fitz
                doc = fitz.open(stream=data, filetype="pdf")
                pix = doc[0].get_pixmap(matrix=fitz.Matrix(150/72, 150/72))
                b64 = base64.b64encode(pix.tobytes("png")).decode("utf-8")
                doc.close()
                media = "image/png"
            except Exception as e:
                continue
        else:
            b64 = base64.b64encode(data).decode("utf-8")
            media = "image/jpeg" if "jpg" in name or "jpeg" in name else "image/png"

        content = [
            {"type":"image","source":{"type":"base64","media_type":media,"data":b64}},
            {"type":"text","text":"Extract from this invoice:\nVendor\\tInvoice#\\tDate\\tDue Date\\tSubtotal\\tHST/GST\\tTotal\\tCategory\nOne line. Category: Materials, Services, Rent, Utilities, Insurance, Other. No explanation."}
        ]
        try:
            resp = ""
            with client.messages.stream(model=model_id, max_tokens=200, system=system_config,
                messages=[{"role":"user","content":content}], timeout=60.0) as stream:
                for txt in stream.text_stream: resp += txt
            cost += (1600*rate_in + (len(resp)//4)*rate_out) / 1e6

            parts = resp.strip().split("\t")
            if len(parts) >= 5:
                try: subtotal = float(re.sub(r'[,$]','',parts[4]))
                except (ValueError, TypeError): subtotal = 0
                try: hst = float(re.sub(r'[,$]','',parts[5])) if len(parts) > 5 else 0
                except (ValueError, TypeError): hst = 0
                try: total = float(re.sub(r'[,$]','',parts[6])) if len(parts) > 6 else subtotal + hst
                except (ValueError, TypeError): total = subtotal + hst

                invoices.append({
                    "file": inv.name, "vendor": parts[0].strip(),
                    "invoice_num": parts[1].strip() if len(parts) > 1 else "",
                    "date": parts[2].strip() if len(parts) > 2 else "",
                    "due_date": parts[3].strip() if len(parts) > 3 else "",
                    "subtotal": subtotal, "hst": hst, "total": total,
                    "category": parts[7].strip() if len(parts) > 7 else "Other"
                })
        except Exception as e:
            invoices.append({"file":inv.name,"vendor":"ERROR","invoice_num":"","date":"","due_date":"","subtotal":0,"hst":0,"total":0,"category":str(e)})

    return invoices, cost

# ═══════════════════════════════════════════════════════════════════
# STATEMENT SUMMARY EXTRACTION — validates totals against PDF
# ═══════════════════════════════════════════════════════════════════

def extract_statement_summary(statement_file, client, model_id, rate_in, rate_out):
    """Extract summary totals from ALL summary pages of a multi-month statement.
    Scans every page for summary keywords and sums totals across all monthly summaries."""
    statement_file.seek(0)

    # Collect ALL pages that contain summary data
    summary_pages_text = []
    try:
        with pdfplumber.open(statement_file) as pdf:
            for i in range(len(pdf.pages)):
                t = pdf.pages[i].extract_text()
                if t and any(kw in t.lower() for kw in [
                    "total purchases", "total charges", "new balance",
                    "your account summary", "account summary", "balance from",
                    "total new charges", "total new activity"
                ]):
                    summary_pages_text.append(t)
    except Exception as e:
        logger.debug(f"PDF summary scan failed: {e}")

    if not summary_pages_text:
        # Fallback: first 3 pages
        statement_file.seek(0)
        try:
            with pdfplumber.open(statement_file) as pdf:
                for i in range(min(3, len(pdf.pages))):
                    t = pdf.pages[i].extract_text()
                    if t: summary_pages_text.append(t)
        except Exception as e:
            logger.debug(f"PDF fallback failed: {e}")

    if not summary_pages_text:
        # Vision fallback for CID pages
        content = []
        try:
            import fitz
            statement_file.seek(0)
            doc = fitz.open(stream=statement_file.read(), filetype="pdf")
            for i in range(min(3, len(doc))):
                pix = doc[i].get_pixmap(matrix=fitz.Matrix(150/72, 150/72))
                b64 = base64.b64encode(pix.tobytes("png")).decode("utf-8")
                content.append({"type":"image","source":{"type":"base64","media_type":"image/png","data":b64}})
            doc.close()
        except Exception as e:
            logger.debug(f"Vision fallback failed: {e}")
            return {}, 0
    else:
        combined = "\n---PAGE BREAK---\n".join(summary_pages_text)
        content = [{"type":"text","text":f"STATEMENT SUMMARY PAGES ({len(summary_pages_text)} pages found):\n{combined[:8000]}"}]

    content.append({"type":"text","text":"""This credit card statement may contain MULTIPLE monthly summaries.
Find ALL monthly summary sections and ADD UP the totals across all months.

Return EXACTLY this format (TSV, one line per field):
MONTHS_FOUND\t<number of monthly summaries found>
OPENING_BALANCE\t<first month's opening balance>
CLOSING_BALANCE\t<last month's closing balance>
TOTAL_PURCHASES\t<SUM of all months' purchases>
TOTAL_PAYMENTS\t<SUM of all months' payments (absolute value)>
TOTAL_INTEREST\t<SUM of all months' interest charges>
TOTAL_FEES\t<SUM of all months' fees>
TOTAL_CREDITS\t<SUM of all months' returns/credits (absolute value)>
CREDIT_LIMIT\t<credit limit if shown>
STATEMENT_PERIOD\t<earliest date> - <latest date>

CRITICAL: If there are 12 monthly summaries, you must SUM all 12 purchase totals.
If a field is not found, use 0 for amounts or UNKNOWN for text.
Output ONLY the TSV lines above. No explanation."""})

    try:
        resp = ""
        with client.messages.stream(model=model_id, max_tokens=500,
            messages=[{"role":"user","content":content}], timeout=90.0) as stream:
            for txt in stream.text_stream: resp += txt

        cost = (max(len(summary_pages_text), 1) * 800 * rate_in + (len(resp)//4) * rate_out) / 1e6
        summary = {}
        for ln in resp.strip().split("\n"):
            parts = ln.strip().split("\t")
            if len(parts) >= 2:
                key = parts[0].strip().upper()
                val = parts[1].strip()
                try:
                    summary[key] = float(re.sub(r'[,$]', '', val))
                except Exception:
                    summary[key] = val
        return summary, cost
    except Exception as e:
        logger.debug(f"Statement summary extraction failed: {e}")
        return {}, 0

def calc_itc(debit, itc_rule, province_code="ON"):
    """
    Calculate ITC using the correct provincial tax rate.
    province_code must be a 2-letter code (ON, AB, QC, etc.).
    Uses get_itc_rate_fraction() from bank_schemas — never hardcodes 13%.
    """
    if not debit or debit <= 0:
        return 0
    code = normalize_province(province_code)   # handles full names safely
    rate = get_itc_rate_fraction(code)          # e.g. ON→0.11504, AB→0.04762
    if "Full" in str(itc_rule):
        return round(debit * rate, 2)
    if "50%" in str(itc_rule) or "50" in str(itc_rule):
        return round(debit * rate * 0.5, 2)
    return 0

def parse_response_line(p, bank_src, province_code="ON"):
    """
    Parse a TSV response line into a transaction dict.
    Handles both old 9-col format (no Type) and new 10-col format (with Type).
    province_code: 2-letter code used for correct ITC calculation.

    TSV validation: lines with fewer than 4 real columns are silently skipped
    by the caller — this function assumes p has already been validated.
    """
    # Detect format: if p[5] is a Type keyword → new 10-col format
    has_type = len(p) > 5 and p[5].strip().upper() in (
        "PURCHASE", "REFUND", "PAYMENT", "FEE", "INTEREST",
        "FEE_REBATE", "CASH_ADVANCE", "OTHER", ""
    )
    if has_type:
        # New format: Date|Source|Description|Debit|Credit|Type|Category|ITC|Confidence|Notes
        try: debit = float(re.sub(r'[,$]', '', p[3])) if p[3].strip() not in ("", "0", "—", "-") else 0
        except (ValueError, TypeError): debit = 0
        try: credit = float(re.sub(r'[,$]', '', p[4])) if p[4].strip() not in ("", "0", "—", "-") else 0
        except (ValueError, TypeError): credit = 0
        txn_type = p[5].strip().upper() if len(p) > 5 else ""
        cat = p[6].strip() if len(p) > 6 else "❓ Uncategorized"
        # ITC: use AI value if present and non-zero, otherwise compute with correct province rate
        try:
            itc_amt = float(re.sub(r'[,$]', '', p[7])) if len(p) > 7 and p[7].strip() not in ("", "—", "-") else 0
        except Exception as e:
            itc_amt = 0
        if itc_amt == 0 and debit > 0 and cat and "Uncategorized" not in cat:
            itc_rule = "50%" if "MEALS" in (p[9] if len(p) > 9 else "").upper() else "Full"
            itc_amt = calc_itc(debit, itc_rule, province_code)
        conf = p[8].strip() if len(p) > 8 else "70"
        notes = p[9].strip() if len(p) > 9 else ""
        itc_rule = "50%" if "MEALS" in notes.upper() else ("Full" if itc_amt > 0 else "No")
        source = p[1].strip() if p[1].strip() else bank_src
        description = p[2].strip() if len(p) > 2 else ""
    else:
        # Old format: Date|Description|Debit|Credit|Category|T2125|ITC|Conf|Notes
        try: debit = float(re.sub(r'[,$]', '', p[2])) if len(p) > 2 and p[2].strip() not in ("", "0", "—", "-") else 0
        except (ValueError, TypeError): debit = 0
        try: credit = float(re.sub(r'[,$]', '', p[3])) if len(p) > 3 and p[3].strip() not in ("", "0", "—", "-") else 0
        except (ValueError, TypeError): credit = 0
        txn_type = "PURCHASE" if debit > 0 else "PAYMENT" if credit > 0 else ""
        cat = p[4].strip() if len(p) > 4 else "❓ Uncategorized"
        itc_rule = p[6].strip() if len(p) > 6 else ""
        itc_amt = calc_itc(debit, itc_rule, province_code)
        conf = p[7].strip() if len(p) > 7 else "🟡"
        notes = p[8].strip() if len(p) > 8 else ""
        source = bank_src
        description = p[1].strip() if len(p) > 1 else ""

    return {
        "date": p[0].strip() if p else "",
        "source": source,
        "description": description,
        "debit": debit,
        "credit": credit,
        "balance": 0,
        "type": txn_type,
        "category": cat,
        "t2125": T2125_LINE_MAP.get(cat, ""),
        "itc_rule": itc_rule,
        "itc_amount": itc_amt,
        "confidence": conf,
        "notes": notes,
    }

# ═══════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# QUICKBOOKS / XERO CSV EXPORT (v3.7)
# Generates import-ready CSV for both platforms.
# QuickBooks Online IIF format uses their standard column headers.
# Xero uses their standard import format.
# ═══════════════════════════════════════════════════════════════════

def build_quickbooks_csv(transactions, biz_name, period_str):
    """
    Export transactions in QuickBooks Online CSV import format.
    Columns: Date, Description, Amount, Account, Tax Code, Memo
    Negatives = expenses, positives = income/credits.
    """
    import io, csv
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Date", "Description", "Amount", "Account", "Tax Code", "Memo"])

    # QBO account name mapping from T2125 categories
    ACCOUNT_MAP = {
        "Motor Vehicle Expense": "Vehicle Expenses",
        "Meals & Entertainment": "Meals and Entertainment",
        "Office Supplies": "Office Expenses",
        "Utilities": "Utilities",
        "Bank Charges": "Bank Service Charges",
        "Insurance": "Insurance Expense",
        "Materials & Supplies": "Job Materials",
        "Rent": "Rent or Lease",
        "Advertising": "Advertising",
        "Travel": "Travel",
        "Professional Fees": "Legal & Professional Fees",
        "Subcontracts": "Subcontractors",
        "Cost of Goods": "Cost of Goods Sold",
        "Repairs & Maintenance": "Repairs & Maintenance",
        "Owner Draw / Personal": "Owner\'s Draw",
        "Shareholder Loan (Debit)": "Shareholder Loans",
    }

    for t in transactions:
        date = t.get("date", "")
        desc = t.get("description", "")
        debit = t.get("debit", 0)
        credit = t.get("credit", 0)
        cat = t.get("category", "Uncategorized")
        notes = t.get("notes", "")

        # QBO: expenses are negative, credits/payments are positive
        if debit > 0:
            amount = -round(debit, 2)
        elif credit > 0:
            amount = round(credit, 2)
        else:
            continue  # skip zero rows

        account = ACCOUNT_MAP.get(cat, cat or "Uncategorized Expense")

        # Tax code: GST/HST if ITC was claimed
        itc = t.get("itc_amount", 0)
        tax_code = "HST/GST" if itc and itc > 0 else "Exempt"

        writer.writerow([date, desc, amount, account, tax_code, notes])

    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


def build_xero_csv(transactions, biz_name, period_str):
    """
    Export transactions in Xero bank statement import format.
    Xero columns: Date, Amount, Payee, Description, Reference, Check Number, Currency
    Xero convention: negative = money out (debit), positive = money in (credit).
    """
    import io, csv
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Date", "Amount", "Payee", "Description", "Reference", "Check Number", "Currency"])

    for t in transactions:
        date = t.get("date", "")
        debit = t.get("debit", 0)
        credit = t.get("credit", 0)
        desc = t.get("description", "")
        cat = t.get("category", "")
        notes = t.get("notes", "")
        source = t.get("source", "")

        if debit > 0:
            amount = -round(debit, 2)  # Xero: outflow is negative
        elif credit > 0:
            amount = round(credit, 2)  # Xero: inflow is positive
        else:
            continue

        # Xero: Payee = merchant, Description = category + notes, Reference = source
        payee = desc
        description = f"{cat} — {notes}".strip(" —") if notes else cat
        reference = source

        writer.writerow([date, amount, payee, description, reference, "", "CAD"])

    return buf.getvalue().encode("utf-8-sig")


def build_excel(transactions, flags, summary, biz_name, industry_str, province_str, period_str, recon_matches=None, recon_unmatched=None, receipt_matches=None, invoice_data=None, validation_results=None, t5018_data=None, validation_report=None, anomalies=None, audit_trail=None):
    wb = Workbook()
    hfill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    dfont = Font(name="Arial", size=10)
    bdr = Border(*(Side(style="thin") for _ in range(4)))
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    def write_header(ws, headers, widths):
        for i, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(1, i, h)
            c.fill, c.font, c.border, c.alignment = hfill, hfont, bdr, Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        ws.freeze_panes = "A2"

    # Tab 1: All Transactions
    ws = wb.active; ws.title = "All Transactions"
    hdrs = ["Date","Source","Description","Debit","Credit","Type","Category","T2125 Line","Biz %","ITC Rule","ITC Amount","Confidence","Notes"]
    write_header(ws, hdrs, [12,14,42,12,12,10,28,10,7,9,11,10,38])
    for r, t in enumerate(transactions, 2):
        ws.cell(r,1,t["date"]).font=dfont; ws.cell(r,2,t["source"]).font=dfont; ws.cell(r,3,t["description"]).font=dfont
        if t["debit"]: c=ws.cell(r,4,t["debit"]); c.number_format='$#,##0.00'; c.font=dfont
        if t["credit"]: c=ws.cell(r,5,t["credit"]); c.number_format='$#,##0.00'; c.font=dfont
        ws.cell(r,6,t.get("type","")).font=dfont
        ws.cell(r,7,t["category"]).font=dfont
        ws.cell(r,8,t.get("t2125","")).font=dfont
        ws.cell(r,9,t.get("biz_pct",100)).font=dfont
        ws.cell(r,10,t.get("itc_rule","")).font=dfont
        if t.get("itc_amount"): c=ws.cell(r,11,t["itc_amount"]); c.number_format='$#,##0.00'; c.font=dfont
        ws.cell(r,12,t.get("confidence","")).font=dfont; ws.cell(r,13,t.get("notes","")).font=dfont
        conf=t.get("confidence","")
        fill = green_fill if str(conf).isdigit() and int(conf)>=85 else red_fill if str(conf).isdigit() and int(conf)<70 else yellow_fill if str(conf).isdigit() and int(conf)<85 else alt_fill if r%2==0 else None
        if fill:
            for col in range(1,14): ws.cell(r,col).fill=fill
        for col in range(1,14): ws.cell(r,col).border=bdr

    # Tab 2: Expense Summary (FORMULA-DRIVEN — updates when accountant edits categories)
    ws2 = wb.create_sheet("Expense Summary")
    ws2.cell(1,1,"📊 Expense Summary — LIVE from All Transactions").font=Font(bold=True,size=12,name="Arial")
    ws2.cell(2,1,"✏️ Edit categories in 'All Transactions' column G — this sheet updates automatically.").font=Font(italic=True,color="006600",name="Arial")
    sum_hdrs = ["Category","T2125 Line","Count","Total","Biz %","Deductible","% of Expenses","ITC Claimable"]
    for i,(h,w) in enumerate(zip(sum_hdrs,[28,10,8,14,7,14,12,14]),1):
        c=ws2.cell(4,i,h); c.fill,c.font,c.border,c.alignment=hfill,hfont,bdr,Alignment(horizontal="center")
        ws2.column_dimensions[get_column_letter(i)].width=w
    ws2.auto_filter.ref = f"A4:H4"
    ws2.freeze_panes = "A5"

    # Get unique categories from transactions
    cats_unique = sorted(set(t.get("category","") for t in transactions if t.get("debit",0) > 0 and t.get("category","")))
    last_data_row = len(transactions) + 1
    txn_sheet = "'All Transactions'"

    # Also build cats dict for HST tab (needed for biz_pct lookup)
    cats = {}
    for t in transactions:
        if t["debit"] > 0:
            cat = t["category"]
            if cat not in cats:
                cats[cat] = {"count":0,"total":0,"itc":0,"t2125":t.get("t2125",""),"biz_pct":t.get("biz_pct",100)}
            cats[cat]["count"] += 1; cats[cat]["total"] += t["debit"]; cats[cat]["itc"] += t.get("itc_amount",0)
    total_exp = sum(c["total"] for c in cats.values())

    for r, cat in enumerate(cats_unique, 5):
        ws2.cell(r,1,cat).font=dfont
        ws2.cell(r,2,T2125_MAP.get(cat,"")).font=dfont
        ws2.cell(r,3).value = f'=COUNTIFS({txn_sheet}!G2:G{last_data_row},A{r},{txn_sheet}!D2:D{last_data_row},">0")'
        ws2.cell(r,3).font = dfont
        ws2.cell(r,4).value = f'=SUMIFS({txn_sheet}!D2:D{last_data_row},{txn_sheet}!G2:G{last_data_row},A{r})'
        ws2.cell(r,4).number_format = '$#,##0.00'; ws2.cell(r,4).font = dfont
        biz = BIZ_USE.get(cat, 100 if "Personal" not in cat else 0)
        ws2.cell(r,5,biz).font=dfont
        ws2.cell(r,6).value = f'=D{r}*E{r}/100'
        ws2.cell(r,6).number_format = '$#,##0.00'; ws2.cell(r,6).font = dfont
        total_row = 5 + len(cats_unique)
        ws2.cell(r,7).value = f'=IF(D${total_row}=0,0,D{r}/D${total_row}*100)'
        ws2.cell(r,7).number_format = '0.0'; ws2.cell(r,7).font = dfont
        ws2.cell(r,8).value = f'=SUMIFS({txn_sheet}!K2:K{last_data_row},{txn_sheet}!G2:G{last_data_row},A{r})'
        ws2.cell(r,8).number_format = '$#,##0.00'; ws2.cell(r,8).font = dfont
        for col in range(1,9): ws2.cell(r,col).border=bdr

    tr = 5 + len(cats_unique)
    ws2.cell(tr,1,"TOTAL").font=Font(bold=True,name="Arial")
    ws2.cell(tr,3).value = f'=SUM(C5:C{tr-1})'; ws2.cell(tr,3).font=Font(bold=True,name="Arial")
    ws2.cell(tr,4).value = f'=SUM(D5:D{tr-1})'; ws2.cell(tr,4).number_format='$#,##0.00'; ws2.cell(tr,4).font=Font(bold=True,name="Arial")
    ws2.cell(tr,6).value = f'=SUM(F5:F{tr-1})'; ws2.cell(tr,6).number_format='$#,##0.00'; ws2.cell(tr,6).font=Font(bold=True,name="Arial")
    ws2.cell(tr,8).value = f'=SUM(H5:H{tr-1})'; ws2.cell(tr,8).number_format='$#,##0.00'; ws2.cell(tr,8).font=Font(bold=True,name="Arial")
    for col in range(1,9): ws2.cell(tr,col).border=bdr

    # Tab 3: HST/ITC Summary (FORMULA-DRIVEN)
    ws3 = wb.create_sheet("HST-GST ITC")
    ws3.cell(1,1,"HST/GST Input Tax Credit Summary").font=Font(bold=True,size=14,name="Arial")
    ws3.cell(2,1,f"Business: {biz_name} | Period: {period_str}").font=dfont
    ws3.cell(3,1,"✏️ Updates automatically when you edit categories in 'All Transactions'.").font=Font(italic=True,color="006600",name="Arial")
    for i,(h,w) in enumerate(zip(["Category","T2125 Line","Expense Total","Biz %","Deductible","ITC Claimable"],[28,10,14,7,14,14]),1):
        c=ws3.cell(5,i,h); c.fill,c.font,c.border=hfill,hfont,bdr; ws3.column_dimensions[get_column_letter(i)].width=w
    total_itc=0
    for r, cat in enumerate(cats_unique, 6):
        ws3.cell(r,1,cat).font=dfont
        ws3.cell(r,2,T2125_MAP.get(cat,"")).font=dfont
        ws3.cell(r,3).value = f'=SUMIFS({txn_sheet}!D2:D{last_data_row},{txn_sheet}!G2:G{last_data_row},A{r})'
        ws3.cell(r,3).number_format='$#,##0.00'; ws3.cell(r,3).font=dfont
        biz = BIZ_USE.get(cat, 100 if "Personal" not in cat else 0)
        ws3.cell(r,4,biz).font=dfont
        ws3.cell(r,5).value = f'=C{r}*D{r}/100'
        ws3.cell(r,5).number_format='$#,##0.00'; ws3.cell(r,5).font=dfont
        ws3.cell(r,6).value = f'=SUMIFS({txn_sheet}!K2:K{last_data_row},{txn_sheet}!G2:G{last_data_row},A{r})'
        ws3.cell(r,6).number_format='$#,##0.00'; ws3.cell(r,6).font=dfont
        for col in range(1,7): ws3.cell(r,col).border=bdr
    tr=6+len(cats_unique)
    ws3.cell(tr,1,"TOTAL ITC CLAIMABLE").font=Font(bold=True,size=11,name="Arial")
    ws3.cell(tr,6).value = f'=SUM(F6:F{tr-1})'
    ws3.cell(tr,6).number_format='$#,##0.00'; ws3.cell(tr,6).font=Font(bold=True,size=12,name="Arial")

    # Tab 4: Needs Review
    ws4 = wb.create_sheet("⚠️ Needs Review")
    review=[t for t in transactions if
        ("Uncategorized" in t.get("category","")) or
        (str(t.get("confidence","0")).isdigit() and int(t.get("confidence","0"))<70) or
        ("FEE_REBATE_PAIR" in t.get("notes",""))]
    write_header(ws4, ["#","Date","Description","Amount","Type","Current Category","Flag/Notes"], [5,13,40,12,10,28,50])
    for r,t in enumerate(review,2):
        ws4.cell(r,1,r-1).font=dfont; ws4.cell(r,2,t["date"]).font=dfont; ws4.cell(r,3,t["description"]).font=dfont
        c=ws4.cell(r,4,t["debit"] or t["credit"]); c.number_format='$#,##0.00'; c.font=dfont
        ws4.cell(r,5,t.get("type","")).font=dfont
        ws4.cell(r,6,t["category"]).font=dfont; ws4.cell(r,7,t.get("notes","")).font=dfont
        for col in range(1,8): ws4.cell(r,col).border=bdr; ws4.cell(r,col).fill=yellow_fill

    # Tab 5: Fixed Assets (CCA) — v3.3: now with proper class assignment
    ws_cca = wb.create_sheet("Fixed Assets CCA")
    cca_items = [t for t in transactions if "CCA_ASSET" in t.get("notes","")]
    ws_cca.cell(1,1,"Potential Capital Assets Review (CCA) — Purchases ≥ $500").font=Font(bold=True,size=14,name="Arial")
    ws_cca.cell(2,1,"Items flagged by AI as potential capital assets (≥ $500). CPA to confirm: claim via CCA over multiple years, NOT as a direct expense. Sub-$500 small tools may qualify for Class 12 (100% deduction).").font=Font(italic=True,color="666666",name="Arial")
    cca_hdrs = ["Date","Description","Amount","Category","CCA Class","Rate","Notes"]
    for i,(h,w) in enumerate(zip(cca_hdrs,[13,42,14,28,22,8,35]),1):
        c=ws_cca.cell(4,i,h); c.fill,c.font,c.border=hfill,hfont,bdr; ws_cca.column_dimensions[get_column_letter(i)].width=w
    for r,t in enumerate(cca_items,5):
        ws_cca.cell(r,1,t["date"]).font=dfont; ws_cca.cell(r,2,t["description"]).font=dfont
        c=ws_cca.cell(r,3,t["debit"]); c.number_format='$#,##0.00'; c.font=dfont
        ws_cca.cell(r,4,t.get("category","")).font=dfont
        # v3.3: Use pre-computed CCA class from notes, or compute it
        cca_label = t.get("cca_label","")
        if not cca_label:
            from bank_schemas import get_cca_class
            _, cca_label = get_cca_class(t.get("description",""), t.get("debit",0))
        # Parse class and rate from label (e.g. "Class 8 (20%) — Equipment")
        class_match = re.match(r'Class (\d+|50) \((\d+%)\)', cca_label)
        cca_class_num = class_match.group(1) if class_match else "8"
        cca_rate = class_match.group(2) if class_match else "20%"
        ws_cca.cell(r,5,cca_label).font=dfont
        ws_cca.cell(r,6,cca_rate).font=dfont
        ws_cca.cell(r,7,t.get("notes","")).font=dfont
        fill_cca = green_fill if cca_class_num == "12" else yellow_fill if cca_class_num == "50" else None
        for col in range(1,8):
            ws_cca.cell(r,col).border=bdr
            if fill_cca: ws_cca.cell(r,col).fill=fill_cca
    if cca_items:
        tr=5+len(cca_items)
        ws_cca.cell(tr,1,"TOTAL CAPITAL PURCHASES").font=Font(bold=True,size=11,name="Arial")
        c=ws_cca.cell(tr,3,sum(t["debit"] for t in cca_items)); c.number_format='$#,##0.00'; c.font=Font(bold=True,size=12,name="Arial")

    # Tab 5b: Reconciliation (if multi-doc)
    # Reconciliation data passed as parameters
    if recon_matches or recon_unmatched:
        ws_rec = wb.create_sheet("🔄 Reconciliation")
        ws_rec.cell(1,1,"Full Cross-Document Reconciliation").font=Font(bold=True,size=14,name="Arial")
        ws_rec.cell(2,1,"Matches: CC↔Chequing, Invoice↔Statement, Receipt↔Transaction, Duplicates").font=Font(italic=True,color="666666",name="Arial")

        if recon_matches:
            rec_hdrs=["Match Type","Doc A","Date A","Amount A","Source A","Doc B","Date B","Amount B","Source B","Days"]
            for i,(h,w) in enumerate(zip(rec_hdrs,[22,35,12,12,16,35,12,12,16,6]),1):
                c=ws_rec.cell(4,i,h); c.fill,c.font,c.border=hfill,hfont,bdr; ws_rec.column_dimensions[get_column_letter(i)].width=w
            for r,m in enumerate(recon_matches,5):
                ws_rec.cell(r,1,m.get("match_type","")).font=Font(bold=True,name="Arial")
                ws_rec.cell(r,2,m.get("doc_a","")).font=dfont
                ws_rec.cell(r,3,m.get("date_a","")).font=dfont
                c=ws_rec.cell(r,4,m.get("amount_a",0)); c.number_format='$#,##0.00'; c.font=dfont
                ws_rec.cell(r,5,m.get("source_a","")).font=dfont
                ws_rec.cell(r,6,m.get("doc_b","")).font=dfont
                ws_rec.cell(r,7,m.get("date_b","")).font=dfont
                c=ws_rec.cell(r,8,m.get("amount_b",0)); c.number_format='$#,##0.00'; c.font=dfont
                ws_rec.cell(r,9,m.get("source_b","")).font=dfont
                ws_rec.cell(r,10,m.get("day_diff",0)).font=dfont
                for col in range(1,11): ws_rec.cell(r,col).border=bdr; ws_rec.cell(r,col).fill=green_fill

        if recon_unmatched:
            ur = (5 + len(recon_matches) + 2) if recon_matches else 4
            ws_rec.cell(ur,1,"UNMATCHED ITEMS:").font=Font(bold=True,size=11,color="CC0000",name="Arial")
            un_hdrs=["Type","Date","Source","Description","Amount","Status"]
            for i,(h,w) in enumerate(zip(un_hdrs,[18,12,16,40,12,12]),1):
                c=ws_rec.cell(ur+1,i,h); c.fill,c.font,c.border=hfill,hfont,bdr
            for r,u in enumerate(recon_unmatched, ur+2):
                ws_rec.cell(r,1,u.get("type","")).font=dfont; ws_rec.cell(r,2,u.get("date","")).font=dfont
                ws_rec.cell(r,3,u.get("source","")).font=dfont; ws_rec.cell(r,4,u.get("desc","")).font=dfont
                c=ws_rec.cell(r,5,u.get("amount",0)); c.number_format='$#,##0.00'; c.font=dfont
                ws_rec.cell(r,6,u.get("status","")).font=dfont
                for col in range(1,7): ws_rec.cell(r,col).border=bdr; ws_rec.cell(r,col).fill=red_fill

    # Tab: Receipt Audit Trail
    if receipt_matches:
        ws_rct = wb.create_sheet("📸 Receipts")
        ws_rct.cell(1,1,"Receipt Audit Trail").font=Font(bold=True,size=14,name="Arial")
        rct_hdrs=["File","Vendor","Date","Total","HST","Items","Matched Transaction","Match Date","Status"]
        for i,(h,w) in enumerate(zip(rct_hdrs,[20,25,12,12,10,30,35,12,10]),1):
            c=ws_rct.cell(3,i,h); c.fill,c.font,c.border=hfill,hfont,bdr; ws_rct.column_dimensions[get_column_letter(i)].width=w
        for r,m in enumerate(receipt_matches,4):
            ws_rct.cell(r,1,m["file"]).font=dfont; ws_rct.cell(r,2,m["vendor"]).font=dfont
            ws_rct.cell(r,3,m["date"]).font=dfont
            if m["total"]: c=ws_rct.cell(r,4,m["total"]); c.number_format='$#,##0.00'; c.font=dfont
            if m["hst"]: c=ws_rct.cell(r,5,m["hst"]); c.number_format='$#,##0.00'; c.font=dfont
            ws_rct.cell(r,6,m.get("items","")).font=dfont
            ws_rct.cell(r,7,m["matched_txn"]).font=dfont; ws_rct.cell(r,8,m["matched_date"]).font=dfont
            ws_rct.cell(r,9,m["status"]).font=Font(bold=True,color="006600" if m["status"]=="MATCHED" else "CC0000",name="Arial")
            fill_r = green_fill if m["status"]=="MATCHED" else red_fill
            for col in range(1,10): ws_rct.cell(r,col).border=bdr; ws_rct.cell(r,col).fill=fill_r

    # Tab: Accounts Payable (Invoices)
    if invoice_data:
        ws_inv = wb.create_sheet("📋 Accounts Payable")
        ws_inv.cell(1,1,"Accounts Payable — Vendor Invoices").font=Font(bold=True,size=14,name="Arial")
        inv_hdrs=["File","Vendor","Invoice #","Date","Due Date","Subtotal","HST/GST","Total","Category"]
        for i,(h,w) in enumerate(zip(inv_hdrs,[20,25,14,12,12,12,12,12,18]),1):
            c=ws_inv.cell(3,i,h); c.fill,c.font,c.border=hfill,hfont,bdr; ws_inv.column_dimensions[get_column_letter(i)].width=w
        for r,inv in enumerate(invoice_data,4):
            ws_inv.cell(r,1,inv["file"]).font=dfont; ws_inv.cell(r,2,inv["vendor"]).font=dfont
            ws_inv.cell(r,3,inv.get("invoice_num","")).font=dfont
            ws_inv.cell(r,4,inv.get("date","")).font=dfont; ws_inv.cell(r,5,inv.get("due_date","")).font=dfont
            if inv["subtotal"]: c=ws_inv.cell(r,6,inv["subtotal"]); c.number_format='$#,##0.00'; c.font=dfont
            if inv["hst"]: c=ws_inv.cell(r,7,inv["hst"]); c.number_format='$#,##0.00'; c.font=dfont
            if inv["total"]: c=ws_inv.cell(r,8,inv["total"]); c.number_format='$#,##0.00'; c.font=dfont
            ws_inv.cell(r,9,inv.get("category","")).font=dfont
            for col in range(1,10): ws_inv.cell(r,col).border=bdr
        # AP totals
        tr = 4 + len(invoice_data)
        ws_inv.cell(tr,1,"TOTAL PAYABLE").font=Font(bold=True,size=11,name="Arial")
        c=ws_inv.cell(tr,8,sum(i["total"] for i in invoice_data)); c.number_format='$#,##0.00'; c.font=Font(bold=True,size=12,name="Arial")

    # Tab: Statement Validation — always created, even when pdf_summary extraction failed
    # This sheet shows both PDF-sourced totals and ledger-computed totals side by side.
    # When PDF extraction fails, shows computed-only rows with ⚠️ PDF N/A status.
    ws_val = wb.create_sheet("✅ Statement Validation")
    ws_val.cell(1,1,"Statement Total Validation").font=Font(bold=True,size=16,name="Arial")
    ws_val.cell(2,1,"PDF statement totals vs computed ledger totals. GREEN = match. RED = mismatch — investigate before filing.").font=Font(italic=True,color="666666",name="Arial")
    ws_val.cell(3,1,"⚠️ If PDF Value shows $0.00, the summary page could not be extracted — verify computed values against your statement manually.").font=Font(italic=True,color="CC6600",name="Arial")

    if validation_results:
        row = 5
        for vr in validation_results:
            ws_val.cell(row,1,f"📄 {vr['file']}").font=Font(bold=True,size=12,color="1F3864",name="Arial")
            row += 1
            val_hdrs = ["Field","PDF Value","Computed Value","Variance","Status"]
            for i,(h,w) in enumerate(zip(val_hdrs,[42,16,16,14,18]),1):
                c=ws_val.cell(row,i,h); c.fill,c.font,c.border=hfill,hfont,bdr
                ws_val.column_dimensions[get_column_letter(i)].width=w
            row += 1
            for v in vr.get("validations",[]):
                ws_val.cell(row,1,v["field"]).font=Font(bold=True,name="Arial")
                # PDF value: show 0.00 with gray color when unavailable
                c=ws_val.cell(row,2,v["pdf"] if v["pdf"] else 0)
                c.number_format='$#,##0.00'
                c.font=Font(name="Arial",color="999999" if not v["pdf"] else "000000")
                c=ws_val.cell(row,3,v["computed"]); c.number_format='$#,##0.00'; c.font=dfont
                c=ws_val.cell(row,4,v["variance"] if v.get("variance") else 0); c.number_format='$#,##0.00'; c.font=dfont
                status = v["status"]
                if "MATCH" in status or "BALANCED" in status:
                    status_color = "006600"; fill_v = green_fill
                elif "CLOSE" in status:
                    status_color = "CC6600"; fill_v = yellow_fill
                elif "PDF N/A" in status:
                    status_color = "666666"; fill_v = yellow_fill
                else:
                    status_color = "CC0000"; fill_v = red_fill
                ws_val.cell(row,5,status).font=Font(bold=True,color=status_color,name="Arial")
                for col in range(1,6):
                    ws_val.cell(row,col).border=bdr
                    ws_val.cell(row,col).fill=fill_v
                row += 1
            row += 1
    else:
        ws_val.cell(5,1,"No statements processed or no PDF files uploaded.").font=Font(italic=True,color="999999",name="Arial")

    # Tab: T5018 Subcontractor Summary (v3.3)
    if t5018_data:
        ws_t5 = wb.create_sheet("📋 T5018 Subcontractors")
        ws_t5.cell(1,1,"T5018 Subcontractor Payment Summary").font=Font(bold=True,size=14,name="Arial")
        ws_t5.cell(2,1,"CRA T5018 required for any subcontractor paid > $500 in the year (Construction/Trades)").font=Font(italic=True,color="666666",name="Arial")
        ws_t5.cell(3,1,"⚠️ Aggregate across ALL statements for the full fiscal year before filing.").font=Font(italic=True,color="CC6600",name="Arial")
        t5_hdrs = ["Payee Name","# Payments","Total Paid","T5018 Required?","Notes"]
        for i,(h,w) in enumerate(zip(t5_hdrs,[30,12,16,16,40]),1):
            c=ws_t5.cell(5,i,h); c.fill,c.font,c.border=hfill,hfont,bdr; ws_t5.column_dimensions[get_column_letter(i)].width=w
        for r_idx,sub in enumerate(t5018_data,6):
            ws_t5.cell(r_idx,1,sub["payee"]).font=Font(bold=True,name="Arial")
            ws_t5.cell(r_idx,2,sub["count"]).font=dfont
            c=ws_t5.cell(r_idx,3,sub["total"]); c.number_format='$#,##0.00'; c.font=Font(bold=True,name="Arial")
            req = "🔴 YES — File T5018" if sub.get("t5018_required") else "✅ Under threshold"
            ws_t5.cell(r_idx,4,req).font=Font(bold=True,color="CC0000" if sub.get("t5018_required") else "006600",name="Arial")
            ws_t5.cell(r_idx,5,"Verify: is this person incorporated? If yes, T5018 not required.").font=dfont
            fill_sub = red_fill if sub.get("t5018_required") else green_fill
            for col in range(1,6): ws_t5.cell(r_idx,col).border=bdr; ws_t5.cell(r_idx,col).fill=fill_sub
        # Totals
        total_r = 6 + len(t5018_data)
        ws_t5.cell(total_r,1,"TOTAL SUBCONTRACTOR PAYMENTS").font=Font(bold=True,size=11,name="Arial")
        c=ws_t5.cell(total_r,3,sum(s["total"] for s in t5018_data)); c.number_format='$#,##0.00'; c.font=Font(bold=True,size=12,name="Arial")

    # Tab: Audit Trail (v3.5/v3.11) — Claude raw vs post-processing changes, split by type
    if audit_trail:
        ws_at = wb.create_sheet("🔍 Audit Trail")
        ws_at.cell(1,1,"Post-Processing Audit Trail").font=Font(bold=True,size=14,name="Arial")
        ws_at.cell(2,1,"FINANCIAL rows affect category/ITC — review before filing. COSMETIC rows are merchant normalization only.").font=Font(italic=True,color="666666",name="Arial")
        ws_at.cell(3,1,"Yellow = category changed | Green = cosmetic only | Review yellow rows before finalising.").font=Font(italic=True,color="CC6600",name="Arial")
        at_hdrs = ["Type","Date","Description","Debit","AI Category","Final Category","AI ITC","Final ITC","AI Conf","Change Reason"]
        at_widths = [12,13,40,12,28,28,12,12,10,55]
        for i,(h,w) in enumerate(zip(at_hdrs,at_widths),1):
            c=ws_at.cell(5,i,h); c.fill,c.font,c.border=hfill,hfont,bdr
            ws_at.column_dimensions[get_column_letter(i)].width=w
        # Sort: FINANCIAL first, then COSMETIC
        sorted_audit = sorted(audit_trail, key=lambda a: (0 if a.get("change_type")=="FINANCIAL" else 1, a.get("date","")))
        for r_idx,a in enumerate(sorted_audit,6):
            is_financial = a.get("change_type") == "FINANCIAL"
            changed_cat = a["ai_category"] != a["final_category"]
            row_fill = PatternFill("solid",fgColor="FFF3CD") if is_financial else PatternFill("solid",fgColor="D4EDDA")
            ws_at.cell(r_idx,1,a.get("change_type","")).font=Font(bold=is_financial,color="CC6600" if is_financial else "006600",name="Arial",size=10)
            ws_at.cell(r_idx,2,a["date"]).font=dfont
            ws_at.cell(r_idx,3,a["description"]).font=dfont
            c=ws_at.cell(r_idx,4,a["debit"]); c.number_format='$#,##0.00'; c.font=dfont
            ws_at.cell(r_idx,5,a["ai_category"]).font=Font(color="666666",name="Arial",size=10)
            ws_at.cell(r_idx,6,a["final_category"]).font=Font(bold=changed_cat,color="000000" if not changed_cat else "0066CC",name="Arial",size=10)
            c=ws_at.cell(r_idx,7,a["ai_itc"]); c.number_format='$#,##0.00'; c.font=Font(color="666666",name="Arial",size=10)
            c=ws_at.cell(r_idx,8,a["final_itc"]); c.number_format='$#,##0.00'; c.font=Font(bold=is_financial,name="Arial",size=10)
            ws_at.cell(r_idx,9,a["ai_conf"]).font=dfont
            ws_at.cell(r_idx,10,a["changes"]).font=Font(italic=True,color="CC6600",name="Arial",size=10)
            for col in range(1,11): ws_at.cell(r_idx,col).border=bdr; ws_at.cell(r_idx,col).fill=row_fill

    # Tab: Validation Report (v3.3)
    if validation_report:
        ws_vr = wb.create_sheet("🔧 Validation Report")
        ws_vr.cell(1,1,"Post-Processing Validation Report").font=Font(bold=True,size=14,name="Arial")
        ws_vr.cell(2,1,"Auto-corrections applied to ledger rows. Review and confirm each fix.").font=Font(italic=True,color="666666",name="Arial")
        vr_hdrs = ["Rule","Severity","Description","Date","Merchant","Amount"]
        for i,(h,w) in enumerate(zip(vr_hdrs,[22,10,40,13,35,12]),1):
            c=ws_vr.cell(4,i,h); c.fill,c.font,c.border=hfill,hfont,bdr; ws_vr.column_dimensions[get_column_letter(i)].width=w
        for r_idx,v in enumerate(validation_report,5):
            ws_vr.cell(r_idx,1,v.get("rule","")).font=Font(bold=True,name="Arial")
            sev = v.get("severity","info")
            sev_color = "CC0000" if sev=="error" else "CC6600" if sev=="warning" else "006600"
            ws_vr.cell(r_idx,2,sev.upper()).font=Font(bold=True,color=sev_color,name="Arial")
            ws_vr.cell(r_idx,3,v.get("description","")).font=dfont
            ws_vr.cell(r_idx,4,v.get("date","")).font=dfont
            ws_vr.cell(r_idx,5,v.get("merchant","")).font=dfont
            if v.get("amount"): c=ws_vr.cell(r_idx,6,v["amount"]); c.number_format='$#,##0.00'; c.font=dfont
            fill_vr = red_fill if sev=="error" else yellow_fill if sev=="warning" else green_fill
            for col in range(1,7): ws_vr.cell(r_idx,col).border=bdr; ws_vr.cell(r_idx,col).fill=fill_vr

    # Tab 6: Summary
    ws5 = wb.create_sheet("Summary")
    ws5.cell(1,1,f"📊 BOOKKEEP AI PRO v{VERSION}").font=Font(bold=True,size=16,name="Arial")
    info = [("Business", biz_name), ("Industry", industry_str), ("Province", province_str),
            ("Period", period_str or summary.get("period","")), ("Transactions", len(transactions)),
            ("Needs Review", len(review)), ("Fixed Assets", len(cca_items))]
    for i,(k,v) in enumerate(info,3):
        ws5.cell(i,1,f"{k}:").font=Font(bold=True,name="Arial"); ws5.cell(i,2,v).font=dfont

    real_income=sum(t["credit"] for t in transactions if t["credit"] and t.get("type") not in ("PAYMENT","REFUND","FEE_REBATE"))
    payments_total=sum(t["credit"] for t in transactions if t["credit"] and t.get("type")=="PAYMENT")
    refunds_total=sum(t["credit"] for t in transactions if t["credit"] and t.get("type")=="REFUND")
    fee_rebates=sum(t["credit"] for t in transactions if t["credit"] and t.get("type")=="FEE_REBATE")
    purchases = [t for t in transactions if t.get("type") == "PURCHASE"]
    purchase_total = sum(t["debit"] for t in purchases if t.get("debit",0))

    r=11
    ws5.cell(r,1,"Purchase Transactions:").font=Font(bold=True,size=12,name="Arial")
    c=ws5.cell(r,2,purchase_total); c.number_format='$#,##0.00'; c.font=Font(bold=True,size=12,color="CC0000",name="Arial")
    r+=1
    ws5.cell(r,1,"Fees + Interest:").font=Font(bold=True,size=11,name="Arial")
    fees_int = sum(t["debit"] for t in transactions if t.get("type") in ("FEE","INTEREST") and t.get("debit",0))
    c=ws5.cell(r,2,fees_int); c.number_format='$#,##0.00'; c.font=Font(size=11,name="Arial")
    r+=1
    ws5.cell(r,1,"Total Expenses (all):").font=Font(bold=True,size=12,name="Arial")
    c=ws5.cell(r,2,total_exp); c.number_format='$#,##0.00'; c.font=Font(bold=True,size=12,color="CC0000",name="Arial")
    r+=1
    ws5.cell(r,1,"Payments to Card:").font=Font(bold=True,size=11,name="Arial")
    c=ws5.cell(r,2,payments_total); c.number_format='$#,##0.00'; c.font=Font(size=11,name="Arial")
    r+=1
    ws5.cell(r,1,"Merchant Refunds:").font=Font(bold=True,size=11,name="Arial")
    c=ws5.cell(r,2,refunds_total); c.number_format='$#,##0.00'; c.font=Font(size=11,color="006600",name="Arial")
    r+=1
    ws5.cell(r,1,"Fee Rebates:").font=Font(bold=True,size=11,name="Arial")
    c=ws5.cell(r,2,fee_rebates); c.number_format='$#,##0.00'; c.font=Font(size=11,name="Arial")
    r+=1
    ws5.cell(r,1,"Net Expenses:").font=Font(bold=True,size=14,name="Arial")
    c=ws5.cell(r,2,total_exp-refunds_total-fee_rebates); c.number_format='$#,##0.00'; c.font=Font(bold=True,size=14,name="Arial")
    r+=2
    ws5.cell(r,1,"Total HST ITC:").font=Font(bold=True,size=12,name="Arial")
    c=ws5.cell(r,2,total_itc); c.number_format='$#,##0.00'; c.font=Font(bold=True,size=12,color="0066CC",name="Arial")
    r+=1
    ws5.cell(r,1,"Capital Purchases (CCA):").font=Font(bold=True,size=11,name="Arial")
    c=ws5.cell(r,2,sum(t["debit"] for t in cca_items)); c.number_format='$#,##0.00'; c.font=Font(size=11,name="Arial")

    # ── Reconciliation Summary Stats ──
    if recon_matches or recon_unmatched or receipt_matches or invoice_data:
        r+=3
        ws5.cell(r,1,"═══ RECONCILIATION SUMMARY ═══").font=Font(bold=True,size=12,color="1F3864",name="Arial")
        r+=1
        recon_stats = [
            ("Documents Processed", len(set(t.get("_recon_source","") for t in transactions if t.get("_recon_source")))),
            ("Receipts Uploaded", len(receipt_matches) if receipt_matches else 0),
            ("Invoices Uploaded", len(invoice_data) if invoice_data else 0),
            ("Fully Matched", len(recon_matches) if recon_matches else 0),
            ("Unmatched Items", len(recon_unmatched) if recon_unmatched else 0),
            ("Receipts Matched", len([r for r in (receipt_matches or []) if r.get("status")=="MATCHED"])),
            ("Receipts Unmatched", len([r for r in (receipt_matches or []) if r.get("status")!="MATCHED"])),
            ("Invoices Matched", len([m for m in (recon_matches or []) if m.get("match_type","").startswith("INVOICE")])),
            ("Duplicate Flags", len([t for t in transactions if "POTENTIAL_DUPLICATE" in t.get("notes","")])),
        ]
        for k, v in recon_stats:
            ws5.cell(r,1,f"{k}:").font=Font(bold=True,name="Arial")
            ws5.cell(r,2,v).font=dfont
            r+=1
        # Reconciliation health score
        total_docs = (len(receipt_matches or [])) + (len(invoice_data or []))
        matched_docs = len([r for r in (receipt_matches or []) if r.get("status")=="MATCHED"]) + len([m for m in (recon_matches or []) if m.get("match_type","").startswith("INVOICE")])
        health = round(matched_docs/total_docs*100) if total_docs > 0 else 0
        r+=1
        ws5.cell(r,1,"Reconciliation Health:").font=Font(bold=True,size=14,name="Arial")
        color = "006600" if health >= 90 else "CC6600" if health >= 70 else "CC0000"
        ws5.cell(r,2,f"{health}%").font=Font(bold=True,size=14,color=color,name="Arial")

    ws5.column_dimensions["A"].width=28; ws5.column_dimensions["B"].width=18
    ws5.cell(r+2,1,f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}").font=Font(italic=True,color="999999",name="Arial")
    ws5.cell(r+3,1,"⚠️ For informational purposes. Consult your accountant for tax filing.").font=Font(italic=True,color="999999",name="Arial")
    return wb

# ═══════════════════════════════════════════════════════════════════
# STEP 3: PROCESS
# ═══════════════════════════════════════════════════════════════════
st.header("3️⃣ Generate Report")

if st.button("🚀 Process Statement(s)", type="primary", use_container_width=True):
    if not statement_files:
        st.error("Upload at least one bank statement."); st.stop()

    start = time.time()
    system_prompt = FULL_INSTRUCTIONS if use_full else SMART_CONDENSED
    system_config = [{"type":"text","text":system_prompt,"cache_control":{"type":"ephemeral"}}]
    all_categorized = []
    all_by_source = {}
    total_cost = 0
    progress = st.empty()
    total_text_count = 0
    total_vision_count = 0

    province = "ON"
    if province_override != "Auto-Detect":
        # normalize_province handles both full names ("Ontario") and codes ("ON")
        province = normalize_province(province_override)

    # Fix #18: Sanitize business name for API prompts — strip anything that could be prompt injection
    safe_biz_name = re.sub(r'[^\w\s&\'-]', '', business_name).strip()[:100]

    # ═══════════════════════════════════════════════════════════════
    # PROCESS EACH UPLOADED FILE
    # ═══════════════════════════════════════════════════════════════
    for file_idx, statement_file in enumerate(statement_files):
        file_label = statement_file.name
        st.subheader(f"📄 File {file_idx+1}/{len(statement_files)}: {file_label}")
        bank_name = bank_override if bank_override != "Auto-Detect" else "Auto"
        file_txns = []

        # ── CSV PATH ──
        if file_label.lower().endswith(('.csv', '.tsv')):
            with st.spinner(f"📊 Parsing CSV: {file_label}..."):
                csv_txns, csv_bank = parse_csv_statement(statement_file)
            if csv_txns:
                st.success(f"✅ CSV: {csv_bank} | {len(csv_txns)} transactions (zero Vision cost)")
                if bank_name == "Auto": bank_name = csv_bank
                raw_txns = csv_txns
                pre_cats = {}
                for i, t in enumerate(raw_txns):
                    hint = pre_categorize_merchant(t.get("description",""))
                    if hint:
                        pre_cats[i] = hint
                        t["_suggested_category"] = hint
                if pre_cats:
                    st.info(f"💡 {len(pre_cats)}/{len(raw_txns)} matched known merchants")
                chunks = chunk_transactions(raw_txns, chunk_size=80)
                offset = 0
                for ci, chunk in enumerate(chunks):
                    progress.info(f"⏳ CSV batch {ci+1}/{len(chunks)}...")
                    parts = []
                    for i, t in enumerate(chunk):
                        amt = t.get('debit', 0) if t.get('debit', 0) else -t.get('credit', 0)
                        ln = f"{t['date']}\t{t['description']}\t{amt:.2f}"
                        hint = pre_cats.get(offset+i)
                        if hint: ln += f"\t[HINT:{hint['category']}]"
                        parts.append(ln)
                    offset += len(chunk)
                    prompt = f"BUSINESS: {safe_biz_name} | INDUSTRY: {industry} | ACCOUNT: {account_type} | SOURCE: {bank_name} | PROVINCE: {province} | PERIOD: {period or 'auto'}\n\n" + "\n".join(parts)
                    try:
                        resp = call_claude_with_retry(
                            client, model_id, 3000, system_config,
                            [{"role":"user","content":prompt}],
                            label=f"CSV batch {ci+1}/{len(chunks)}"
                        )
                        resp_lines = [l.strip() for l in resp.strip().split("\n") if l.strip()]
                        # TSV validation: only accept lines with enough columns
                        valid_lines = [l for l in resp_lines if validate_tsv_line(l.split("\t"))]
                        for idx, t in enumerate(chunk):
                            if idx < len(valid_lines):
                                p = valid_lines[idx].split("\t")
                                try:
                                    txn = parse_response_line(p, bank_name, province)
                                    txn["date"] = t["date"]; txn["description"] = t["description"]
                                    txn["debit"] = t.get("debit",0); txn["credit"] = t.get("credit",0)
                                    file_txns.append(txn)
                                except Exception as row_err:
                                    file_txns.append({"date":t['date'],"source":bank_name,"description":t['description'],
                                        "debit":t.get('debit',0),"credit":t.get('credit',0),"balance":0,"type":"",
                                        "category":"❓ Uncategorized","t2125":"","itc_rule":"","itc_amount":0,
                                        "confidence":"50","notes":f"Parse error: {row_err}"})
                            else:
                                file_txns.append({"date":t['date'],"source":bank_name,"description":t['description'],
                                    "debit":t.get('debit',0),"credit":t.get('credit',0),"balance":0,"type":"",
                                    "category":"❓ Uncategorized","t2125":"","itc_rule":"","itc_amount":0,
                                    "confidence":"50","notes":"No AI response"})
                        cd = 0.1 if ci > 0 else 1.0
                        total_cost += (sys_tokens*rate_in*cd/1e6) + (len(prompt)//4*rate_in/1e6) + ((len(resp)//4)*rate_out/1e6)
                    except Exception as e:
                        st.error(f"❌ CSV batch {ci+1} failed permanently: {e}")
                        # Add uncategorized stubs so the batch isn't silently dropped
                        for t in chunk:
                            file_txns.append({"date":t['date'],"source":bank_name,"description":t['description'],
                                "debit":t.get('debit',0),"credit":t.get('credit',0),"balance":0,"type":"",
                                "category":"❓ Uncategorized","t2125":"","itc_rule":"","itc_amount":0,
                                "confidence":"0","notes":f"API_ERROR: {str(e)[:60]}"})
                source_key = bank_name if bank_name != "Auto" else file_label
                all_by_source[source_key] = file_txns
                all_categorized.extend(file_txns)
                total_text_count += len(file_txns)
                continue

        # ── PDF PATH ──
        with st.spinner(f"📄 Reading {file_label}..."):
            good_text, bad_pages, total_pages, page_texts = extract_pdf_hybrid(statement_file, max_pages)

        all_text_pages = total_pages - len(bad_pages)
        if bad_pages:
            st.warning(f"⚠️ {len(bad_pages)}/{total_pages} pages CID/garbled → Vision")
        if all_text_pages > 0 and len(good_text) > 50:
            st.success(f"✅ Text: {all_text_pages} pages, {len(good_text):,} chars")

        is_fully_scanned = len(good_text.strip()) < 50 and len(bad_pages) == total_pages

        # ── PHASE 1: Text pages ──
        if good_text and len(good_text.strip()) > 50:
            with st.spinner("🔍 Detecting bank..."):
                detected_bank, schema_txns, schema_key = detect_and_parse(good_text)
                if province_override == "Auto-Detect":
                    province = detect_province_from_text(good_text)

            raw_txns = []
            if detected_bank and len(schema_txns) > 0:
                st.success(f"✅ **{detected_bank}** | {len(schema_txns)} txns | Province: {province}")
                raw_txns = schema_txns
                if bank_name == "Auto": bank_name = detected_bank
                st.session_state.detected_bank = detected_bank
            else:
                with st.spinner("🔍 Legacy parser..."):
                    raw_txns = parse_td_statement(good_text)
                if raw_txns:
                    st.success(f"✅ Legacy: {len(raw_txns)} txns")

            if raw_txns:
                pre_cats = {i: t["_suggested_category"] for i, t in enumerate(raw_txns) if t.get("_suggested_category")}
                if pre_cats:
                    st.info(f"💡 {len(pre_cats)}/{len(raw_txns)} matched merchants")
                chunks = chunk_transactions(raw_txns, chunk_size=80)
                offset = 0
                for ci, chunk in enumerate(chunks):
                    progress.info(f"⏳ Text batch {ci+1}/{len(chunks)}...")
                    parts = []
                    for i, t in enumerate(chunk):
                        amt = t.get('debit', 0) if t.get('debit', 0) else -t.get('credit', 0)
                        ln = f"{t['date']}\t{t['description']}\t{amt:.2f}"
                        hint = pre_cats.get(offset+i)
                        if hint: ln += f"\t[HINT:{hint['category']}]"
                        if t.get("tax_note"): ln += f"\t[{t['tax_note']}]"
                        parts.append(ln)
                    offset += len(chunk)
                    prompt = f"BUSINESS: {safe_biz_name} | INDUSTRY: {industry} | ACCOUNT: {account_type} | SOURCE: {bank_name} | PROVINCE: {province} | PERIOD: {period or 'auto'}\n\n" + "\n".join(parts)
                    try:
                        resp = call_claude_with_retry(
                            client, model_id, 3000, system_config,
                            [{"role":"user","content":prompt}],
                            label=f"Text batch {ci+1}/{len(chunks)}"
                        )
                        resp_lines=[l.strip() for l in resp.strip().split("\n") if l.strip()]
                        valid_lines = [l for l in resp_lines if validate_tsv_line(l.split("\t"))]
                        for idx,t in enumerate(chunk):
                            if idx<len(valid_lines):
                                p=valid_lines[idx].split("\t")
                                try:
                                    txn = parse_response_line(p, bank_name, province)
                                    txn["date"]=t["date"]; txn["description"]=t["description"]
                                    txn["debit"]=t.get("debit",0); txn["credit"]=t.get("credit",0); txn["balance"]=t.get("balance",0)
                                    file_txns.append(txn)
                                except Exception as row_err:
                                    file_txns.append({"date":t['date'],"source":bank_name,"description":t['description'],
                                        "debit":t.get('debit',0),"credit":t.get('credit',0),"balance":0,"type":"",
                                        "category":"❓ Uncategorized","t2125":"","itc_rule":"","itc_amount":0,
                                        "confidence":"0","notes":f"Row parse error: {row_err}"})
                            else:
                                file_txns.append({"date":t['date'],"source":bank_name,"description":t['description'],
                                    "debit":t.get('debit',0),"credit":t.get('credit',0),"balance":0,"type":"",
                                    "category":"Uncategorized","t2125":"","itc_rule":"","itc_amount":0,"confidence":"50","notes":"No response"})
                        cd=0.1 if ci>0 else 1.0
                        total_cost+=(sys_tokens*rate_in*cd/1e6)+(len(prompt)//4*rate_in/1e6)+((len(resp)//4)*rate_out/1e6)
                    except Exception as e:
                        st.warning(f"⚠️ Text batch {ci+1}: {e}")

            elif not bad_pages:
                st.info("📝 No parser matched — Claude fallback")
                text_chunks = [good_text[i:i+4000] for i in range(0, len(good_text), 4000)]
                for ci, chunk_text in enumerate(text_chunks):
                    if not chunk_text.strip(): continue
                    progress.info(f"⏳ Fallback {ci+1}/{len(text_chunks)}...")
                    prompt = f"BUSINESS: {safe_biz_name} | INDUSTRY: {industry} | ACCOUNT: {account_type} | BANK: {bank_name} | PROVINCE: {province} | PERIOD: {period or 'auto'}\n\nParse ALL transactions:\n\n{chunk_text}\n\nOutput TSV: Date\\tSource\\tDescription\\tDebit\\tCredit\\tType\\tCategory\\tITC Amount\\tConfidence\\tNotes\nType: PURCHASE,REFUND,PAYMENT,FEE,INTEREST. NO headers. Just rows."
                    try:
                        resp=""
                        with client.messages.stream(model=model_id,max_tokens=4000,system=system_config,
                            messages=[{"role":"user","content":prompt}],timeout=300.0) as stream:
                            for txt in stream.text_stream: resp+=txt
                        for ln in resp.strip().split("\n"):
                            ln=ln.strip()
                            if not ln or ln.startswith("Date\t"): continue
                            p=ln.split("\t")
                            if len(p)>=4 and validate_tsv_line(p):
                                try:
                                    file_txns.append(parse_response_line(p, bank_name, province))
                                except Exception as row_err:
                                    pass  # skip malformed row, continue processing remaining rows
                        cd=0.1 if ci>0 else 1.0
                        total_cost+=(sys_tokens*rate_in*cd/1e6)+(len(prompt)//4*rate_in/1e6)+((len(resp)//4)*rate_out/1e6)
                    except Exception as e:
                        st.warning(f"⚠️ Fallback {ci+1}: {e}")

        text_count = len(file_txns)

        # ── PHASE 2: Vision for CID/scanned pages ──
        if bad_pages or is_fully_scanned:
            vision_pages = bad_pages if bad_pages else list(range(min(total_pages, max_pages)))
            try:
                import fitz
            except ImportError:
                st.error("❌ PyMuPDF not installed")
                if is_fully_scanned: st.stop()
                else: continue

            st.info(f"📸 {len(vision_pages)} CID pages → Vision...")
            try:
                page_images = render_pages_to_images(statement_file, vision_pages, dpi=150)
            except Exception as e:
                st.error(f"❌ Rendering: {e}"); page_images = {}

            if page_images:
                image_list = sorted(page_images.items())
                VISION_COST_CAP = 2.00  # Max $2 vision spend per statement
                vision_spend = 0.0
                for bs in range(0, len(image_list), 5):
                    if vision_spend >= VISION_COST_CAP:
                        st.warning(f"⚠️ Vision cost cap ${VISION_COST_CAP:.2f} reached — {len(image_list)-bs} pages skipped. Split large PDFs to process remainder.")
                        break
                    batch = image_list[bs:bs+5]
                    page_nums = [pn for pn, _ in batch]
                    progress.info(f"⏳ Vision: p{page_nums[0]+1}–{page_nums[-1]+1}...")
                    content = [{"type":"image","source":{"type":"base64","media_type":"image/png","data":b64}} for _, b64 in batch]
                    content.append({"type":"text","text":f"BUSINESS: {safe_biz_name} | INDUSTRY: {industry} | ACCOUNT: {account_type} | SOURCE: {bank_name} | PROVINCE: {province} | PERIOD: {period or 'auto'}\n\nCID pages. Extract ALL transactions.\nTSV: Date\\tSource\\tDescription\\tDebit\\tCredit\\tType\\tCategory\\tITC Amount\\tConfidence\\tNotes\nType: PURCHASE,REFUND,PAYMENT,FEE,INTEREST. Skip summaries. Just rows."})
                    try:
                        resp=""
                        with client.messages.stream(model=model_id,max_tokens=4000,system=system_config,
                            messages=[{"role":"user","content":content}],timeout=300.0) as stream:
                            for txt in stream.text_stream: resp+=txt
                        for ln in resp.strip().split("\n"):
                            ln=ln.strip()
                            if not ln or ln.startswith("Date\t") or ln.startswith("---"): continue
                            p=ln.split("\t")
                            if len(p)>=4 and validate_tsv_line(p):
                                try:
                                    txn = parse_response_line(p, bank_name, province)
                                    txn["notes"] = (txn.get("notes","") + " CID_VISION").strip()
                                    file_txns.append(txn)
                                except Exception as row_err:
                                    pass  # skip malformed vision row, continue remaining
                        batch_cost=(len(batch)*1600*rate_in+(len(resp)//4)*rate_out)/1e6
                        total_cost+=batch_cost; vision_spend+=batch_cost
                    except Exception as e:
                        st.warning(f"⚠️ Vision p{page_nums[0]+1}–{page_nums[-1]+1}: {e}")

        vision_count = len(file_txns) - text_count
        total_text_count += text_count
        total_vision_count += vision_count
        st.success(f"📄 {file_label}: {len(file_txns)} txns ({text_count} text + {vision_count} vision)")

        source_key = bank_name if bank_name != "Auto" else file_label
        all_by_source[source_key] = file_txns
        all_categorized.extend(file_txns)
    # ═══════════════════════════════════════════════════════════════
    # END OF PER-FILE LOOP
    # ═══════════════════════════════════════════════════════════════

    # ── STATEMENT VALIDATION — extract summary totals from each PDF ──
    # Seek all files back to start — they were consumed during extraction above
    for _sf in statement_files:
        try: _sf.seek(0)
        except Exception:
            pass
    validation_results = []
    for statement_file in statement_files:
        fname = statement_file.name
        if fname.lower().endswith(('.csv', '.tsv')):
            continue  # CSV has no summary page
        with st.spinner(f"🔎 Extracting summary from {fname} for validation..."):
            pdf_summary, val_cost = extract_statement_summary(statement_file, client, model_id, rate_in, rate_out)
            total_cost += val_cost
        # Always compute ledger-side totals regardless of whether PDF extraction worked
        # Uses all_categorized (pre-dedup) here; the dedup'd totals go in the Excel sheet
        # after post-processing completes. The sheet is populated in build_excel() below.
        source_key = None
        for sk, _sk_txns in all_by_source.items():
            if sk in fname or fname in sk or sk.split()[0] in fname:
                source_key = sk; break
        if not source_key:
            source_key = list(all_by_source.keys())[0] if len(all_by_source) == 1 else None
        file_txns_for_val = all_by_source.get(source_key, all_categorized) if source_key else all_categorized

        computed_purchases = round(sum(t.get("debit",0) for t in file_txns_for_val if t.get("type") in ("PURCHASE","")), 2)
        computed_payments  = round(sum(t.get("credit",0) for t in file_txns_for_val if t.get("type") == "PAYMENT"), 2)
        computed_fees      = round(sum(t.get("debit",0) for t in file_txns_for_val if t.get("type") in ("FEE","INTEREST")), 2)
        computed_credits   = round(sum(t.get("credit",0) for t in file_txns_for_val if t.get("type") in ("REFUND","FEE_REBATE")), 2)

        if not pdf_summary:
            st.warning(f"⚠️ {fname}: Could not extract summary page — PDF may be image-only. "
                       f"Ledger totals computed: Purchases ${computed_purchases:,.2f} | "
                       f"Payments ${computed_payments:,.2f} | Credits ${computed_credits:,.2f}. "
                       f"Verify against statement manually.")
            # Still add to validation_results so the sheet shows computed totals
            validation_results.append({
                "file": fname,
                "pdf_summary": {},
                "validations": [
                    {"field": "Total Purchases (computed only — PDF unreadable)", "pdf": 0, "computed": computed_purchases, "variance": 0, "status": "⚠️ PDF N/A"},
                    {"field": "Total Payments (computed only — PDF unreadable)",  "pdf": 0, "computed": computed_payments,  "variance": 0, "status": "⚠️ PDF N/A"},
                    {"field": "Total Credits (computed only — PDF unreadable)",   "pdf": 0, "computed": computed_credits,   "variance": 0, "status": "⚠️ PDF N/A"},
                ],
            })
        else:
            # Fail loudly if critical fields are missing
            critical_fields = ["TOTAL_PURCHASES"]
            missing_critical = [f for f in critical_fields
                if f not in pdf_summary or not isinstance(pdf_summary.get(f), (int,float)) or pdf_summary.get(f) == 0]
            if missing_critical:
                st.error(f"🔴 {fname}: Could not extract {', '.join(missing_critical)} from PDF summary. Validation unreliable — verify manually.")

            pdf_purchases = pdf_summary.get("TOTAL_PURCHASES", 0)
            pdf_payments  = pdf_summary.get("TOTAL_PAYMENTS", 0)
            pdf_interest  = pdf_summary.get("TOTAL_INTEREST", 0)
            pdf_fees      = pdf_summary.get("TOTAL_FEES", 0)
            pdf_opening   = pdf_summary.get("OPENING_BALANCE", 0)
            pdf_closing   = pdf_summary.get("CLOSING_BALANCE", 0)
            pdf_credits   = pdf_summary.get("TOTAL_CREDITS", 0)

            validations = []
            def add_val(field, pdf_val, computed_val):
                if isinstance(pdf_val, (int, float)) and pdf_val > 0:
                    variance = round(abs(pdf_val - computed_val), 2)
                    status = "✅ MATCH" if variance <= 0.05 else "⚠️ CLOSE" if variance <= 1.0 else "❌ MISMATCH"
                    validations.append({"field": field, "pdf": pdf_val, "computed": computed_val, "variance": variance, "status": status})

            add_val("Total Purchases", pdf_purchases, computed_purchases)
            add_val("Total Payments", pdf_payments, computed_payments)
            add_val("Interest + Fees",
                    (pdf_interest if isinstance(pdf_interest,(int,float)) else 0) +
                    (pdf_fees if isinstance(pdf_fees,(int,float)) else 0),
                    computed_fees)
            add_val("Total Credits/Refunds", pdf_credits, computed_credits)

            # Balance equation check: Opening + Purchases + Fees − Payments − Credits = Closing
            if isinstance(pdf_opening,(int,float)) and isinstance(pdf_closing,(int,float)) and pdf_opening > 0:
                expected_closing = (pdf_opening
                    + (pdf_purchases if isinstance(pdf_purchases,(int,float)) else 0)
                    + (pdf_fees if isinstance(pdf_fees,(int,float)) else 0)
                    + (pdf_interest if isinstance(pdf_interest,(int,float)) else 0)
                    - (pdf_payments if isinstance(pdf_payments,(int,float)) else 0)
                    - (pdf_credits if isinstance(pdf_credits,(int,float)) else 0))
                bal_variance = round(abs(expected_closing - pdf_closing), 2)
                bal_status = "✅ BALANCED" if bal_variance <= 0.05 else "⚠️ CLOSE" if bal_variance <= 1.0 else "❌ IMBALANCED"
                validations.append({"field": "Balance Equation", "pdf": pdf_closing, "computed": round(expected_closing,2), "variance": bal_variance, "status": bal_status})

            # If no validations (all pdf values were 0 or missing), add computed-only rows
            if not validations:
                validations = [
                    {"field": "Total Purchases (PDF fields returned 0)", "pdf": 0, "computed": computed_purchases, "variance": 0, "status": "⚠️ PDF N/A"},
                    {"field": "Total Payments (PDF fields returned 0)",  "pdf": 0, "computed": computed_payments,  "variance": 0, "status": "⚠️ PDF N/A"},
                    {"field": "Total Credits (PDF fields returned 0)",   "pdf": 0, "computed": computed_credits,   "variance": 0, "status": "⚠️ PDF N/A"},
                ]

            validation_results.append({"file": fname, "pdf_summary": pdf_summary, "validations": validations})

            # Show validation in UI
            all_match = all(v["status"].startswith("✅") for v in validations)
            has_imbalance = any(v["status"] == "❌ IMBALANCED" for v in validations)
            if validations:
                if all_match:
                    st.success(f"🔎 {fname}: All totals validated ✅")
                else:
                    # Top-level banner for balance integrity failure
                    if has_imbalance:
                        bal_v = next(v for v in validations if v["status"] == "❌ IMBALANCED")
                        st.error(
                            f"⚠️ **STATEMENT INTEGRITY FAILURE — {fname}**\n\n"
                            f"Opening + debits − credits does not equal the closing balance on the PDF.\n"
                            f"Expected closing: **${bal_v['computed']:,.2f}** | "
                            f"PDF closing: **${bal_v['pdf']:,.2f}** | "
                            f"Variance: **${bal_v['variance']:,.2f}**\n\n"
                            f"Possible causes: missing transactions, PDF truncation, or round-off. "
                            f"Do not file this statement without manual review."
                        )
                    for v in validations:
                        if not v["status"].startswith("✅"):
                            icon = "❌" if v["status"] == "❌ IMBALANCED" else "⚠️"
                            st.warning(f"{icon} {fname}: {v['field']} — PDF: ${v['pdf']:,.2f} vs Computed: ${v['computed']:,.2f} (Δ ${v['variance']:,.2f}) {v['status']}")

    # ── RECEIPTS (before reconciliation) ──
    receipt_matches = []
    if receipt_files:
        with st.spinner(f"📸 Processing {len(receipt_files)} receipt(s)..."):
            receipt_matches, rcpt_cost = process_receipts(receipt_files, all_categorized, client, model_id, system_config, rate_in, rate_out, period_hint=period_hint)
            total_cost += rcpt_cost
        matched_ct = len([r for r in receipt_matches if r["status"] == "MATCHED"])
        st.success(f"📸 Receipts: {matched_ct}/{len(receipt_matches)} matched")

    # ── INVOICES (before reconciliation) ──
    invoice_data = []
    if invoice_files:
        with st.spinner(f"📋 Parsing {len(invoice_files)} invoice(s)..."):
            invoice_data, inv_cost = process_invoices(invoice_files, client, model_id, system_config, rate_in, rate_out)
            total_cost += inv_cost
        st.success(f"📋 Invoices: {len(invoice_data)} | AP: ${sum(i.get('total',0) for i in invoice_data):,.2f}")

    # ── FULL RECONCILIATION ──
    recon_matches = []
    recon_unmatched = []
    recon_duplicates = []
    if len(all_by_source) > 1 or receipt_matches or invoice_data:
        with st.spinner("🔄 Cross-matching all documents..."):
            recon_matches, recon_unmatched, recon_duplicates = reconcile_all(all_by_source, receipt_matches, invoice_data)
        if recon_matches:
            st.success(f"🔄 {len(recon_matches)} matches across documents")
        if recon_unmatched:
            st.warning(f"⚠️ {len(recon_unmatched)} unmatched — verify manually")
        if recon_duplicates:
            st.error(f"🔴 {len(recon_duplicates)} potential duplicates")

    elapsed = time.time() - start
    progress.empty()

    if not all_categorized:
        st.error("❌ No transactions found."); st.stop()

    # ── POST-PROCESSING ──
    # T2125 line mapping
    T2125_MAP = {
        "Motor Vehicle Expense": "9281", "Meals & Entertainment": "8523",
        "Office Supplies": "8810", "Utilities": "8220", "Bank Charges": "8710",
        "Insurance": "8690", "Materials & Supplies": "8811", "Rent": "8910",
        "Delivery & Shipping": "8730", "Shipping": "8730", "Advertising": "8520",
        "Travel": "9200", "Professional Fees": "8860", "Subcontracts": "8590",
        "Cost of Goods": "8320", "Repairs & Maintenance": "8960",
        "Government Remittances": "",  # Not a T2125 line but valid category
    }
    # Business use % defaults (CPA can override in Excel)
    BIZ_USE = {
        "Motor Vehicle Expense": 100,  # default 100%, CPA adjusts
        "Meals & Entertainment": 50,   # CRA 50% rule
        "Owner Draw / Personal": 0,
    }

    # Date normalization: ensure all dates have year and are sortable
    def normalize_date(date_str, period_hint=""):
        """Convert dates like 'Dec 23' to '2024-12-23' using period context.
        Falls back to statement-detected period rather than current year."""
        if not date_str or not date_str.strip():
            return date_str
        d = date_str.strip()
        # Already has year in YYYY-MM-DD format
        if re.match(r'^\d{4}-\d{2}-\d{2}$', d):
            return d
        # Try parsing with known formats
        for fmt, out in [
            ("%b %d, %Y", None), ("%d-%b-%Y", None), ("%Y-%m-%d", None),
            ("%m/%d/%Y", None), ("%d/%m/%Y", None), ("%m/%d/%y", None)]:
            try:
                dt = datetime.strptime(d, fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue
        # No year — try "Dec 23" or "Jan 09" format
        for fmt in ("%b %d", "%b. %d", "%d %b"):
            try:
                dt = datetime.strptime(d, fmt)
                # Determine year from period hint, NOT current year
                year = None
                if period_hint:
                    # Extract years from period like "Dec 2024 - Jan 2025"
                    years = re.findall(r'20\d{2}', period_hint)
                    if years:
                        if dt.month >= 10:  # Oct-Dec → first year
                            year = int(years[0])
                        elif dt.month <= 3 and len(years) > 1:  # Jan-Mar → second year
                            year = int(years[-1])
                        else:
                            year = int(years[0])
                # Fallback: infer from other transactions already parsed
                if year is None:
                    existing_years = set()
                    for t in all_categorized:
                        td = t.get("date", "")
                        if td and len(td) >= 4 and td[:4].isdigit():
                            existing_years.add(int(td[:4]))
                    if existing_years:
                        year = max(existing_years)  # use most recent year found
                    else:
                        year = datetime.now().year
                return dt.replace(year=year).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return d  # return as-is if nothing worked

    GHOST_PATTERNS = re.compile(
        r'(?i)^('
        # ── Universal balance/summary lines ──────────────────────────────────
        r'opening balance|previous balance|closing balance|balance forward|'
        r'subtotal|sub-total|sub total|'
        r'previous statement|new balance|minimum payment|payment due|'
        r'credit limit|available credit|annual percentage|APR\b|'
        # ── Credit total lines (all major Canadian banks) ─────────────────────
        r'total charges|total credits|total payments|total new charges|total new activity|'
        r'total payments received|total new transactions|total purchases|'
        r'total interest|total fees|total amount due|total balance|'
        # ── Triangle Mastercard / Canadian Tire Bank section headers ──────────
        r'payments received .{3,30} to .{3,30}|'   # "Payments received Apr 17 to May 16"
        r'returns and other credits|'               # section header
        # ── TD section headers ────────────────────────────────────────────────
        r'new charges and credits|activity since last statement|'
        r'total new charges and credits|previous statement balance|'
        # ── RBC section headers ───────────────────────────────────────────────
        r'new activity|previous month balance|total new activity|'
        r'total credits and payments|rbcroyalbank|royal bank of canada|'
        # ── BMO section headers ───────────────────────────────────────────────
        r'previous statement balance|total new activity - mastercard|'
        r'bmo bank of montreal|bank of montreal statement|'
        # ── CIBC section headers ──────────────────────────────────────────────
        r'your new charges and credits|cibc credit card statement|'
        r'total charges and credits|'
        # ── Scotiabank section headers ────────────────────────────────────────
        r'previous balance carried forward|total transactions|'
        r'scotiabank visa statement|telescotia statement|'
        # ── Amex section headers ──────────────────────────────────────────────
        r'total new charges - american express|'
        r'membership rewards|amex canada|american express canada|'
        # ── Desjardins / National Bank French section headers ─────────────────
        r'solde precedent|nouveau solde|paiement recu|'
        r'total des achats|total des paiements|'
        r'banque nationale|desjardins statement|'
        # ── Generic financial statement noise ─────────────────────────────────
        r'account summary|statement summary|account overview|'
        r'annual interest rate|daily interest rate|'
        r'grace period|payment received - thank you|'
        r'for billing inquiries|questions about your account'
        r')$')
    # v3.8: Apply permanent vendor rules (deterministic override before AI results are used)
    all_categorized, rules_applied = apply_vendor_rules(all_categorized)
    if rules_applied:
        st.info(f"📋 {rules_applied} transaction(s) overridden by your saved vendor rules.")

    # v3.12: MERCHANT PATTERN AUTO-FILL — catch Uncategorized that match known patterns
    # This runs after Claude + vendor rules, catching Vision-extracted transactions
    # that Claude couldn't categorize but our merchant dictionary knows
    pattern_fills = 0
    for t in all_categorized:
        cat = t.get("category", "")
        if "Uncategorized" in cat or cat.strip() == "" or cat is None:
            if t.get("type") in ("PAYMENT", "FEE_REBATE"):
                continue  # Payments should stay blank
            desc = t.get("description", "")
            hint = pre_categorize_merchant(desc)
            if hint and hint.get("category"):
                t["category"] = hint["category"]
                t["notes"] = (t.get("notes","") + " AUTO_PATTERN_FILL").strip()
                pattern_fills += 1
            else:
                # v3.12: Try vendor memory lookup
                mem_cat = recall_vendor(desc)
                if mem_cat:
                    t["category"] = mem_cat
                    t["notes"] = (t.get("notes","") + " VENDOR_MEMORY").strip()
                    pattern_fills += 1
    if pattern_fills:
        st.info(f"🔍 {pattern_fills} Uncategorized entries auto-filled from merchant patterns + vendor memory")

    # v3.12: CONSISTENCY PASS — if a merchant appears 3+ times and most are categorized,
    # fill the Uncategorized ones with the majority category
    from collections import Counter
    merchant_cats = {}  # merchant_key → [categories]
    for t in all_categorized:
        desc = clean_description(t.get("description","")).upper()[:25]
        if desc and t.get("type") not in ("PAYMENT", "FEE_REBATE"):
            if desc not in merchant_cats:
                merchant_cats[desc] = []
            merchant_cats[desc].append(t.get("category", ""))
    consistency_fills = 0
    for t in all_categorized:
        cat = t.get("category", "")
        if "Uncategorized" in cat or cat.strip() == "":
            if t.get("type") in ("PAYMENT", "FEE_REBATE"):
                continue
            desc = clean_description(t.get("description","")).upper()[:25]
            if desc in merchant_cats:
                cats_for_merchant = [c for c in merchant_cats[desc] if c and "Uncategorized" not in c and c.strip()]
                if len(cats_for_merchant) >= 2:
                    most_common = Counter(cats_for_merchant).most_common(1)[0][0]
                    t["category"] = most_common
                    t["notes"] = (t.get("notes","") + " CONSISTENCY_FILL").strip()
                    consistency_fills += 1
    if consistency_fills:
        st.info(f"🔄 {consistency_fills} entries auto-filled by consistency (same merchant = same category)")

    # v3.12: AMOUNT-BASED PYTHON HEURISTICS — deterministic, zero API cost
    # Catches obvious patterns that don't need AI at all
    heuristic_fills = 0
    for t in all_categorized:
        cat = t.get("category", "")
        if cat and "Uncategorized" not in cat and cat.strip() != "":
            continue
        if t.get("type") in ("PAYMENT", "FEE_REBATE"):
            continue
        desc_upper = t.get("description", "").upper()
        debit = t.get("debit", 0) or 0
        credit = t.get("credit", 0) or 0
        filled = False

        # Rule 1: INTEREST in description → Bank Charges
        if "INTEREST" in desc_upper and debit > 0:
            t["category"] = "Bank Charges"; t["type"] = "INTEREST"; filled = True
        # Rule 2: FEE / CHARGE in description → Bank Charges
        elif any(kw in desc_upper for kw in ("ANNUAL FEE", "SERVICE FEE", "MONTHLY FEE", "LATE FEE", "NSF", "OVERLIMIT")) and debit > 0:
            t["category"] = "Bank Charges"; t["type"] = "FEE"; filled = True
        # Rule 3: E-transfer $500+ for Construction → Subcontracts
        elif industry == "Construction/Trades" and debit >= 500 and any(kw in desc_upper for kw in ("E-TRANSFER", "ETRANSFER", "E TRANSFER", "INTERAC", "E-TFR")):
            t["category"] = "Subcontracts"; filled = True
        # Rule 4: CREDIT ADJUSTMENT / OTHER CREDITS → leave as Uncategorized but set type
        elif any(kw in desc_upper for kw in ("CREDIT ADJUSTMENT", "OTHER CREDITS", "CREDIT VILLAGE")):
            if credit > 0: t["type"] = "REFUND"
            # Don't fill category — genuinely ambiguous
        # Rule 5: Small amounts $1-$15 at unknown merchants → Meals & Entertainment
        elif 1 <= debit <= 15 and t.get("type") == "PURCHASE":
            t["category"] = "Meals & Entertainment"; filled = True
        # Rule 6: PAYPAL / TIKTOK / online marketplace → Office Supplies
        elif any(kw in desc_upper for kw in ("PAYPAL", "TIKTOK", "FIVERR", "UPWORK")):
            t["category"] = "Office Supplies"; filled = True
        # Rule 7: HEALTH / MEDICAL / PHARMACY → Owner Draw / Personal
        elif any(kw in desc_upper for kw in ("HEALTH", "MEDICAL", "PHARMACY", "MACKENZIE HEALTH", "HOSPITAL", "CLINIC")):
            t["category"] = "Owner Draw / Personal"; filled = True
        # Rule 8: DEFENDER / security service → Office Supplies
        elif "DEFENDER" in desc_upper:
            t["category"] = "Office Supplies"; filled = True
        # Rule 9: MARINA / recreation → Owner Draw / Personal
        elif any(kw in desc_upper for kw in ("MARINA", "GOLF", "BOWLING", "CINEMA", "MOVIE", "THEATRE")):
            t["category"] = "Owner Draw / Personal"; filled = True
        # Rule 10: ND GRAPHICS / print shop → Advertising
        elif any(kw in desc_upper for kw in ("GRAPHICS", "PRINT", "SIGN", "BANNER")):
            t["category"] = "Advertising"; filled = True

        if filled:
            t["notes"] = (t.get("notes","") + " HEURISTIC_FILL").strip()
            heuristic_fills += 1

    if heuristic_fills:
        st.info(f"🧮 {heuristic_fills} entries auto-categorized by amount/keyword heuristics (zero API cost)")

    # v3.12: TWO-PASS CLAUDE — send remaining Uncategorized back for focused categorization
    # Claude performs much better with 20 focused rows than 80 in a batch
    remaining_uncat = [t for t in all_categorized
                       if ("Uncategorized" in t.get("category","") or t.get("category","").strip() == "")
                       and t.get("type") not in ("PAYMENT", "FEE_REBATE")
                       and t.get("debit", 0) > 0]

    if remaining_uncat and len(remaining_uncat) <= 60:
        with st.spinner(f"🔄 Pass 2: Re-categorizing {len(remaining_uncat)} remaining Uncategorized with focused context..."):
            # Build a focused prompt with only the uncategorized rows
            pass2_lines = []
            for t in remaining_uncat:
                amt = t.get("debit", 0)
                desc = clean_description(t.get("description", ""))
                pass2_lines.append(f"{t.get('date','')}\t{desc}\t{amt:.2f}")

            pass2_prompt = (
                f"BUSINESS: {business_name} | INDUSTRY: {industry} | PROVINCE: {province}\n"
                f"ACCOUNT: {account_type} | PERIOD: {period or 'auto'}\n\n"
                f"These {len(remaining_uncat)} transactions could NOT be categorized in the first pass.\n"
                f"Give your BEST guess for each. Use ONLY these categories:\n"
                f"Motor Vehicle Expense, Meals & Entertainment, Office Supplies, Utilities, "
                f"Bank Charges, Insurance, Materials & Supplies, Rent, Advertising, Travel, "
                f"Professional Fees, Subcontracts, Cost of Goods, Repairs & Maintenance, "
                f"Delivery & Shipping, Government Remittances, Owner Draw / Personal, Uncategorized\n\n"
                f"For Construction/Trades: hardware stores and building suppliers = Materials & Supplies.\n"
                f"Grocery stores and personal shops = Owner Draw / Personal.\n"
                f"Restaurants and food = Meals & Entertainment.\n\n"
                f"Output ONLY TSV rows: Date\\tSource\\tDescription\\tDebit\\tCredit\\tType\\tCategory\\tITC Amount\\tConfidence\\tNotes\n\n"
                + "\n".join(pass2_lines)
            )

            try:
                resp = call_claude_with_retry(
                    client, model_id, max_tokens=2000,
                    system_config=system_config,
                    messages=[{"role":"user","content":pass2_prompt}],
                    timeout=120.0, label="Pass 2 categorization"
                )
                pass2_filled = 0
                resp_lines = [l.strip() for l in resp.strip().split("\n") if l.strip()]
                for idx, t in enumerate(remaining_uncat):
                    if idx < len(resp_lines):
                        p = resp_lines[idx].split("\t")
                        if len(p) >= 7:
                            new_cat = p[6].strip() if len(p) > 6 else ""
                            # Strip T2125 appendages
                            new_cat = re.sub(r'\s*\(T2125.*?\)', '', new_cat).strip()
                            new_cat = re.sub(r'\s*\(Line\s*\d+\)', '', new_cat).strip()
                            if new_cat and "Uncategorized" not in new_cat and new_cat in [v for v in VALID_CATEGORIES if v != "❓ Uncategorized"]:
                                t["category"] = new_cat
                                t["notes"] = (t.get("notes","") + " PASS2_FILL").strip()
                                # Use Pass 2 confidence if lower
                                try:
                                    p2_conf = int(p[8].strip()) if len(p) > 8 else 65
                                    t["confidence"] = str(min(p2_conf, int(t.get("confidence","70"))))
                                except (ValueError, TypeError):
                                    pass
                                pass2_filled += 1
                if pass2_filled:
                    st.success(f"🔄 Pass 2: {pass2_filled}/{len(remaining_uncat)} re-categorized by focused Claude pass")
                    total_cost += (len(pass2_prompt)//4 * rate_in + len(resp)//4 * rate_out) / 1e6
                else:
                    st.info(f"🔄 Pass 2: Claude could not resolve remaining {len(remaining_uncat)} — these are genuinely ambiguous")
            except Exception as e:
                st.warning(f"⚠️ Pass 2 failed: {e} — {len(remaining_uncat)} entries remain Uncategorized")
    elif remaining_uncat:
        st.info(f"ℹ️ {len(remaining_uncat)} Uncategorized entries — too many for Pass 2 (limit 60). Use vendor rules for bulk correction.")

    filtered = []
    audit_trail = []   # v3.5: tracks Claude raw → post-processing changes
    ghost_count = 0
    period_hint = period or ""
    # Fix #19: If no period entered, try to infer from validation/PDF extraction
    if not period_hint and validation_results:
        for vr in validation_results:
            ps = vr.get("pdf_summary", {})
            if isinstance(ps, dict):
                sp = ps.get("STATEMENT_PERIOD", "")
                if sp and sp != "UNKNOWN" and isinstance(sp, str):
                    period_hint = sp
                    break
    for t in all_categorized:
        if GHOST_PATTERNS.search(t.get("description", "")):
            ghost_count += 1; continue
        # v3.5: snapshot Claude's raw output before any post-processing
        _ai_category = t.get("category", "")
        _ai_notes    = t.get("notes", "")
        _ai_itc      = t.get("itc_amount", 0)
        _ai_conf     = t.get("confidence", "")
        _changes     = []
        # Fix #3: Ensure PAYMENT/FEE_REBATE rows have clean empty category, not None
        if t.get("type") in ("PAYMENT", "FEE_REBATE"):
            if t.get("category") is None or t.get("category", "").strip() == "":
                t["category"] = ""
            t["itc_amount"] = 0
            t["itc_rule"] = "No"
        if t.get("type") == "REFUND" and t.get("credit", 0) > 0:
            t["notes"] = (t.get("notes", "") + " EXPENSE_REDUCTION").strip()
        if t.get("debit", 0) >= 500 and t.get("type") in ("PURCHASE", ""):
            if "CCA_ASSET" not in t.get("notes", ""):
                cca_class, cca_label = get_cca_class(t.get("description",""), t["debit"])
                t["notes"] = (t.get("notes", "") + f" CCA_ASSET CCA_CLASS_{cca_class}").strip()
                t["cca_class"] = cca_class
                t["cca_label"] = cca_label
        elif t.get("debit", 0) < 500:
            # Deterministic override: strip any AI-hallucinated CCA_ASSET tags
            # AI sometimes flags electronics/equipment descriptions below threshold
            import re as _re
            t["notes"] = _re.sub(r'\bCCA_ASSET\b', '', t.get("notes", "")).strip()
            t["notes"] = _re.sub(r'\bCCA_CLASS_\S+', '', t["notes"]).strip()
            t["notes"] = _re.sub(r'\bPOTENTIAL_CCA_ASSET\b', '', t["notes"]).strip()
            t["notes"] = ' '.join(t["notes"].split())  # collapse extra spaces
            t.pop("cca_class", None)
            t.pop("cca_label", None)
        # Normalize date
        t["date"] = normalize_date(t.get("date",""), period_hint)
        # Apply T2125 line mapping
        cat = t.get("category", "")
        # Fix: Strip T2125 appendages Claude sometimes adds (e.g. "Office Supplies (T2125 Line 8810)")
        cat = re.sub(r'\s*\(T2125.*?\)', '', cat).strip()
        cat = re.sub(r'\s*\(Line\s*\d+\)', '', cat).strip()
        cat = re.sub(r'\s*-\s*T2125.*$', '', cat).strip()
        # Normalize to valid category list
        if cat and cat not in VALID_CATEGORIES and cat != "Uncategorized" and cat != "":
            # Try substring match against valid categories
            matched = False
            for vc in VALID_CATEGORIES:
                if vc.lower() in cat.lower() or cat.lower() in vc.lower():
                    _changes.append(f"Category normalised: '{cat}' → '{vc}'")
                    cat = vc
                    matched = True
                    break
            if not matched:
                _changes.append(f"Unknown category normalised: '{cat}' → 'Uncategorized'")
                cat = "Uncategorized"
        t["category"] = cat
        t["t2125"] = T2125_MAP.get(cat, "")
        # Apply business use %
        t["biz_pct"] = BIZ_USE.get(cat, 100 if cat and "Uncategorized" not in cat and "Personal" not in cat else 0)
        # Adjust ITC for business use %
        if t.get("itc_amount", 0) > 0 and t["biz_pct"] < 100:
            t["itc_amount"] = round(t["itc_amount"] * t["biz_pct"] / 100, 2)

        # v3.3: Province override from description suffix (e.g. "STORE NAME AB")
        desc_province = detect_province_from_description(t.get("description",""))
        if desc_province and desc_province != province:
            from bank_schemas import get_itc_rate_fraction
            itc_frac = get_itc_rate_fraction(desc_province)
            debit = t.get("debit", 0)
            if debit > 0 and t.get("category") not in ("Insurance", "Bank Charges", "Owner Draw / Personal", "Shareholder Loan (Debit)", "", None):
                _old_itc = t.get("itc_amount", 0)
                t["itc_amount"] = round(debit * itc_frac, 2)
                t["notes"] = (t.get("notes","") + f" ITC_PROV={desc_province}").strip()
                _changes.append(f"ITC overridden for province {desc_province}: ${_old_itc:.2f} → ${t['itc_amount']:.2f}")

        # v3.3: Merchant normalization
        raw_desc = t.get("description", "")
        norm_desc, was_changed = normalize_merchant(raw_desc)
        if was_changed:
            t["description_raw"] = raw_desc
            t["description"] = norm_desc
            _changes.append(f"Merchant normalised: '{raw_desc}' → '{norm_desc}'")

        # v3.11: Amazon confidence penalty for Construction/Trades
        # Amazon descriptions are opaque codes — can't distinguish tools from personal
        desc_upper = t.get("description", "").upper()
        if industry == "Construction/Trades" and ("AMAZON" in desc_upper or "AMZN" in desc_upper):
            if t.get("category") == "Office Supplies" and t.get("type") == "PURCHASE":
                # Cap confidence at 70 — forces into Needs Review tab
                try:
                    curr_conf = int(t.get("confidence", "70"))
                    if curr_conf > 70:
                        t["confidence"] = "70"
                        _changes.append(f"Confidence capped at 70 (Amazon + Construction — verify with receipt)")
                except (ValueError, TypeError):
                    pass
                t["notes"] = (t.get("notes","") + " VERIFY_RECEIPT_AMAZON").strip()

        # v3.3: Shareholder loan detection for corporations
        new_cat = detect_shareholder_loan(t.get("description",""), t.get("category",""), business_structure)
        if new_cat != t.get("category",""):
            _changes.append(f"Shareholder loan reclassification: '{t.get('category','')}' → '{new_cat}'")
            t["category"] = new_cat
            t["notes"] = (t.get("notes","") + " SHAREHOLDER_LOAN").strip()

        # v3.3: T5018 subcontractor flag
        if detect_subcontractor_payment(t.get("description",""), t.get("debit",0), industry):
            t["notes"] = (t.get("notes","") + " T5018_CANDIDATE").strip()
            t["category"] = "Subcontracts"
            t["t2125"] = "8590"

        # v3.5: Record audit trail entry if any post-processing changes were made
        if _changes or t.get("category","") != _ai_category or t.get("itc_amount",0) != _ai_itc:
            # Fix #20: Classify change as FINANCIAL or COSMETIC for audit trail filtering
            is_financial = (
                t.get("category","") != _ai_category or
                abs((t.get("itc_amount",0) or 0) - (_ai_itc or 0)) > 0.01 or
                any(kw in c for c in _changes for kw in ("reclassification", "ITC overridden", "Confidence capped", "Shareholder"))
            )
            change_type = "FINANCIAL" if is_financial else "COSMETIC"
            audit_trail.append({
                "date":      t.get("date",""),
                "description": t.get("description_raw", t.get("description","")),
                "debit":     t.get("debit",0),
                "ai_category": _ai_category,
                "final_category": t.get("category",""),
                "ai_itc":    _ai_itc,
                "final_itc": t.get("itc_amount",0),
                "ai_conf":   _ai_conf,
                "changes":   " | ".join(_changes) if _changes else "Category/ITC adjusted by validation engine",
                "change_type": change_type,
            })
        filtered.append(t)

    # Sort by date
    filtered.sort(key=lambda t: t.get("date",""))
    if ghost_count > 0:
        st.info(f"🧹 Filtered {ghost_count} ghost entries")

    # ── v3.9: CREDIT-SIDE DEDUPLICATION ENGINE ───────────────────────────────
    # Fixes two confirmed bugs from Triangle Mastercard audit:
    #
    # BUG A — Phantom summary-row payments ($1,875.47 over-counted):
    #   Triangle MC prints "Payments received Apr 17 to May 16 — $1,875.47" as a
    #   section header. The Vision AI extracted this as a real transaction dated Apr 17.
    #   Real payments (Apr 23 $1,100 + May 13 $775.47) also extracted correctly,
    #   giving 3 rows when only 2 exist. Total payments overstated by $1,875.47.
    #
    # BUG B — Phantom refund row ($33.54 over-counted):
    #   Same pattern in "Returns and other credits" section. Vision AI extracted
    #   section total as Apr 17 Amazon $33.54. Real May 04 Amazon $33.54 also present.
    #
    # ALGORITHM — per billing cycle (critical: must work per-cycle, not globally):
    #   1. Group PAYMENT/REFUND rows into billing cycles (gap > 20 days = new cycle).
    #   2. Within each cycle: if any row's amount == sum of ALL OTHER rows in that
    #      cycle, AND at least 2 other rows exist → phantom summary row, remove it.
    #   3. Pass 1 (before cycle analysis): collapse exact duplicates within 2 days
    #      (same type + amount + description prefix → keep later-dated one).

    def _dedup_credits(txns):
        """
        Two-pass credit deduplication engine.

        Pass 1 — Exact duplicate collapse (35-day window):
          Same type + same amount + same description prefix within 35 days →
          keep the later-dated one, remove the earlier.
          Catches cross-page re-extractions (e.g. Amazon refund $33.54 appearing
          on two consecutive statement pages 17 days apart).

        Pass 2 — Forward-only, 30-day summary-row detection:
          For each PAYMENT/REFUND row P, look FORWARD (strictly later dates only)
          within a 30-day window. If P.amount == sum(forward rows) AND ≥2 forward
          rows in window → P is a phantom section-total row.
          Forward-only + 30-day window avoids pulling in rows from adjacent billing
          cycles. Triangle MC billing cycles are 30 days (17th to 16th), so a
          phantom on Apr 17 with real payments Apr 23 and May 13 is within window,
          while the next cycle's May 18 payment (31 days out) is excluded.
        """
        from datetime import datetime as _dt

        def _d(s):
            try: return _dt.strptime(str(s)[:10], "%Y-%m-%d").date()
            except (ValueError, TypeError): return None

        removals = set()

        # ── Pass 1: exact duplicate collapse ──────────────────────────────────
        # Guard: skip amounts that appear 3+ times with same desc prefix — those
        # are likely recurring monthly charges, not extraction duplicates.
        for txn_type in ("PAYMENT", "REFUND"):
            type_rows = [(i, t) for i, t in enumerate(txns) if t.get("type") == txn_type]
            dated = [(i, t, _d(t.get("date"))) for i, t in type_rows]
            dated = [(i, t, d) for i, t, d in dated
                     if d is not None and round(t.get("credit", 0) or 0, 2) > 0]
            dated.sort(key=lambda x: x[2])

            # Count occurrences of each (description_prefix, amount) pair
            from collections import Counter
            pair_counts = Counter()
            for _, t, _ in dated:
                key = ((t.get("description") or "").strip().upper()[:20],
                       round(t.get("credit", 0) or 0, 2))
                pair_counts[key] += 1

            for a in range(len(dated)):
                if dated[a][0] in removals:
                    continue
                for b in range(a + 1, len(dated)):
                    if dated[b][0] in removals:
                        continue
                    i_a, t_a, d_a = dated[a]
                    i_b, t_b, d_b = dated[b]
                    if (d_b - d_a).days > 35:
                        break
                    amt_a = round(t_a.get("credit", 0) or 0, 2)
                    amt_b = round(t_b.get("credit", 0) or 0, 2)
                    desc_a = (t_a.get("description") or "").strip().upper()[:20]
                    desc_b = (t_b.get("description") or "").strip().upper()[:20]
                    if abs(amt_a - amt_b) < 0.02 and desc_a == desc_b:
                        # Skip if this (desc, amount) pair appears 3+ times — likely recurring
                        if pair_counts.get((desc_a, amt_a), 0) >= 3:
                            break
                        removals.add(i_a)  # remove earlier-dated copy
                        break

        # ── Pass 2: forward-only 30-day summary-row detection ─────────────────
        for txn_type in ("PAYMENT", "REFUND"):
            type_rows = [(i, t) for i, t in enumerate(txns)
                         if t.get("type") == txn_type and i not in removals]
            dated = [(i, t, _d(t.get("date"))) for i, t in type_rows]
            dated = [(i, t, d) for i, t, d in dated if d is not None]
            for i_a, t_a, d_a in dated:
                if i_a in removals:
                    continue
                amt_a = round(t_a.get("credit", 0) or 0, 2)
                if amt_a <= 0:
                    continue
                # FORWARD only, 30-day window (matches one billing cycle, excludes next)
                forward = [(i_b, round(t_b.get("credit", 0) or 0, 2))
                           for i_b, t_b, d_b in dated
                           if i_b != i_a and i_b not in removals
                           and d_b > d_a and (d_b - d_a).days <= 30]
                if len(forward) < 2:
                    continue
                forward_sum = round(sum(amt for _, amt in forward), 2)
                if abs(amt_a - forward_sum) < 0.02:
                    removals.add(i_a)

        kept = [t for i, t in enumerate(txns) if i not in removals]
        return kept, len(txns) - len(kept)




    filtered, dedup_removed = _dedup_credits(filtered)
    if dedup_removed > 0:
        st.info(f"🔁 Removed {dedup_removed} phantom credit/payment row(s) — "
                f"section-header summary lines detected and excluded")
    # ── END DEDUPLICATION ENGINE ──────────────────────────────────────────────

    # ── v3.9: POSTING DATE CYCLE ASSIGNMENT ───────────────────────────────────
    # Triangle MC (and most credit cards) show BOTH transaction date and posting
    # date. For cycle assignment the POSTING DATE is authoritative — it determines
    # which monthly statement a transaction belongs to.
    # When the AI only extracts the transaction date (e.g. Feb 15), a transaction
    # that posts Feb 17 gets mis-assigned to the previous cycle.
    # We can't fix this in post-processing without the posting date, but we flag
    # any transaction whose date falls within 3 days BEFORE a cycle boundary
    # (identified by gaps in the date sequence) so the accountant can review them.
    # (Full fix requires the AI prompt to extract posting date — see instructions.txt)
    # ── END POSTING DATE NOTE ─────────────────────────────────────────────────

    # v3.3: Year-bound check for cross-year statements
    filtered = apply_year_bound(filtered, period_hint)

    # v3.3: Refund-match logic (prevent merchant refunds being classified as revenue)
    filtered = match_refunds_to_purchases(filtered)

    # v3.3: Industry remaps (Retail → COGS, etc.)
    filtered = apply_industry_remaps(filtered, industry)

    # v3.3: Post-processing validation engine
    filtered, validation_report = run_validation(filtered)
    if validation_report:
        errors = [r for r in validation_report if r["severity"] == "error"]
        warnings = [r for r in validation_report if r["severity"] == "warning"]
        if errors:
            st.error(f"🔧 Auto-fixed {len(errors)} ledger errors: {', '.join(set(r['rule'] for r in errors))}")
        if warnings:
            st.warning(f"⚠️ {len(warnings)} validation warnings auto-corrected")

    # v3.3: Coverage check
    coverage_info = {}
    if all_categorized:
        # Estimate coverage from raw text of first statement
        try:
            statement_files[0].seek(0)
            raw_sample = statement_files[0].read(50000).decode("utf-8", errors="ignore")
            est, cov_pct, cov_warn = check_transaction_coverage(raw_sample, len(filtered))
            coverage_info = {"estimated": est, "parsed": len(filtered), "coverage": cov_pct}
            if cov_warn:
                st.warning(cov_warn)
            else:
                st.success(f"✅ Coverage: {cov_pct}% ({len(filtered)} transactions parsed vs ~{est} estimated)")
        except Exception:
            pass

    # v3.3: Expense anomaly detection
    anomalies = []
    if anomaly_detection and len(filtered) >= 5:
        anomalies = detect_expense_anomalies(filtered)
        if anomalies:
            st.warning(f"📊 {len(anomalies)} expense anomalies flagged (unusually large vs category median)")

    # v3.5: T5018 aggregation — gated by industry
    CONSTRUCTION_INDUSTRIES = {"Construction/Trades"}
    is_construction = industry in CONSTRUCTION_INDUSTRIES
    t5018_data = aggregate_t5018(filtered) if is_construction else []
    if not is_construction:
        # Still detect high-value e-transfers in other industries for manual review
        possible_subs = [t for t in filtered if detect_subcontractor_payment(t.get("description",""), t.get("debit",0), industry)]
        if possible_subs:
            st.info(f"💡 {len(possible_subs)} large e-transfer/subcontractor-style payments detected. "
                    f"T5018 reporting applies only if >50% of business income is from construction. "
                    f"Flagged for manual review.")
    if t5018_data:
        reportable = [r for r in t5018_data if r["t5018_required"]]
        if reportable:
            st.warning(f"📋 T5018: {len(reportable)} subcontractor(s) with payments ≥ $500 — requires CRA T5018 filing. "
                       f"⚠️ Applies only if construction activities exceed 50% of total business income. Confirm before filing.")

    # Count merchant normalizations
    norm_count = sum(1 for t in filtered if t.get("description_raw"))

    all_categorized = filtered

    detail = f"✅ {len(all_categorized)} transactions in {elapsed:.0f}s | Cost: ${total_cost:.4f}"
    if total_vision_count > 0:
        detail += f" | 📝 {total_text_count} text + 📸 {total_vision_count} vision"
    if ghost_count > 0:
        detail += f" | 🧹 {ghost_count} ghosts"
    if norm_count > 0:
        detail += f" | 🏷 {norm_count} merchants normalized"
    st.success(detail)

    st.session_state.transactions = all_categorized
    st.session_state.flags = []
    st.session_state.summary = {"period": period or period_hint or 'auto', "transactions": str(len(all_categorized))}
    st.session_state.total_cost = total_cost
    st.session_state.recon_matches = recon_matches
    st.session_state.recon_unmatched = recon_unmatched
    st.session_state.recon_duplicates = recon_duplicates
    st.session_state.receipt_matches = receipt_matches
    st.session_state.invoice_data = invoice_data
    st.session_state.validation_results = validation_results
    st.session_state.audit_trail = audit_trail
    # v3.3 new fields
    st.session_state.validation_report = validation_report
    st.session_state.anomalies = anomalies
    st.session_state.t5018_data = t5018_data
    st.session_state.coverage_info = coverage_info
    st.session_state.normalization_count = norm_count

# ═══════════════════════════════════════════════════════════════════
# DISPLAY RESULTS
# ═══════════════════════════════════════════════════════════════════
if st.session_state.transactions:
    txns = st.session_state.transactions
    st.divider()
    st.header("📋 Results")

    real_income = sum(t["credit"] for t in txns if t["credit"] and t.get("type") not in ("PAYMENT","REFUND","FEE_REBATE"))
    payments = sum(t["credit"] for t in txns if t["credit"] and t.get("type") == "PAYMENT")
    refunds = sum(t["credit"] for t in txns if t["credit"] and t.get("type") == "REFUND")
    total_exp = sum(t["debit"] for t in txns if t["debit"])
    total_itc = sum(t.get("itc_amount",0) for t in txns)
    uncategorized = len([t for t in txns if "Uncategorized" in t.get("category","")])
    low_conf = len([t for t in txns if str(t.get("confidence","0")).isdigit() and int(t.get("confidence","0"))<70])
    cca_assets = [t for t in txns if "CCA_ASSET" in t.get("notes","")]

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Transactions", len(txns)); c2.metric("Total Expenses", f"${total_exp:,.2f}")
    c3.metric("Payments to Card", f"${payments:,.2f}"); c4.metric("HST ITC", f"${total_itc:,.2f}")
    c5,c6,c7,c8 = st.columns(4)
    c5.metric("Refunds/Credits", f"${refunds:,.2f}"); c6.metric("Uncategorized", uncategorized)
    c7.metric("CCA Assets (≥$500)", len(cca_assets)); c8.metric("API Cost", f"${st.session_state.total_cost:.4f}")

    if st.session_state.detected_bank:
        st.info(f"🏦 Auto-detected: **{st.session_state.detected_bank}**")

    # ── Extraction Confidence Score ──
    val_results = st.session_state.get("validation_results", [])
    conf_factors = []
    # Factor 1: What % of transactions have confidence >= 70
    high_conf = len([t for t in txns if str(t.get("confidence","0")).isdigit() and int(t.get("confidence","0")) >= 70])
    conf_factors.append(high_conf / max(len(txns),1))
    # Factor 2: Validation match rate
    if val_results:
        all_vals = [v for vr in val_results for v in vr.get("validations",[])]
        matched_vals = [v for v in all_vals if "✅" in v.get("status","")]
        conf_factors.append(len(matched_vals) / max(len(all_vals),1))
    # Factor 3: % categorized (not Uncategorized)
    categorized = len([t for t in txns if "Uncategorized" not in t.get("category","")])
    conf_factors.append(categorized / max(len(txns),1))
    # Factor 4: No duplicates
    dup_count = len([t for t in txns if "POTENTIAL_DUPLICATE" in t.get("notes","")])
    conf_factors.append(1.0 - min(dup_count / max(len(txns),1), 0.2))

    extraction_confidence = round(sum(conf_factors) / len(conf_factors) * 100)
    conf_color = "green" if extraction_confidence >= 85 else "orange" if extraction_confidence >= 70 else "red"
    st.markdown(f"### Extraction Confidence: :{conf_color}[**{extraction_confidence}%**]")

    prov_display = province_override if province_override != "Auto-Detect" else "ON"
    wb = build_excel(txns, [], st.session_state.summary, business_name, industry, prov_display, period,
                     st.session_state.get("recon_matches"), st.session_state.get("recon_unmatched"),
                     st.session_state.get("receipt_matches"), st.session_state.get("invoice_data"),
                     st.session_state.get("validation_results"),
                     st.session_state.get("t5018_data"), st.session_state.get("validation_report"),
                     st.session_state.get("anomalies"), st.session_state.get("audit_trail"))
    safe = "".join(c for c in business_name if c.isalnum() or c in " -_").strip()
    fname = f"{safe}_BookKeepAI_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)

    tab_count = 7 + sum(1 for x in [st.session_state.get("recon_matches"), st.session_state.get("receipt_matches"),
        st.session_state.get("invoice_data"), st.session_state.get("validation_results"),
        st.session_state.get("t5018_data"), st.session_state.get("validation_report")] if x)
    st.download_button(f"⬇️ Download Excel — {len(txns)} transactions, {tab_count} tabs",
        data=buf, file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary", use_container_width=True)

    # v3.7: QuickBooks / Xero CSV export
    with st.expander("📤 Export to Accounting Software (QuickBooks / Xero)"):
        st.caption(
            "Download your transactions in the native import format for QuickBooks Online or Xero. "            "Import via **Banking → Upload Statement** in QuickBooks, or **Accounting → Bank Accounts → Import** in Xero."
        )
        col_qb, col_xero = st.columns(2)
        with col_qb:
            qb_bytes = build_quickbooks_csv(txns, business_name, period)
            qb_fname = f"{safe}_QuickBooks_{datetime.now().strftime('%Y%m%d')}.csv"
            st.download_button(
                label="⬇️ QuickBooks Online CSV",
                data=qb_bytes, file_name=qb_fname, mime="text/csv",
                use_container_width=True,
                help="Import via Banking → Upload Statement → Map Columns in QuickBooks Online"
            )
        with col_xero:
            xero_bytes = build_xero_csv(txns, business_name, period)
            xero_fname = f"{safe}_Xero_{datetime.now().strftime('%Y%m%d')}.csv"
            st.download_button(
                label="⬇️ Xero Bank Statement CSV",
                data=xero_bytes, file_name=xero_fname, mime="text/csv",
                use_container_width=True,
                help="Import via Accounting → Bank Accounts → Manage Account → Import Statement in Xero"
            )

    # v3.3: Additional metrics row
    c9, c10, c11, c12 = st.columns(4)
    norm_count = st.session_state.get("normalization_count", 0)
    val_report = st.session_state.get("validation_report", [])
    t5018 = st.session_state.get("t5018_data", [])
    anomaly_list = st.session_state.get("anomalies", [])
    c9.metric("Merchants Normalized", norm_count)
    c10.metric("Validation Fixes", len(val_report))
    c11.metric("T5018 Subcontractors", len([r for r in t5018 if r.get("t5018_required")]))
    c12.metric("Anomalies Flagged", len(anomaly_list))

    if st.session_state.get("t5018_data"):
        with st.expander("📋 T5018 Subcontractor Summary (Construction/Trades only)"):
            st.caption("⚠️ T5018 reporting is mandatory only when construction activities exceed 50% of total business income. "                       "Verify eligibility with your client before filing. Incorporated payees are exempt.")
            import pandas as pd
            st.dataframe(pd.DataFrame([{"Payee": r["payee"], "Payments": r["count"],
                "Total": f'${r["total"]:,.2f}', "T5018 Required": "🔴 YES" if r["t5018_required"] else "✅ No"}
                for r in st.session_state.t5018_data]), use_container_width=True)

    if st.session_state.get("anomalies"):
        with st.expander("📊 Expense Anomalies"):
            import pandas as pd
            st.dataframe(pd.DataFrame(st.session_state.anomalies), use_container_width=True)

    # v3.5: Audit Trail — split into Financial and Cosmetic changes
    audit = st.session_state.get("audit_trail", [])
    if audit:
        financial_audit = [r for r in audit if r.get("change_type") == "FINANCIAL"]
        cosmetic_audit = [r for r in audit if r.get("change_type") != "FINANCIAL"]
        with st.expander(f"🔍 Post-Processing Audit Trail ({len(financial_audit)} financial, {len(cosmetic_audit)} cosmetic)"):
            import pandas as pd
            st.caption(
                "**Financial Changes** affect category, ITC, or confidence — review before filing. "
                "**Cosmetic Changes** are merchant normalizations only — informational."
            )
            if financial_audit:
                st.markdown("#### 💰 Financial Changes")
                df_fin = pd.DataFrame([{
                    "Date":           r["date"],
                    "Description":    r["description"],
                    "Debit":          f'${r["debit"]:,.2f}' if r["debit"] else "",
                    "AI Category":    r["ai_category"],
                    "Final Category": r["final_category"],
                    "AI ITC":         f'${r["ai_itc"]:,.2f}',
                    "Final ITC":      f'${r["final_itc"]:,.2f}',
                    "AI Confidence":  r["ai_conf"],
                    "Change Reason":  r["changes"],
                } for r in financial_audit])
                st.dataframe(df_fin, use_container_width=True)
            else:
                st.success("No financial changes — all AI categories and ITC amounts accepted.")
            if cosmetic_audit:
                st.markdown("#### 🏷 Cosmetic Changes (merchant normalization)")
                df_cos = pd.DataFrame([{
                    "Date":           r["date"],
                    "Description":    r["description"],
                    "Change":         r["changes"],
                } for r in cosmetic_audit])
                st.dataframe(df_cos, use_container_width=True)
            # CSV download of full audit trail
            df_audit = pd.DataFrame([{
                "Date": r["date"], "Description": r["description"],
                "Debit": r["debit"], "AI Category": r["ai_category"],
                "Final Category": r["final_category"], "AI ITC": r["ai_itc"],
                "Final ITC": r["final_itc"], "AI Confidence": r["ai_conf"],
                "Change Reason": r["changes"], "Change Type": r.get("change_type",""),
            } for r in audit])
            csv_bytes = df_audit.to_csv(index=False).encode("utf-8")
            st.download_button(
                label="⬇️ Download Full Audit Trail (.csv)",
                data=csv_bytes,
                file_name=f"audit_trail_{business_name.replace(' ','_')}.csv",
                mime="text/csv"
            )

    with st.expander("📊 Transaction Preview"):
        import pandas as pd
        st.dataframe(pd.DataFrame(txns), use_container_width=True)
